#!/usr/bin/env python3
"""
GUI Moderna estilo Web para Comparador de Campos Específicos V2
Usando CustomTkinter para estilos CSS modernos
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import json
import os
from datetime import datetime
import sys
from PIL import Image, ImageTk
import threading
import xml.etree.ElementTree as ET
import pandas as pd
import logging

# Configurar tema moderno
ctk.set_appearance_mode("light")  # "light" o "dark"
ctk.set_default_color_theme("blue")  # "blue", "green", "dark-blue"

class ModernWebGUI:
    def __init__(self):
        # Configurar ventana principal
        self.root = ctk.CTk()
        self.root.title("XML Validation Fields - Interfaz Web Moderna")
        self.root.geometry("1400x900")
        
        # MAXIMIZAR VENTANA AUTOMÁTICAMENTE
        self.root.state('zoomed')  # Windows
        
        # Configurar logging
        self.setup_logging()
        
        # Variables de estado
        self.selected_json_path = None
        self.current_step = None
        self.config = {}
        self.db_config = {}
        
        # Cargar configuración
        self.load_config()
        self.load_db_config()  # Cargar configuración de BD
        
        # Variables de los pasos
        self.batt_dept_ids = []  # Lista para almacenar múltiples IDs
        self.batt_dept_sql_filter = None  # Filtro SQL formateado
        self.xref_xpath = None  # XPath del xref_id encontrado
        self.xref_id = None
        self.upload_status = None
        self.verification_result = None
        self.filtered_data = None
        
        # Variables para campos objetivos (Paso 4)
        self.target_fields = []  # Lista de campos: [{"db_field": "nombre", "xpath": "xpath"}, ...]
        
        # Variables para carpeta XML (Paso 5)
        self.xml_folder_path = None
        
        # Variables para xref_ids extraídos de XMLs (Paso 5)
        self.extracted_xref_ids = []  # Lista de xref_ids encontrados en XMLs
        
        # Variables para filtros (Paso 6)
        self.filter_criteria = []  # Lista de filtros: [{"field": "campo", "operator": "=", "value": "valor"}, ...]
        
        # Estilos CSS-like
        self.colors = {
            'primary': '#2563eb',      # Azul moderno
            'secondary': '#64748b',    # Gris azulado
            'accent': '#f59e0b',       # Naranja dorado
            'success': '#10b981',      # Verde éxito
            'warning': '#f59e0b',      # Naranja advertencia
            'danger': '#ef4444',       # Rojo peligro
            'white': '#ffffff',
            'gray_50': '#f8fafc',
            'gray_100': '#f1f5f9',
            'gray_200': '#e2e8f0',
            'gray_300': '#cbd5e1',
            'gray_400': '#94a3b8',
            'gray_500': '#64748b',
            'gray_600': '#475569',
            'gray_700': '#334155',
            'gray_800': '#1e293b',
            'gray_900': '#0f172a'
        }
        
        # Configurar layout principal
        self.setup_layout()
        
        # Mostrar vista inicial
        self.show_welcome_view()
    
    def setup_logging(self):
        """Configurar logging para la aplicación"""
        try:
            # Crear directorio de logs si no existe
            log_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'logs')
            os.makedirs(log_dir, exist_ok=True)
            
            # Configurar logging
            log_filename = f"web_gui_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
            log_path = os.path.join(log_dir, log_filename)
            
            logging.basicConfig(
                level=logging.INFO,
                format='%(asctime)s - %(levelname)s - %(message)s',
                handlers=[
                    logging.FileHandler(log_path, encoding='utf-8'),
                    logging.StreamHandler()
                ]
            )
            
            self.logger = logging.getLogger(__name__)
            self.logger.info("Sistema de logging iniciado")
            
        except Exception as e:
            print(f"Error configurando logging: {e}")
            self.logger = logging.getLogger(__name__)
    
    def load_config(self):
        """Cargar configuración desde config.json"""
        try:
            config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'config.json')
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    self.config = json.load(f)
                    self.db_config = self.config.get('database', {})
        except Exception as e:
            print(f"Error cargando configuración: {e}")
            self.config = {}
            self.db_config = {}
    
    def setup_layout(self):
        """Configurar el layout principal estilo web moderno"""
        # Frame principal con grid - CONFIGURACIÓN MEJORADA
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_columnconfigure(0, weight=0)  # Sidebar izquierdo fijo
        self.root.grid_columnconfigure(1, weight=1)  # Contenido principal
        self.root.grid_columnconfigure(2, weight=0, minsize=750)  # Panel de seguimiento con ancho mínimo
        
        # Header superior
        self.create_top_menu()
        
        # Sidebar izquierdo compacto
        self.create_modern_sidebar()
        
        # Área de contenido principal
        self.create_main_content_area()
        
        # Panel de seguimiento derecho
        self.create_tracking_panel()
    
    def open_bd_config(self):
        """Abrir formulario de configuración de BD PostgreSQL"""
        self.show_bd_config_dialog()
    
    def open_reports(self):
        """Abrir carpeta de reportes"""
        import subprocess
        import platform
        
        reports_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reportes")
        
        # Crear carpeta si no existe
        if not os.path.exists(reports_path):
            os.makedirs(reports_path)
        
        try:
            if platform.system() == "Windows":
                os.startfile(reports_path)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", reports_path])
            else:  # Linux
                subprocess.run(["xdg-open", reports_path])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la carpeta de reportes: {str(e)}")
    
    def open_logs(self):
        """Abrir carpeta de logs"""
        import subprocess
        import platform
        
        logs_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "logs")
        
        # Crear carpeta si no existe
        if not os.path.exists(logs_path):
            os.makedirs(logs_path)
        
        try:
            if platform.system() == "Windows":
                os.startfile(logs_path)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", logs_path])
            else:  # Linux
                subprocess.run(["xdg-open", logs_path])
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir la carpeta de logs: {str(e)}")
    
    def show_client_code(self):
        """Mostrar configuración de Client Code - Funcionalidad futura"""
        messagebox.showinfo("Client Code", "🚧 Funcionalidad en desarrollo\n\nEsta característica será implementada próximamente.")
    
    def show_utc_config(self):
        """Mostrar configuración UTC - Funcionalidad futura"""
        messagebox.showinfo("UTC Config", "🚧 Funcionalidad en desarrollo\n\nConfiguración de zona horaria UTC será implementada próximamente.")
    
    def show_new_integration(self):
        """Mostrar Nueva Integración - Funcionalidad futura"""
        messagebox.showinfo("Nueva Integración", "🚧 Funcionalidad en desarrollo\n\nHerramienta para crear nuevas integraciones será implementada próximamente.")
    
    def show_xml_modifier(self):
        """Mostrar Modificador XML - Funcionalidad futura"""
        messagebox.showinfo("Modificador XML", "🚧 Funcionalidad en desarrollo\n\nEditor de archivos XML será implementado próximamente.")
    
    def show_bd_config_dialog(self):
        """Mostrar diálogo de configuración de BD"""
        # Crear ventana de configuración
        config_window = ctk.CTkToplevel(self.root)
        config_window.title("Configuración Base de Datos PostgreSQL")
        config_window.geometry("550x650")
        config_window.transient(self.root)
        config_window.grab_set()
        
        # Centrar ventana
        config_window.update_idletasks()
        x = (config_window.winfo_screenwidth() // 2) - (550 // 2)
        y = (config_window.winfo_screenheight() // 2) - (650 // 2)
        config_window.geometry(f"550x650+{x}+{y}")
        
        # Frame principal
        main_frame = ctk.CTkFrame(config_window, fg_color="white")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        title_label = ctk.CTkLabel(
            main_frame,
            text="🗄️ Configuración PostgreSQL",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="#1e293b"
        )
        title_label.pack(pady=(20, 30))
        
        # Campos de configuración
        fields_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        fields_frame.pack(fill="x", padx=20)
        
        # Host
        host_label = ctk.CTkLabel(fields_frame, text="Host:", font=ctk.CTkFont(size=14, weight="bold"))
        host_label.pack(anchor="w", pady=(10, 5))
        host_entry = ctk.CTkEntry(fields_frame, width=400, height=35, placeholder_text="localhost")
        host_entry.pack(fill="x", pady=(0, 10))
        host_entry.insert(0, self.db_config.get('host', 'localhost'))
        
        # Puerto
        port_label = ctk.CTkLabel(fields_frame, text="Puerto:", font=ctk.CTkFont(size=14, weight="bold"))
        port_label.pack(anchor="w", pady=(10, 5))
        port_entry = ctk.CTkEntry(fields_frame, width=400, height=35, placeholder_text="5432")
        port_entry.pack(fill="x", pady=(0, 10))
        port_entry.insert(0, str(self.db_config.get('port', 5432)))
        
        # Base de datos
        db_label = ctk.CTkLabel(fields_frame, text="Base de Datos:", font=ctk.CTkFont(size=14, weight="bold"))
        db_label.pack(anchor="w", pady=(10, 5))
        db_entry = ctk.CTkEntry(fields_frame, width=400, height=35, placeholder_text="nombre_bd")
        db_entry.pack(fill="x", pady=(0, 10))
        db_entry.insert(0, self.db_config.get('database', ''))
        
        # Usuario
        user_label = ctk.CTkLabel(fields_frame, text="Usuario:", font=ctk.CTkFont(size=14, weight="bold"))
        user_label.pack(anchor="w", pady=(10, 5))
        user_entry = ctk.CTkEntry(fields_frame, width=400, height=35, placeholder_text="usuario")
        user_entry.pack(fill="x", pady=(0, 10))
        user_entry.insert(0, self.db_config.get('user', ''))
        
        # Contraseña
        password_label = ctk.CTkLabel(fields_frame, text="Contraseña:", font=ctk.CTkFont(size=14, weight="bold"))
        password_label.pack(anchor="w", pady=(10, 5))
        password_entry = ctk.CTkEntry(fields_frame, width=400, height=35, placeholder_text="contraseña", show="*")
        password_entry.pack(fill="x", pady=(0, 20))
        password_entry.insert(0, self.db_config.get('password', ''))
        
        # Botones
        buttons_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        buttons_frame.pack(pady=30, padx=20)
        
        def save_config():
            self.db_config = {
                'host': host_entry.get(),
                'port': int(port_entry.get()) if port_entry.get().isdigit() else 5432,
                'database': db_entry.get(),
                'user': user_entry.get(),
                'password': password_entry.get()
            }
            self.save_db_config()
            messagebox.showinfo("Éxito", "Configuración guardada correctamente")
            config_window.destroy()
        
        def test_connection():
            try:
                import psycopg2
                conn = psycopg2.connect(
                    host=host_entry.get(),
                    port=int(port_entry.get()) if port_entry.get().isdigit() else 5432,
                    database=db_entry.get(),
                    user=user_entry.get(),
                    password=password_entry.get()
                )
                conn.close()
                messagebox.showinfo("Éxito", "Conexión exitosa a la base de datos")
            except Exception as e:
                messagebox.showerror("Error", f"Error de conexión: {str(e)}")
        
        save_btn = ctk.CTkButton(
            buttons_frame,
            text="💾 Guardar",
            width=120,
            height=40,
            fg_color="#2563eb",
            hover_color="#1d4ed8",
            command=save_config
        )
        save_btn.pack(side="left", padx=(0, 10))
        
        test_btn = ctk.CTkButton(
            buttons_frame,
            text="🔍 Probar Conexión",
            width=140,
            height=40,
            fg_color="#10b981",
            hover_color="#059669",
            command=test_connection
        )
        test_btn.pack(side="left", padx=(0, 10))
        
        cancel_btn = ctk.CTkButton(
            buttons_frame,
            text="❌ Cancelar",
            width=120,
            height=40,
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=config_window.destroy
        )
        cancel_btn.pack(side="left")
    
    def save_db_config(self):
        """Guardar configuración de BD en archivo"""
        config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config")
        if not os.path.exists(config_dir):
            os.makedirs(config_dir)
        
        db_config_path = os.path.join(config_dir, "db_config.json")
        
        try:
            with open(db_config_path, 'w', encoding='utf-8') as f:
                json.dump(self.db_config, f, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Error guardando configuración de BD: {e}")
    
    def load_db_config(self):
        """Cargar configuración de BD desde archivo"""
        config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "config")
        db_config_path = os.path.join(config_dir, "db_config.json")
        
        try:
            if os.path.exists(db_config_path):
                with open(db_config_path, 'r', encoding='utf-8') as f:
                    self.db_config = json.load(f)
        except Exception as e:
            print(f"Error cargando configuración de BD: {e}")
            self.db_config = {}
    
    def create_top_menu(self):
        """Crear menú superior estilo web moderna"""
        # Frame superior fijo - MÁS COMPACTO
        self.top_menu = ctk.CTkFrame(self.root, height=60, corner_radius=0, fg_color="#ffffff")
        self.top_menu.grid(row=0, column=0, columnspan=3, sticky="ew")
        self.top_menu.grid_propagate(False)
        
        # Frame izquierdo para título y botones de configuración
        left_frame = ctk.CTkFrame(self.top_menu, fg_color="transparent")
        left_frame.pack(side="left", padx=20, pady=15)
        
        # Logo/Título de la aplicación
        app_title = ctk.CTkLabel(
            left_frame, 
            text="🔬 XML Validation Fields", 
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="#1e293b"
        )
        app_title.pack(side="left", padx=(0, 10))
        
        # Botones de configuración al lado del título
        config_buttons_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
        config_buttons_frame.pack(side="left")
        
        # Botón de BD Config
        bd_config_btn = ctk.CTkButton(
            config_buttons_frame,
            text="🗄️ Config",
            width=90,
            height=30,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color="#3b82f6",
            hover_color="#2563eb",
            command=self.open_bd_config
        )
        bd_config_btn.pack(side="left", padx=(0, 5))
        
        # Client Code
        client_code_btn = ctk.CTkButton(
            config_buttons_frame,
            text="🏢 Client Code",
            width=110,
            height=30,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color="#8b5cf6",
            hover_color="#7c3aed",
            command=self.show_client_code
        )
        client_code_btn.pack(side="left", padx=(0, 5))
        
        # UTC
        utc_btn = ctk.CTkButton(
            config_buttons_frame,
            text="🌍 UTC",
            width=80,
            height=30,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color="#06b6d4",
            hover_color="#0891b2",
            command=self.show_utc_config
        )
        utc_btn.pack(side="left", padx=(0, 5))
        
        # Nueva Integración
        integration_btn = ctk.CTkButton(
            config_buttons_frame,
            text="� Nueva Integración",
            width=140,
            height=30,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color="#10b981",
            hover_color="#059669",
            command=self.show_new_integration
        )
        integration_btn.pack(side="left", padx=(0, 5))
        
        # Modificador XML
        xml_modifier_btn = ctk.CTkButton(
            config_buttons_frame,
            text="📝 Modificador XML",
            width=140,
            height=30,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color="#f59e0b",
            hover_color="#d97706",
            command=self.show_xml_modifier
        )
        xml_modifier_btn.pack(side="left", padx=(0, 5))
        
        # Botón de Reportes
        reports_btn = ctk.CTkButton(
            config_buttons_frame,
            text="📊 Reportes",
            width=100,
            height=30,
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color="#ef4444",
            hover_color="#dc2626",
            command=self.open_reports
        )
        reports_btn.pack(side="left", padx=(0, 5))
        
        # Botón de Logs
        logs_btn = ctk.CTkButton(
            config_buttons_frame,
            text="� Logs",
            width=80,
            height=30,
            font=ctk.CTkFont(size=11, weight="bold"),
            fg_color="#10b981",
            hover_color="#059669",
            command=self.open_logs
        )
        logs_btn.pack(side="left")
        
        # Frame derecho para la barra de progreso
        right_frame = ctk.CTkFrame(self.top_menu, fg_color="transparent")
        right_frame.pack(side="right", padx=20, pady=15)
        
        # Barra de progreso en el top
        progress_frame = ctk.CTkFrame(right_frame, fg_color="transparent")
        progress_frame.pack(side="right")
        
        progress_label = ctk.CTkLabel(
            progress_frame,
            text="Progreso:",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#64748b"
        )
        progress_label.pack(side="left", padx=(0, 5))
        
        self.progress_bar = ctk.CTkProgressBar(
            progress_frame,
            width=200,
            height=15,
            progress_color="#10b981"
        )
        self.progress_bar.pack(side="left", padx=(0, 5))
        
        self.progress_text = ctk.CTkLabel(
            progress_frame,
            text="0/8",
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="#64748b"
        )
        self.progress_text.pack(side="left")
    
    def show_config_dialog(self):
        """Mostrar diálogo de configuración"""
        messagebox.showinfo("Configuración", "Funcionalidad de configuración en desarrollo")
    
    def show_db_config(self):
        """Mostrar configuración de base de datos"""
        messagebox.showinfo("Base de Datos", "Configuración de BD:\n• Host: localhost\n• Puerto: 5432\n• Usuario: postgres")
    
    def show_reports(self):
        """Mostrar reportes disponibles"""
        messagebox.showinfo("Reportes", "Reportes disponibles:\n• Comparación de campos\n• Análisis de errores\n• Estadísticas de procesamiento")
        
    def open_config(self):
        """Abrir configuración general"""
        self.show_config_dialog()
        
    def open_database(self):
        """Abrir configuración de base de datos"""
        self.show_db_config()
        
    def update_progress(self, step_number):
        """Actualizar la barra de progreso en el header"""
        progress = step_number / 7.0  # Cambiar de 8 a 7 pasos
        self.progress_bar.set(progress)
        self.progress_text.configure(text=f"{step_number}/7")
    
    def create_modern_sidebar(self):
        """Crear sidebar moderno con iconos únicamente - DISEÑO OSCURO OPTIMIZADO"""
        # Sidebar con diseño oscuro profesional - ALTURA OPTIMIZADA
        self.sidebar = ctk.CTkFrame(
            self.root, 
            width=75,  # Ancho optimizado
            corner_radius=0, 
            fg_color="#1e293b"  # Color oscuro profesional
        )
        self.sidebar.grid(row=1, column=0, sticky="nsew")
        self.sidebar.grid_propagate(False)
        
        # Frame para los iconos - CENTRADO ABSOLUTO CON GRID
        icons_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        icons_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Configurar el frame para centrado perfecto
        icons_frame.grid_columnconfigure(0, weight=1)  # Centrar horizontalmente
        
        # LISTA DE ICONOS - TODOS IDÉNTICOS EN FORMATO (7 pasos)
        icons_data = [
            "📂",  # step1
            "🗂️",  # step2  
            "🔍",  # step3
            "📋",  # step4
            "📤",  # step5
            "🎯",  # step6 - Definir filtros
            "⚖️"   # step7 - Comparar
        ]
        
        step_ids = ["step1", "step2", "step3", "step4", "step5", "step6", "step7"]
        
        # Crear todos los iconos de forma idéntica usando grid para mejor control
        for i, icon in enumerate(icons_data):
            self.create_uniform_icon_grid(icons_frame, icon, step_ids[i], i)
    
    def create_uniform_icon_grid(self, parent, icon, step_id, row):
        """Crear icono COMPLETAMENTE UNIFORME usando GRID - SIN EXCEPCIONES"""
        # Forzar todos los iconos a tener el mismo ancho visual
        centered_icon = f" {icon} "  # Espacios antes y después para centrado perfecto
        
        btn = ctk.CTkButton(
            parent,
            text=centered_icon,
            width=55,
            height=50,
            corner_radius=10,
            font=ctk.CTkFont(size=18),  # Reducir tamaño para compensar espacios
            fg_color="#334155",
            hover_color="#475569",
            text_color="#ffffff",
            border_width=0,
            command=lambda: self.handle_step_click(step_id)
        )
        btn.grid(row=row, column=0, pady=2, padx=0, sticky="")  # Grid centrado
    
    def handle_step_click(self, step_id):
        """Manejar clic en paso del workflow"""
        self.current_step = step_id
        
        # Ejecutar paso correspondiente
        if step_id == "step1":
            self.show_step1_view()
        elif step_id == "step2":
            self.show_step2_view()
        elif step_id == "step3":
            self.show_step3_view()
        elif step_id == "step4":
            self.show_step4_view()
        elif step_id == "step5":
            self.show_step5_view()
        elif step_id == "step6":
            self.show_step6_view()
        elif step_id == "step7":
            self.show_step7_view()
        else:
            self.show_step1_view()  # Default
    
    def create_main_content_area(self):
        """Crear área principal de contenido"""
        # Frame principal para contenido
        self.main_frame = ctk.CTkFrame(self.root, corner_radius=0, fg_color="#f8fafc")
        self.main_frame.grid(row=1, column=1, sticky="nsew", padx=0, pady=0)
        
        # Configurar grid para centrado vertical
        self.main_frame.grid_rowconfigure(0, weight=1)  # Espacio superior
        self.main_frame.grid_rowconfigure(1, weight=0)  # Contenido 
        self.main_frame.grid_rowconfigure(2, weight=1)  # Espacio inferior
        self.main_frame.grid_columnconfigure(0, weight=1)
    
    def create_tracking_panel(self):
        """Crear panel de seguimiento de cambios en la derecha"""
        # Frame principal del panel de seguimiento - ANCHO FIJO GARANTIZADO
        self.tracking_frame = ctk.CTkFrame(self.root, corner_radius=0, fg_color="#ffffff", width=750)
        self.tracking_frame.grid(row=1, column=2, sticky="nsew", padx=(0, 0), pady=0)
        self.tracking_frame.grid_propagate(False)
        
        # Header del panel
        header_frame = ctk.CTkFrame(self.tracking_frame, corner_radius=0, fg_color="#2563eb", height=60)
        header_frame.pack(fill="x", padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        title_label = ctk.CTkLabel(
            header_frame,
            text="📊 Seguimiento de Progreso",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="white"
        )
        title_label.pack(expand=True, pady=15)
        
        # Área de contenido del seguimiento con scroll
        self.tracking_content = ctk.CTkScrollableFrame(
            self.tracking_frame,
            fg_color="#f8fafc",
            corner_radius=0
        )
        self.tracking_content.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Inicializar el contenido del seguimiento
        self.update_tracking_display()
    
    def show_home_view(self):
        """Mostrar la vista principal/inicio"""
        self.clear_main_content()
        self.create_content_header("🏠 Comparador de Campos XML vs BD", "Sistema de validación de campos específicos")
        
        # Contenido principal de inicio
        content_frame = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        content_frame.grid_columnconfigure(0, weight=1)
        
        welcome_label = ctk.CTkLabel(
            content_frame,
            text="Bienvenido al Sistema de Comparación de Campos\n\n📋 Proceso de 7 pasos para validar campos XML vs Base de Datos\n\n👈 Usa el menú lateral para navegar entre los pasos",
            font=ctk.CTkFont(size=16),
            text_color="#1f2937",
            justify="center"
        )
        welcome_label.pack(expand=True, fill="both", padx=40, pady=40)
    
    def update_tracking_display(self):
        """Actualizar la visualización del seguimiento"""
        # Limpiar contenido anterior
        for widget in self.tracking_content.winfo_children():
            widget.destroy()
        
        # Título de resumen
        summary_label = ctk.CTkLabel(
            self.tracking_content,
            text="Estado Actual del Proceso",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#1f2937"
        )
        summary_label.pack(pady=(10, 15), padx=10)
        
        # Paso 1: Selección de JSON
        self.create_tracking_step("1", "📁 Selección JSON", 
                                 self.selected_json_path is not None,
                                 os.path.basename(self.selected_json_path) if self.selected_json_path else "No seleccionado")
        
        # Paso 2: batt_dept_ids
        has_batt_ids = hasattr(self, 'batt_dept_ids') and self.batt_dept_ids
        batt_summary = f"{len(self.batt_dept_ids)} IDs encontrados" if has_batt_ids else "Pendiente"
        self.create_tracking_step("2", "🗂️ batt_dept_id", has_batt_ids, batt_summary)
        
        if has_batt_ids:
            # Mostrar filtro SQL
            sql_frame = ctk.CTkFrame(self.tracking_content, fg_color="#e0f2fe", corner_radius=5)
            sql_frame.pack(fill="x", padx=15, pady=(0, 10))
            
            sql_label = ctk.CTkLabel(
                sql_frame,
                text=f"IN ({', '.join(self.batt_dept_ids)})" if hasattr(self, 'batt_dept_ids') and self.batt_dept_ids else "IN: Generando...",
                font=ctk.CTkFont(size=10, family="Courier"),
                text_color="#0277bd",
                wraplength=300
            )
            sql_label.pack(pady=8, padx=8)
        
        # Paso 3: XPath del xref_id
        has_xpath = hasattr(self, 'xref_xpath') and self.xref_xpath
        xpath_summary = "XPath encontrado" if has_xpath else "Pendiente"
        self.create_tracking_step("3", "🔍 XPath xref_id", has_xpath, xpath_summary)
        
        if has_xpath:
            # Mostrar XPath
            xpath_frame = ctk.CTkFrame(self.tracking_content, fg_color="#f3e5f5", corner_radius=5)
            xpath_frame.pack(fill="x", padx=15, pady=(0, 10))
            
            xpath_label = ctk.CTkLabel(
                xpath_frame,
                text=f"XPath: {self.xref_xpath}",
                font=ctk.CTkFont(size=10, family="Courier"),
                text_color="#7b1fa2",
                wraplength=300
            )
            xpath_label.pack(pady=8, padx=8)
        
        # Paso 4: Campos objetivos definidos
        has_target_fields = hasattr(self, 'target_fields') and len(self.target_fields) > 0
        fields_summary = f"{len(self.target_fields)} campos configurados" if has_target_fields else "Pendiente"
        self.create_tracking_step("4", "🎯 Campos objetivos", has_target_fields, fields_summary)
        
        if has_target_fields:
            # Mostrar resumen de campos con scroll horizontal
            fields_frame = ctk.CTkFrame(self.tracking_content, fg_color="#f0f9ff", corner_radius=5)
            fields_frame.pack(fill="x", padx=15, pady=(0, 10))
            
            for i, field in enumerate(self.target_fields[:3]):  # Mostrar máximo 3 campos
                # Frame individual para cada campo
                single_field_frame = ctk.CTkFrame(fields_frame, fg_color="transparent")
                single_field_frame.pack(fill="x", padx=5, pady=2)
                
                # Información básica del campo
                field_info = ctk.CTkLabel(
                    single_field_frame,
                    text=f"{field['table']}.{field['db_field']}",
                    font=ctk.CTkFont(size=9, weight="bold"),
                    text_color="#0369a1"
                )
                field_info.pack(anchor="w")
                
                # XPath en texto scrollable pequeño
                xpath_text = ctk.CTkTextbox(
                    single_field_frame,
                    height=20,
                    wrap="none",
                    font=ctk.CTkFont(size=8, family="Courier"),
                    text_color="#7c2d12",
                    fg_color="#f0f9ff"
                )
                xpath_text.pack(fill="x", pady=(1, 0))
                xpath_text.insert("1.0", field['xpath'])
                xpath_text.configure(state="disabled")
            
            if len(self.target_fields) > 3:
                more_label = ctk.CTkLabel(
                    fields_frame,
                    text=f"... y {len(self.target_fields) - 3} más",
                    font=ctk.CTkFont(size=9),
                    text_color="#6b7280"
                )
                more_label.pack(anchor="w", padx=8, pady=(0, 8))
        
        # Paso 5: Carpeta XML seleccionada
        has_xml_folder = hasattr(self, 'xml_folder_path') and self.xml_folder_path
        xml_summary = "Carpeta seleccionada" if has_xml_folder else "Pendiente"
        self.create_tracking_step("5", "📁 Carpeta XMLs", has_xml_folder, xml_summary)
        
        if has_xml_folder:
            # Mostrar información de la carpeta
            xml_frame = ctk.CTkFrame(self.tracking_content, fg_color="#ecfdf5", corner_radius=5)
            xml_frame.pack(fill="x", padx=15, pady=(0, 10))
            
            folder_name = os.path.basename(self.xml_folder_path) if self.xml_folder_path else ""
            folder_label = ctk.CTkLabel(
                xml_frame,
                text=f"� {folder_name}",
                font=ctk.CTkFont(size=10, family="Courier"),
                text_color="#047857"
            )
            folder_label.pack(pady=8, padx=8)
            
            # Contar archivos XML
            try:
                xml_files = [f for f in os.listdir(self.xml_folder_path) if f.lower().endswith('.xml')]
                count_label = ctk.CTkLabel(
                    xml_frame,
                    text=f"{len(xml_files)} archivos XML",
                    font=ctk.CTkFont(size=9),
                    text_color="#059669"
                )
                count_label.pack(pady=(0, 4), padx=8)
                
                # Mostrar información de xref_ids encontrados
                if hasattr(self, 'extracted_xref_ids') and self.extracted_xref_ids:
                    xref_count_label = ctk.CTkLabel(
                        xml_frame,
                        text=f"🔍 {len(self.extracted_xref_ids)} XRef IDs únicos extraídos",
                        font=ctk.CTkFont(size=9, weight="bold"),
                        text_color="#1d4ed8"
                    )
                    xref_count_label.pack(pady=(0, 4), padx=8)
                    
                    # Mostrar algunos valores de ejemplo
                    sample_xrefs = [item['value'] for item in self.extracted_xref_ids[:3]]
                    if sample_xrefs:
                        sample_text = ", ".join(sample_xrefs)
                        if len(self.extracted_xref_ids) > 3:
                            sample_text += f"... (+{len(self.extracted_xref_ids) - 3})"
                        
                        sample_label = ctk.CTkLabel(
                            xml_frame,
                            text=f"Ejemplos: {sample_text}",
                            font=ctk.CTkFont(size=8),
                            text_color="#6366f1",
                            wraplength=280
                        )
                        sample_label.pack(pady=(0, 8), padx=8)
                else:
                    # Indicar que aún no se han extraído XRef IDs
                    no_xref_label = ctk.CTkLabel(
                        xml_frame,
                        text="⚠️ XRef IDs no extraídos aún",
                        font=ctk.CTkFont(size=9),
                        text_color="#d97706"
                    )
                    no_xref_label.pack(pady=(0, 8), padx=8)
            except Exception:
                pass
        
        # Paso 6: Filtros definidos
        has_filters = hasattr(self, 'filter_criteria') and len(self.filter_criteria) > 0
        filter_summary = f"{len(self.filter_criteria)} filtros definidos" if has_filters else "Pendiente"
        self.create_tracking_step("6", "🎯 Filtros BD", has_filters, filter_summary)
        
        if has_filters:
            # Mostrar resumen de filtros
            filters_frame = ctk.CTkFrame(self.tracking_content, fg_color="#fef3c7", corner_radius=5)
            filters_frame.pack(fill="x", padx=15, pady=(0, 10))
            
            # Mostrar filtros
            for i, filter_item in enumerate(self.filter_criteria[:3]):  # Mostrar máximo 3
                filter_info = ctk.CTkLabel(
                    filters_frame,
                    text=f"🔍 {filter_item['table']}.{filter_item['field']} = {filter_item['value']}",
                    font=ctk.CTkFont(size=10),
                    text_color="#92400e"
                )
                filter_info.pack(anchor="w", padx=8, pady=2)
            
            if len(self.filter_criteria) > 3:
                more_filters_label = ctk.CTkLabel(
                    filters_frame,
                    text=f"... y {len(self.filter_criteria) - 3} más",
                    font=ctk.CTkFont(size=9),
                    text_color="#6b7280"
                )
                more_filters_label.pack(anchor="w", padx=8, pady=(0, 8))
        
        # Paso 7: Comparación
        comparison_completed = False  # Esta variable se actualizará cuando se implemente
        comparison_summary = "Pendiente"
        self.create_tracking_step("7", "⚖️ Comparar", comparison_completed, comparison_summary)
    
    def create_tracking_step(self, step_num, step_name, completed, summary):
        """Crear un elemento de seguimiento para un paso"""
        # Frame del paso
        step_frame = ctk.CTkFrame(
            self.tracking_content,
            fg_color="#ffffff" if completed else "#f9fafb",
            corner_radius=8,
            border_width=2,
            border_color="#10b981" if completed else "#e5e7eb"
        )
        step_frame.pack(fill="x", padx=10, pady=5)
        
        # Header del paso
        header_frame = ctk.CTkFrame(step_frame, fg_color="transparent")
        header_frame.pack(fill="x", padx=10, pady=(8, 4))
        
        # Número del paso e icono
        status_icon = "✅" if completed else "⭕"
        step_label = ctk.CTkLabel(
            header_frame,
            text=f"{status_icon} Paso {step_num}",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#10b981" if completed else "#6b7280"
        )
        step_label.pack(side="left")
        
        # Nombre del paso
        name_label = ctk.CTkLabel(
            step_frame,
            text=step_name,
            font=ctk.CTkFont(size=11),
            text_color="#374151"
        )
        name_label.pack(anchor="w", padx=10, pady=(0, 4))
        
        # Resumen/Estado
        summary_label = ctk.CTkLabel(
            step_frame,
            text=summary,
            font=ctk.CTkFont(size=10),
            text_color="#6b7280",
            wraplength=280
        )
        summary_label.pack(anchor="w", padx=10, pady=(0, 8))
    
    def clear_main_content(self):
        """Limpiar contenido principal"""
        for widget in self.main_frame.winfo_children():
            widget.destroy()
            
        # Reconfigurar grid para centrado vertical
        self.main_frame.grid_rowconfigure(0, weight=1)  # Espacio superior
        self.main_frame.grid_rowconfigure(1, weight=0)  # Contenido 
        self.main_frame.grid_rowconfigure(2, weight=1)  # Espacio inferior
        self.main_frame.grid_columnconfigure(0, weight=1)
    
    def create_content_header(self, title, subtitle=""):
        """Crear header para contenido con título y subtítulo"""
        # Header del contenido - COMPACTO
        header = ctk.CTkFrame(self.main_frame, height=40, fg_color="white", corner_radius=6)
        header.grid(row=0, column=0, sticky="ew", padx=20, pady=(20, 10))
        header.grid_propagate(False)
        
        # Título principal
        self.title_label = ctk.CTkLabel(
            header,
            text=title,
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="#1e293b"
        )
        self.title_label.pack(side="left", padx=20, pady=10)
        
        # Subtítulo si se proporciona
        if subtitle:
            subtitle_label = ctk.CTkLabel(
                header,
                text=subtitle,
                font=ctk.CTkFont(size=12),
                text_color="#64748b"
            )
            subtitle_label.pack(side="left", padx=(0, 20), pady=10)
    
    def show_welcome_view(self):
        """Mostrar vista de bienvenida CENTRADA VERTICALMENTE"""
        self.clear_main_content()
        
        # Container principal centrado vertical y horizontalmente
        welcome_container = ctk.CTkFrame(
            self.main_frame, 
            fg_color="white", 
            corner_radius=12
        )
        welcome_container.grid(row=1, column=0, padx=50, pady=0, sticky="")
        
        # Contenido de bienvenida
        welcome_content = ctk.CTkFrame(welcome_container, fg_color="transparent")
        welcome_content.pack(padx=60, pady=60)
        
        # Título de bienvenida
        welcome_title = ctk.CTkLabel(
            welcome_content,
            text="🔬 XML Validation Fields",
            font=ctk.CTkFont(size=32, weight="bold"),
            text_color="#1e293b"
        )
        welcome_title.pack(pady=(0, 15))
        
        # Subtítulo
        subtitle = ctk.CTkLabel(
            welcome_content,
            text="Herramienta moderna para validación y análisis de campos XML",
            font=ctk.CTkFont(size=16),
            text_color="#64748b"
        )
        subtitle.pack(pady=(0, 30))
        
        # Iconos de características principales
        features_frame = ctk.CTkFrame(welcome_content, fg_color="transparent")
        features_frame.pack(pady=(0, 30))
        
        features = [
            ("📊", "Análisis\nAvanzado"),
            ("🔄", "Proceso\nAutomatizado"),
            ("💾", "Base de\nDatos"),
            ("📈", "Reportes\nDetallados")
        ]
        
        for i, (icon, text) in enumerate(features):
            feature_frame = ctk.CTkFrame(features_frame, fg_color="transparent")
            feature_frame.grid(row=0, column=i, padx=20, pady=10)
            
            icon_label = ctk.CTkLabel(
                feature_frame,
                text=icon,
                font=ctk.CTkFont(size=24)
            )
            icon_label.pack()
            
            text_label = ctk.CTkLabel(
                feature_frame,
                text=text,
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color="#475569"
            )
            text_label.pack(pady=(5, 0))
        
        # Botón de inicio
        start_btn = ctk.CTkButton(
            welcome_content,
            text="📂 Comenzar - Seleccionar JSON",
            width=250,
            height=50,
            font=ctk.CTkFont(size=14, weight="bold"),
            corner_radius=8,
            command=lambda: self.handle_step_click("step1")
        )
        start_btn.pack(pady=20)
        
        # Información de versión
        version_label = ctk.CTkLabel(
            welcome_content,
            text="v2.0 - Interfaz Web Moderna",
            font=ctk.CTkFont(size=10),
            text_color="#94a3b8"
        )
        version_label.pack(pady=(20, 0))

    # ==== VISTAS DE PASOS ====
    
    def create_navigation_frame(self, parent, current_step):
        """Crear frame de navegación estándar"""
        nav_frame = ctk.CTkFrame(parent, fg_color="transparent")
        nav_frame.grid(row=10, column=0, columnspan=3, sticky="ew", pady=(20, 10), padx=20)
        nav_frame.grid_columnconfigure(1, weight=1)
        
        # Botón Anterior
        if current_step > 1:
            prev_btn = ctk.CTkButton(
                nav_frame,
                text=f"⬅️ Paso {current_step-1}",
                width=120,
                height=40,
                font=ctk.CTkFont(size=12),
                command=lambda: self.handle_step_click(f"step{current_step-1}")
            )
            prev_btn.grid(row=0, column=0, padx=(0, 10), sticky="w")
        
        # Botón Home en el centro
        home_btn = ctk.CTkButton(
            nav_frame,
            text="🏠 Inicio",
            width=100,
            height=40,
            font=ctk.CTkFont(size=12),
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=self.show_home_view
        )
        home_btn.grid(row=0, column=1, sticky="")
        
        # Botón Siguiente
        if current_step < 7:
            next_btn = ctk.CTkButton(
                nav_frame,
                text=f"Paso {current_step+1} ➡️",
                width=120,
                height=40,
                font=ctk.CTkFont(size=12, weight="bold"),
                fg_color="#059669",
                hover_color="#047857",
                command=lambda: self.handle_step_click(f"step{current_step+1}")
            )
            next_btn.grid(row=0, column=2, padx=(10, 0), sticky="e")
        
        return nav_frame

    def show_step1_view(self):
        """Vista del Paso 1: Seleccionar JSON"""
        self.clear_main_content()
        self.create_content_header("📂 Paso 1: Seleccionar Archivo JSON", "Selecciona el archivo JSON para procesar")
        
        # Frame principal con mejor gestión de altura
        content_frame = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        content_frame.grid_columnconfigure(0, weight=1)
        content_frame.grid_rowconfigure(0, weight=1)
        
        # Frame contenido (que se expande)
        main_content = ctk.CTkFrame(content_frame, fg_color="transparent")
        main_content.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        main_content.grid_columnconfigure(0, weight=1)
        
        # Instrucciones
        instruction_label = ctk.CTkLabel(
            main_content,
            text="Selecciona el archivo JSON que contiene los datos a procesar",
            font=ctk.CTkFont(size=14),
            text_color="#64748b"
        )
        instruction_label.grid(row=0, column=0, pady=20)
        
        # Botón de selección
        select_btn = ctk.CTkButton(
            main_content,
            text="📁 Seleccionar Archivo JSON",
            width=200,
            height=50,
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self.select_json_file
        )
        select_btn.grid(row=1, column=0, pady=20)
        
        # Mostrar archivo seleccionado
        if self.selected_json_path:
            file_label = ctk.CTkLabel(
                main_content,
                text=f"Archivo seleccionado: {os.path.basename(self.selected_json_path) if self.selected_json_path else 'No seleccionado'}",
                font=ctk.CTkFont(size=12),
                text_color="#10b981"
            )
            file_label.grid(row=2, column=0, pady=10)
        
        # Botones de navegación (fijos al final)
        self.create_navigation_frame(content_frame, 1)
    
    def select_json_file(self):
        """Seleccionar archivo JSON"""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo JSON",
            filetypes=[("Archivos JSON", "*.json"), ("Todos los archivos", "*.*")]
        )
        
        if file_path:
            self.selected_json_path = file_path
            self.update_progress(1)  # Actualizar barra de progreso
            self.show_step1_view()  # Refrescar vista
            self.update_tracking_display()  # Actualizar panel de seguimiento
            messagebox.showinfo("Éxito", f"Archivo seleccionado: {os.path.basename(file_path)}")
    
    def show_step2_view(self):
        """Vista del Paso 2: Buscar batt_dept_id"""
        self.clear_main_content()
        self.create_content_header("🗂️ Paso 2: Buscar batt_dept_id", "Buscar ID del Departamento")
        
        content_frame = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=5)
        
        # Información del archivo seleccionado
        info_label = ctk.CTkLabel(
            content_frame,
            text=f"📁 Archivo: {os.path.basename(self.selected_json_path) if self.selected_json_path else 'No seleccionado'}",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        info_label.pack(pady=(15, 5))
        
        # Botón ejecutar (solo mostrar si no hay resultados previos)
        if not hasattr(self, 'batt_dept_ids') or not self.batt_dept_ids:
            execute_btn = ctk.CTkButton(
                content_frame,
                text="🔍 Buscar batt_dept_id",
                width=200,
                height=50,
                font=ctk.CTkFont(size=14, weight="bold"),
                command=self.execute_step2
            )
            execute_btn.pack(pady=10)
        
        # Frame para mostrar resultados
        self.results_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        self.results_frame.pack(fill="x", padx=15, pady=5)
        
        # Si ya hay resultados, mostrarlos automáticamente
        if hasattr(self, 'batt_dept_ids') and self.batt_dept_ids:
            self.display_step2_results()
    
    def execute_step2(self):
        """Ejecutar paso 2"""
        try:
            # Validar que hay un archivo JSON seleccionado
            if not self.selected_json_path or not os.path.exists(self.selected_json_path):
                # Limpiar resultados anteriores
                for widget in self.results_frame.winfo_children():
                    widget.destroy()
                error_label = ctk.CTkLabel(
                    self.results_frame,
                    text="❌ No hay archivo JSON seleccionado o no existe",
                    font=ctk.CTkFont(size=14),
                    text_color="#dc3545"
                )
                error_label.pack(pady=15)
                return
            
            # Limpiar resultados anteriores
            for widget in self.results_frame.winfo_children():
                widget.destroy()
            
            # Leer el archivo JSON y buscar batt_dept_ids
            import json
            with open(self.selected_json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Buscar batt_dept_ids en la estructura del JSON
            batt_dept_ids = []
            
            def search_batt_dept_ids(obj, path=""):
                if isinstance(obj, dict):
                    # Buscar específicamente la sección batt_depts
                    if 'batt_depts' in obj:
                        batt_depts_value = obj['batt_depts']
                        if isinstance(batt_depts_value, list):
                            # Si es una lista, agregar todos los valores
                            for item in batt_depts_value:
                                if isinstance(item, (str, int)):
                                    batt_dept_ids.append(str(item))
                        elif isinstance(batt_depts_value, (str, int)):
                            # Si es un valor único
                            batt_dept_ids.append(str(batt_depts_value))
                    
                    # Continuar búsqueda recursiva solo si no hemos encontrado batt_depts
                    if not batt_dept_ids:
                        for key, value in obj.items():
                            if isinstance(value, (dict, list)):
                                search_batt_dept_ids(value, f"{path}.{key}" if path else key)
                            
                elif isinstance(obj, list):
                    for i, item in enumerate(obj):
                        search_batt_dept_ids(item, f"{path}[{i}]" if path else f"[{i}]")
            
            search_batt_dept_ids(data)
            
            # Eliminar duplicados manteniendo orden
            unique_ids = list(dict.fromkeys(batt_dept_ids))
            
            # Mostrar resultados en la interfaz
            if unique_ids:
                # Guardar los IDs para pasos posteriores
                self.batt_dept_ids = unique_ids
                
                # Crear filtro SQL formateado para usar en consultas posteriores
                formatted_ids = ", ".join(unique_ids)
                self.batt_dept_sql_filter = f"batt_dept_id IN ({formatted_ids})"
                
                # Mostrar resultados usando función separada
                self.display_step2_results()
                
                # Actualizar progreso y panel de seguimiento
                self.update_progress(2)
                self.update_tracking_display()
                
            else:
                # No se encontraron IDs
                no_results_label = ctk.CTkLabel(
                    self.results_frame,
                    text="❌ No se encontraron batt_dept_id en el archivo JSON",
                    font=ctk.CTkFont(size=14),
                    text_color="#dc3545"
                )
                no_results_label.pack(pady=15)
                
        except Exception as e:
            # Limpiar resultados en caso de error
            for widget in self.results_frame.winfo_children():
                widget.destroy()
                
            error_label = ctk.CTkLabel(
                self.results_frame,
                text=f"❌ Error al buscar batt_dept_id: {str(e)}",
                font=ctk.CTkFont(size=14),
                text_color="#dc3545"
            )
            error_label.pack(pady=15)
    
    def display_step2_results(self):
        """Mostrar los resultados del paso 2"""
        if not hasattr(self, 'batt_dept_ids') or not self.batt_dept_ids:
            return
        
        # Limpiar resultados anteriores
        for widget in self.results_frame.winfo_children():
            widget.destroy()
        
        unique_ids = self.batt_dept_ids
        
        # Título de resultados
        title_label = ctk.CTkLabel(
            self.results_frame,
            text="🎯 batt_dept_id encontrados:",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#2c5aa0"
        )
        title_label.pack(pady=(5, 10))
        
        # Formatear como consulta SQL IN con el nombre del campo
        formatted_ids = ", ".join(unique_ids)
        sql_format = f"batt_dept_id IN ({formatted_ids})"
        
        # Frame para el resultado SQL
        sql_frame = ctk.CTkFrame(
            self.results_frame,
            fg_color="#f8f9fa",
            corner_radius=8,
            height=80
        )
        sql_frame.pack(fill="x", padx=5, pady=5)
        sql_frame.pack_propagate(False)
        
        # Label con el formato SQL
        sql_label = ctk.CTkLabel(
            sql_frame,
            text=sql_format,
            font=ctk.CTkFont(size=14, family="Courier"),
            text_color="#2c5aa0",
            wraplength=700
        )
        sql_label.pack(expand=True, pady=10, padx=15)
        
        # Resumen
        summary_label = ctk.CTkLabel(
            self.results_frame,
            text=f"✅ Total encontrados: {len(unique_ids)} batt_dept_id(s)",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#28a745"
        )
        summary_label.pack(pady=10)
        
        # Frame para mostrar valores IN
        in_frame = ctk.CTkFrame(
            self.results_frame,
            fg_color="#e0f2fe",
            corner_radius=8,
            height=60
        )
        in_frame.pack(fill="x", padx=5, pady=5)
        in_frame.pack_propagate(False)
        
        # Label con los valores IN
        in_values = f"IN ({formatted_ids})"
        in_label = ctk.CTkLabel(
            in_frame,
            text=in_values,
            font=ctk.CTkFont(size=14, family="Courier", weight="bold"),
            text_color="#0277bd"
        )
        in_label.pack(expand=True, pady=10, padx=15)
        
        # Frame para botones de acción
        action_frame = ctk.CTkFrame(self.results_frame, fg_color="transparent")
        action_frame.pack(pady=15)
        
        # Botón para seleccionar otro JSON
        new_search_btn = ctk.CTkButton(
            action_frame,
            text="📁 Seleccionar otro JSON",
            width=180,
            height=40,
            font=ctk.CTkFont(size=12),
            command=self.reset_and_go_step1
        )
        new_search_btn.pack(side="left", padx=5)
        
        # Botón para ir al siguiente paso
        next_step_btn = ctk.CTkButton(
            action_frame,
            text="➡️ Paso 3: Buscar XPath",
            width=180,
            height=40,
            font=ctk.CTkFont(size=12),
            command=lambda: self.handle_step_click("step3")
        )
        next_step_btn.pack(side="left", padx=5)
    
    def reset_and_go_step1(self):
        """Resetear los resultados del paso 2 y ir al paso 1"""
        if hasattr(self, 'batt_dept_ids'):
            delattr(self, 'batt_dept_ids')
        self.handle_step_click("step1")
    
    def show_step3_view(self):
        """Vista del Paso 3: Buscar xpath del xref_id"""
        self.clear_main_content()
        self.create_content_header("🔍 Paso 3: Buscar XPath del xref_id", "Buscar ruta XPath del xref_id/dispatch_number")
        
        content_frame = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=5)
        
        # Información del archivo y paso anterior
        archivo_nombre = os.path.basename(self.selected_json_path) if hasattr(self, 'selected_json_path') and self.selected_json_path else "No seleccionado"
        batt_dept_count = len(self.batt_dept_ids) if hasattr(self, 'batt_dept_ids') and self.batt_dept_ids else 0
        info_label = ctk.CTkLabel(
            content_frame,
            text=f"📁 Archivo: {archivo_nombre}\n✅ batt_dept_ids: {batt_dept_count} encontrados",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        info_label.pack(pady=(15, 5))
        
        # Frame para mostrar resultados
        self.results_frame_step3 = ctk.CTkFrame(content_frame, fg_color="transparent")
        self.results_frame_step3.pack(fill="x", padx=15, pady=5)
        
        # Ejecución automática sin botón visible
        if not hasattr(self, 'xref_xpath') or not self.xref_xpath:
            execute_btn = ctk.CTkButton(
                content_frame,
                text="� Buscar XPath del xref_id",
                width=200,
                height=50,
                font=ctk.CTkFont(size=14, weight="bold"),
                command=self.execute_step3
            )
            # Botón eliminado - ejecutar automáticamente
            # Ejecutar automáticamente después de crear la interfaz
            self.root.after(100, self.execute_step3)
        
        # Si ya hay resultados, mostrarlos automáticamente
        if hasattr(self, 'xref_xpath') and self.xref_xpath:
            self.display_step3_results()

    def execute_step3(self):
        """Ejecutar paso 3: Buscar xpath del xref_id en el JSON"""
        try:
            # Validar que hay un archivo JSON seleccionado
            if not self.selected_json_path or not os.path.exists(self.selected_json_path):
                self.show_step3_error("❌ No hay archivo JSON seleccionado o no existe")
                return
            
            # Limpiar resultados anteriores
            for widget in self.results_frame_step3.winfo_children():
                widget.destroy()
            
            # Leer el archivo JSON y buscar xpath del xref_id
            import json
            with open(self.selected_json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Buscar xpath del xref_id en la estructura del JSON
            xref_xpath = None
            
            def search_xref_xpath(obj, path=""):
                nonlocal xref_xpath
                
                if isinstance(obj, dict):
                    # Buscar específicamente en la sección xpaths
                    if 'xpaths' in obj:
                        xpaths_obj = obj['xpaths']
                        if isinstance(xpaths_obj, dict):
                            # Buscar xref_id dentro de xpaths
                            if 'xref_id' in xpaths_obj:
                                xref_xpath = xpaths_obj['xref_id']
                                return True
                            # También buscar dispatch_number como alternativa
                            if 'dispatch_number' in xpaths_obj:
                                xref_xpath = xpaths_obj['dispatch_number']
                                return True
                    
                    # Si no encontramos en xpaths, buscar recursivamente
                    if not xref_xpath:
                        for key, value in obj.items():
                            if key in ['xref_id', 'dispatch_number'] and isinstance(value, str):
                                xref_xpath = value
                                return True
                            elif isinstance(value, (dict, list)):
                                if search_xref_xpath(value, f"{path}.{key}" if path else key):
                                    return True
                            
                elif isinstance(obj, list):
                    for i, item in enumerate(obj):
                        if search_xref_xpath(item, f"{path}[{i}]" if path else f"[{i}]"):
                            return True
                
                return False
            
            search_xref_xpath(data)
            
            # Mostrar resultados en la interfaz
            if xref_xpath:
                # Guardar el xpath para pasos posteriores
                self.xref_xpath = xref_xpath
                
                # Mostrar resultados usando función separada
                self.display_step3_results()
                
                # Actualizar progreso y panel de seguimiento
                self.update_progress(3)
                self.update_tracking_display()
                
            else:
                # No se encontró xpath
                self.show_step3_error("❌ No se encontró xpath del xref_id/dispatch_number en el archivo JSON")
                
        except Exception as e:
            self.show_step3_error(f"❌ Error al buscar xpath del xref_id: {str(e)}")
    
    def show_step3_error(self, message):
        """Mostrar error en el paso 3"""
        # Limpiar resultados en caso de error
        for widget in self.results_frame_step3.winfo_children():
            widget.destroy()
            
        error_label = ctk.CTkLabel(
            self.results_frame_step3,
            text=message,
            font=ctk.CTkFont(size=14),
            text_color="#dc3545"
        )
        error_label.pack(pady=15)
    
    def display_step3_results(self):
        """Mostrar los resultados del paso 3"""
        if not hasattr(self, 'xref_xpath') or not self.xref_xpath:
            return
        
        # Limpiar resultados anteriores
        for widget in self.results_frame_step3.winfo_children():
            widget.destroy()
        
        # Título de resultados
        title_label = ctk.CTkLabel(
            self.results_frame_step3,
            text="🎯 XPath del xref_id encontrado:",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#2c5aa0"
        )
        title_label.pack(pady=(5, 10))
        
        # Frame para el resultado XPath
        xpath_frame = ctk.CTkFrame(
            self.results_frame_step3,
            fg_color="#f8f9fa",
            corner_radius=8,
            height=80
        )
        xpath_frame.pack(fill="x", padx=5, pady=5)
        xpath_frame.pack_propagate(False)
        
        # Label con el XPath
        xpath_label = ctk.CTkLabel(
            xpath_frame,
            text=self.xref_xpath,
            font=ctk.CTkFont(size=14, family="Courier"),
            text_color="#2c5aa0",
            wraplength=700
        )
        xpath_label.pack(expand=True, pady=10, padx=15)
        
        # Resumen
        summary_label = ctk.CTkLabel(
            self.results_frame_step3,
            text="✅ XPath del xref_id almacenado para filtros posteriores",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#28a745"
        )
        summary_label.pack(pady=10)
        
        # Frame para botones de acción
        action_frame = ctk.CTkFrame(self.results_frame_step3, fg_color="transparent")
        action_frame.pack(pady=15)
        
        # Botón para escoger otro JSON (Paso 1)
        new_json_btn = ctk.CTkButton(
            action_frame,
            text="📁 Escoger otro JSON (Paso 1)",
            width=200,
            height=40,
            font=ctk.CTkFont(size=12),
            command=lambda: self.handle_step_click("step1")
        )
        new_json_btn.pack(side="left", padx=5)
        
        # Botón para volver al paso 2
        back_btn = ctk.CTkButton(
            action_frame,
            text="⬅️ Volver al Paso 2",
            width=160,
            height=40,
            font=ctk.CTkFont(size=12),
            command=lambda: self.handle_step_click("step2")
        )
        back_btn.pack(side="left", padx=5)
        
        # Botón para ir al siguiente paso
        next_step_btn = ctk.CTkButton(
            action_frame,
            text="➡️ Paso 4: Extraer xref_id",
            width=180,
            height=40,
            font=ctk.CTkFont(size=12),
            command=lambda: self.handle_step_click("step4")
        )
        next_step_btn.pack(side="left", padx=5)
    
    def show_step4_view(self):
        """Vista del Paso 4: Definir xpath de campos objetivos"""
        self.clear_main_content()
        self.create_content_header("📋 Paso 4: Definir xpath de campos objetivos", "Configurar campos para comparación XML vs BD")
        
        content_frame = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        content_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=10)  # Cambié nsew por ew
        content_frame.grid_columnconfigure(0, weight=1)
        
        # Información del paso anterior
        info_label = ctk.CTkLabel(
            content_frame,
            text=f"✅ XPath del xref_id configurado: {self.xref_xpath}",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        info_label.pack(pady=(15, 20))
        
        # Título del formulario
        form_title = ctk.CTkLabel(
            content_frame,
            text="🎯 Configuración de Campos Objetivos",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#1f2937"
        )
        form_title.pack(pady=(0, 15))
        
        # Frame para el formulario de campos (altura ajustada)
        self.fields_form_frame = ctk.CTkFrame(content_frame, fg_color="#f8fafc", corner_radius=8, height=350)
        self.fields_form_frame.pack(fill="x", padx=20, pady=10)
        self.fields_form_frame.pack_propagate(False)  # Evitar que se expanda
        
        # Frame para agregar nuevo campo
        add_field_frame = ctk.CTkFrame(self.fields_form_frame, fg_color="white", corner_radius=6)
        add_field_frame.pack(fill="x", padx=15, pady=15)
        
        add_title = ctk.CTkLabel(
            add_field_frame,
            text="➕ Agregar Nuevo Campo",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#2563eb"
        )
        add_title.pack(pady=(10, 5))
        
        # Inputs para nuevo campo
        input_frame = ctk.CTkFrame(add_field_frame, fg_color="transparent")
        input_frame.pack(fill="x", padx=15, pady=10)
        input_frame.grid_columnconfigure(1, weight=1)
        input_frame.grid_columnconfigure(3, weight=1)
        input_frame.grid_columnconfigure(5, weight=1)
        
        # Tabla BD
        ctk.CTkLabel(input_frame, text="Tabla BD:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=0, padx=(0, 5), pady=5, sticky="w")
        self.table_entry = ctk.CTkComboBox(
            input_frame, 
            values=["dispatch", "nfirs_notification", "nfirs_notification_apparatus", "nfirs_notification_personel", "otro"],
            width=150,
            state="readonly"
        )
        self.table_entry.grid(row=0, column=1, padx=(0, 10), pady=5, sticky="w")
        self.table_entry.set("dispatch")  # Valor por defecto
        
        # Campo BD
        ctk.CTkLabel(input_frame, text="Campo BD:", font=ctk.CTkFont(weight="bold")).grid(row=0, column=2, padx=(0, 5), pady=5, sticky="w")
        self.db_field_entry = ctk.CTkEntry(input_frame, placeholder_text="nombre_campo_bd", width=120)
        self.db_field_entry.grid(row=0, column=3, padx=(0, 10), pady=5, sticky="w")
        
        # XPath
        ctk.CTkLabel(input_frame, text="XPath:", font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, padx=(0, 5), pady=5, sticky="w")
        self.xpath_entry = ctk.CTkEntry(input_frame, placeholder_text="//xpath/al/campo", width=350)
        self.xpath_entry.grid(row=1, column=1, columnspan=4, padx=(0, 10), pady=5, sticky="ew")
        
        # Botón agregar
        add_btn = ctk.CTkButton(
            input_frame,
            text="➕ Agregar",
            width=100,
            height=32,
            command=self.add_target_field
        )
        add_btn.grid(row=1, column=5, padx=10, pady=5)
        
        # Frame para mostrar campos configurados (con altura muy limitada)
        self.configured_fields_frame = ctk.CTkScrollableFrame(
            self.fields_form_frame,
            height=150,  # Altura más reducida para dejar espacio a botones
            fg_color="white",
            corner_radius=6
        )
        self.configured_fields_frame.pack(fill="x", padx=15, pady=(0, 10))
        
        # Mostrar campos existentes
        self.update_fields_display()
        
        # Frame para botones de acción
        action_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        action_frame.pack(side="bottom", fill="x", padx=20, pady=20)
        
        # Botón continuar (solo si hay al menos 1 campo)
        if len(self.target_fields) > 0:
            continue_btn = ctk.CTkButton(
                action_frame,
                text="➡️ Paso 5: Seleccionar XMLs",
                width=200,
                height=40,
                font=ctk.CTkFont(size=12, weight="bold"),
                command=lambda: self.handle_step_click("step5")
            )
            continue_btn.pack(side="right", padx=5)
        
        # Botón volver
        back_btn = ctk.CTkButton(
            action_frame,
            text="⬅️ Volver al Paso 3",
            width=160,
            height=40,
            font=ctk.CTkFont(size=12),
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=lambda: self.handle_step_click("step3")
        )
        back_btn.pack(side="left", padx=5)
    
    def add_target_field(self):
        """Agregar un nuevo campo objetivo"""
        table = self.table_entry.get().strip()
        db_field = self.db_field_entry.get().strip()
        xpath = self.xpath_entry.get().strip()
        
        if not table or not db_field or not xpath:
            messagebox.showwarning("Advertencia", "Todos los campos son obligatorios")
            return
        
        # Verificar que no exista ya
        for field in self.target_fields:
            if field['table'] == table and field['db_field'] == db_field:
                messagebox.showwarning("Advertencia", f"El campo '{table}.{db_field}' ya existe")
                return
        
        # Agregar el campo
        self.target_fields.append({
            'table': table,
            'db_field': db_field,
            'xpath': xpath
        })
        
        # Si es el primer campo agregado, actualizar progreso
        if len(self.target_fields) == 1:
            self.update_progress(4)  # Marcar Paso 4 como completado
        
        # Limpiar inputs
        self.table_entry.set("dispatch")  # Resetear ComboBox al valor por defecto
        self.db_field_entry.delete(0, 'end')
        self.xpath_entry.delete(0, 'end')
        
        # Actualizar display
        self.update_fields_display()
        self.update_tracking_display()  # Actualizar panel de seguimiento
        self.show_step4_view()  # Refrescar para mostrar botón continuar
        
        messagebox.showinfo("Éxito", f"Campo '{table}.{db_field}' agregado correctamente")
    
    def remove_target_field(self, index):
        """Eliminar un campo objetivo"""
        if 0 <= index < len(self.target_fields):
            removed_field = self.target_fields.pop(index)
            
            # Si no quedan campos, actualizar progreso
            if len(self.target_fields) == 0:
                self.update_progress(3)  # Retroceder a paso 3
            
            self.update_fields_display()
            self.update_tracking_display()  # Actualizar panel de seguimiento
            self.show_step4_view()  # Refrescar vista
            messagebox.showinfo("Campo eliminado", f"Campo '{removed_field['db_field']}' eliminado")
    
    def update_fields_display(self):
        """Actualizar la visualización de campos configurados"""
        # Limpiar contenido anterior
        for widget in self.configured_fields_frame.winfo_children():
            widget.destroy()
        
        if not self.target_fields:
            # Mensaje cuando no hay campos
            no_fields_label = ctk.CTkLabel(
                self.configured_fields_frame,
                text="📝 No hay campos configurados.\nAgrega al menos 1 campo para continuar.",
                font=ctk.CTkFont(size=14),
                text_color="#6b7280"
            )
            no_fields_label.pack(pady=50)
            return
        
        # Título de la lista
        list_title = ctk.CTkLabel(
            self.configured_fields_frame,
            text=f"📋 Campos Configurados ({len(self.target_fields)})",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#1f2937"
        )
        list_title.pack(pady=(10, 15))
        
        # Mostrar cada campo
        for i, field in enumerate(self.target_fields):
            field_frame = ctk.CTkFrame(self.configured_fields_frame, fg_color="#f1f5f9", corner_radius=6)
            field_frame.pack(fill="x", padx=10, pady=5)
            
            # Contenido del campo
            content_frame = ctk.CTkFrame(field_frame, fg_color="transparent")
            content_frame.pack(fill="x", padx=15, pady=10)
            content_frame.grid_columnconfigure(1, weight=1)
            
            # Número del campo
            number_label = ctk.CTkLabel(
                content_frame,
                text=f"{i+1}.",
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color="#2563eb",
                width=35
            )
            number_label.grid(row=0, column=0, sticky="w")
            
            # Información del campo
            info_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
            info_frame.grid(row=0, column=1, sticky="ew", padx=(10, 0))
            
            # Tabla BD
            table_label = ctk.CTkLabel(
                info_frame,
                text=f"🏢 Tabla: {field['table']}",
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color="#3b82f6"
            )
            table_label.pack(anchor="w")
            
            # Campo BD
            db_label = ctk.CTkLabel(
                info_frame,
                text=f"🗄️ Campo: {field['db_field']}",
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color="#059669"
            )
            db_label.pack(anchor="w")
            
            # XPath con scroll horizontal
            xpath_frame = ctk.CTkFrame(info_frame, fg_color="#fff7ed", corner_radius=4)
            xpath_frame.pack(fill="x", pady=(5, 0))
            
            # Crear un texto scrollable para XPath largo
            xpath_scrollable = ctk.CTkTextbox(
                xpath_frame,
                height=30,
                wrap="none",  # No hacer wrap para permitir scroll horizontal
                font=ctk.CTkFont(size=12, family="Courier"),
                text_color="#7c2d12",
                fg_color="#fff7ed"
            )
            xpath_scrollable.pack(fill="x", padx=5, pady=5)
            
            # Insertar el XPath completo
            xpath_scrollable.insert("1.0", f"🔍 XPath: {field['xpath']}")
            xpath_scrollable.configure(state="disabled")  # Solo lectura
            
            # Botón eliminar
            delete_btn = ctk.CTkButton(
                content_frame,
                text="🗑️",
                width=30,
                height=30,
                font=ctk.CTkFont(size=12),
                fg_color="#dc2626",
                hover_color="#b91c1c",
                command=lambda idx=i: self.remove_target_field(idx)
            )
            delete_btn.grid(row=0, column=2, padx=(10, 0))
    
    def show_step5_view(self):
        """Vista del Paso 5: Seleccionar carpeta de XMLs"""
        self.clear_main_content()
        self.create_content_header("� Paso 5: Seleccionar carpeta de XMLs", "Seleccionar carpeta con archivos XML para comparar")
        
        content_frame = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        
        # Información del paso anterior (más compacta)
        info_label = ctk.CTkLabel(
            content_frame,
            text=f"✅ Campos configurados: {len(self.target_fields)} campos listos para comparar",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        info_label.pack(pady=(10, 5))
        
        # Instrucciones
        instruction_label = ctk.CTkLabel(
            content_frame,
            text="📂 Selecciona la carpeta que contiene los archivos XML\nque se van a comparar con la base de datos",
            font=ctk.CTkFont(size=14),
            text_color="#374151"
        )
        instruction_label.pack(pady=15)
        
        # Botón de selección
        select_folder_btn = ctk.CTkButton(
            content_frame,
            text="📁 Seleccionar Carpeta XML",
            width=250,
            height=50,
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self.select_xml_folder
        )
        select_folder_btn.pack(pady=20)
        
        # Mostrar carpeta seleccionada si existe
        if self.xml_folder_path:
            folder_info_frame = ctk.CTkFrame(content_frame, fg_color="#ecfdf5", corner_radius=8)
            folder_info_frame.pack(fill="x", padx=20, pady=10)
            
            folder_title = ctk.CTkLabel(
                folder_info_frame,
                text="✅ Carpeta XML seleccionada:",
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color="#059669"
            )
            folder_title.pack(pady=(10, 5))
            
            folder_path_label = ctk.CTkLabel(
                folder_info_frame,
                text=self.xml_folder_path,
                font=ctk.CTkFont(size=11, family="Courier"),
                text_color="#047857",
                wraplength=600
            )
            folder_path_label.pack(padx=15, pady=(0, 5))
            
            # Contar archivos XML en la carpeta
            try:
                xml_files = [f for f in os.listdir(self.xml_folder_path) if f.lower().endswith('.xml')]
                count_label = ctk.CTkLabel(
                    folder_info_frame,
                    text=f"📄 Archivos XML encontrados: {len(xml_files)}",
                    font=ctk.CTkFont(size=12, weight="bold"),
                    text_color="#047857"
                )
                count_label.pack(pady=(0, 10))
                
                # Mostrar xref_ids encontrados si existen
                if hasattr(self, 'extracted_xref_ids') and self.extracted_xref_ids:
                    xref_frame = ctk.CTkFrame(folder_info_frame, fg_color="#f0f9ff", corner_radius=6)
                    xref_frame.pack(fill="x", padx=10, pady=(10, 0))
                    
                    xref_title = ctk.CTkLabel(
                        xref_frame,
                        text="🔍 XRef IDs encontrados en XMLs:",
                        font=ctk.CTkFont(size=12, weight="bold"),
                        text_color="#1d4ed8"
                    )
                    xref_title.pack(pady=(10, 5))
                    
                    # Frame scrollable para los xref_ids (reducido para mejor visibilidad de botones)
                    xref_scroll_frame = ctk.CTkScrollableFrame(
                        xref_frame,
                        height=120,  # Reducido de 250 a 120 para dar más espacio a botones
                        fg_color="#ffffff"
                    )
                    xref_scroll_frame.pack(fill="x", padx=10, pady=(0, 10))
                    
                    # Mostrar cada xref_id (reducido para mejor performance)
                    for i, xref_data in enumerate(self.extracted_xref_ids[:15]):  # Reducido de 50 a 15
                        xref_item_frame = ctk.CTkFrame(xref_scroll_frame, fg_color="#f8fafc", corner_radius=4)
                        xref_item_frame.pack(fill="x", pady=2)
                        
                        # Información del xref_id
                        xref_info = ctk.CTkLabel(
                            xref_item_frame,
                            text=f"{i+1}. {xref_data['value']}",
                            font=ctk.CTkFont(size=11, weight="bold"),
                            text_color="#059669"
                        )
                        xref_info.pack(anchor="w", padx=8, pady=2)
                        
                        # Archivo fuente
                        file_info = ctk.CTkLabel(
                            xref_item_frame,
                            text=f"   📄 Archivo: {xref_data['file']}",
                            font=ctk.CTkFont(size=9),
                            text_color="#6b7280"
                        )
                        file_info.pack(anchor="w", padx=8, pady=(0, 2))
                    
                    if len(self.extracted_xref_ids) > 15:
                        more_label = ctk.CTkLabel(
                            xref_scroll_frame,
                            text=f"... y {len(self.extracted_xref_ids) - 15} más",
                            font=ctk.CTkFont(size=10),
                            text_color="#6b7280"
                        )
                        more_label.pack(pady=5)
                    
                    # Resumen
                    summary_label = ctk.CTkLabel(
                        xref_frame,
                        text=f"📊 Total: {len(self.extracted_xref_ids)} XRef IDs únicos encontrados",
                        font=ctk.CTkFont(size=11, weight="bold"),
                        text_color="#1d4ed8"
                    )
                    summary_label.pack(pady=(5, 10))
            except Exception:
                pass
        
        # Frame para botones de acción (siempre visible al final con fondo destacado)
        action_frame = ctk.CTkFrame(content_frame, fg_color="#f8f9fa", corner_radius=8)
        action_frame.pack(side="bottom", fill="x", padx=10, pady=(20, 15))
        
        # Contenedor interno para centrar botones
        buttons_container = ctk.CTkFrame(action_frame, fg_color="transparent")
        buttons_container.pack(expand=True, pady=15)
        
        # Botón volver (siempre visible)
        back_btn = ctk.CTkButton(
            buttons_container,
            text="⬅️ Volver al Paso 4",
            width=160,
            height=45,
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=lambda: self.handle_step_click("step4")
        )
        back_btn.pack(side="left", padx=(0, 15))
        
        # Botón continuar (solo si hay carpeta seleccionada)
        if self.xml_folder_path:
            continue_btn = ctk.CTkButton(
                buttons_container,
                text="➡️ Paso 6: Verificar BD",
                width=200,
                height=45,
                font=ctk.CTkFont(size=13, weight="bold"),
                fg_color="#059669",
                hover_color="#047857",
                command=lambda: self.handle_step_click("step6")
            )
            continue_btn.pack(side="left", padx=(15, 0))
    
    def select_xml_folder(self):
        """Seleccionar carpeta de archivos XML"""
        folder_path = filedialog.askdirectory(
            title="Seleccionar carpeta con archivos XML"
        )
        
        if folder_path:
            # Verificar que la carpeta contenga archivos XML
            try:
                xml_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.xml')]
                if len(xml_files) == 0:
                    messagebox.showwarning("Advertencia", "La carpeta seleccionada no contiene archivos XML")
                    return
                
                self.xml_folder_path = folder_path
                self.update_progress(5)  # Actualizar barra de progreso
                
                # Extraer xref_ids de los XMLs
                self.extract_xref_ids_from_xmls()
                
                self.show_step5_view()  # Refrescar vista
                self.update_tracking_display()  # Actualizar panel de seguimiento
                messagebox.showinfo("Éxito", f"Carpeta seleccionada: {len(xml_files)} archivos XML encontrados")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al acceder a la carpeta: {str(e)}")
    
    def extract_xref_ids_from_xmls(self):
        """Extraer xref_ids de los archivos XML usando XPath específico del JSON"""
        if not self.xml_folder_path:
            return
        
        self.extracted_xref_ids = []
        
        # Solo procesar si tenemos el XPath específico del JSON
        if not hasattr(self, 'xref_xpath') or not self.xref_xpath:
            print("No se encontró XPath para xref_id en el JSON")
            return
        
        try:
            # Obtener todos los archivos XML de la carpeta
            xml_files = [f for f in os.listdir(self.xml_folder_path) if f.lower().endswith('.xml')]
            print(f"Procesando {len(xml_files)} archivos XML con XPath: {self.xref_xpath}")
            
            for xml_file in xml_files:
                xml_path = os.path.join(self.xml_folder_path, xml_file)
                
                try:
                    # Parsear el archivo XML
                    tree = ET.parse(xml_path)
                    root = tree.getroot()
                    
                    # Usar el XPath específico del JSON
                    try:
                        # Extraer valores usando diferentes métodos de búsqueda
                        xref_values = self.extract_values_from_xml(root, self.xref_xpath)
                        
                        for value in xref_values:
                            if value and value.strip():
                                xref_data = {
                                    'value': value.strip(),
                                    'file': xml_file,
                                    'type': 'xpath_match',
                                    'xpath': self.xref_xpath
                                }
                                # Verificar que no esté duplicado
                                if not any(item['value'] == value.strip() for item in self.extracted_xref_ids):
                                    self.extracted_xref_ids.append(xref_data)
                                    print(f"✅ Encontrado XRef ID: {value.strip()} en {xml_file}")
                    
                    except Exception as e:
                        print(f"Error procesando XPath {self.xref_xpath} en {xml_file}: {e}")
                
                except ET.ParseError as e:
                    print(f"Error parsing XML file {xml_file}: {e}")
                except Exception as e:
                    print(f"Error processing XML file {xml_file}: {e}")
        
        except Exception as e:
            print(f"Error accessing XML folder: {e}")
            messagebox.showerror("Error", f"Error al acceder a los archivos XML: {str(e)}")
    
        print(f"Total XRef IDs extraídos usando XPath específico: {len(self.extracted_xref_ids)}")
    
    def extract_values_from_xml(self, root, xpath):
        """Extraer valores del XML usando diferentes métodos de búsqueda"""
        values = []
        
        try:
            # Método 1: Buscar usando XPath directo con findall
            if xpath.startswith('//'):
                # XPath absoluto: //FMPDSORESULT/ROW[1]/CAD
                try:
                    elements = root.findall(xpath[2:])  # Quitar // del inicio
                    for elem in elements:
                        if elem.text:
                            values.append(elem.text)
                except Exception as e:
                    print(f"Error con findall: {e}")
            
            # Método 2: Buscar paso a paso manualmente
            if not values:
                values.extend(self.manual_xpath_search(root, xpath))
            
            # Método 3: Buscar por tag final en todo el árbol
            if not values:
                final_tag = xpath.split('/')[-1]
                if '[' in final_tag:
                    final_tag = final_tag.split('[')[0]
                
                for elem in root.iter():
                    if elem.tag.endswith(final_tag) or elem.tag == final_tag:
                        if elem.text and elem.text.strip():
                            values.append(elem.text)
                            break  # Solo tomar el primero encontrado
            
        except Exception as e:
            print(f"Error en extract_values_from_xml: {e}")
        
        return values
    
    def manual_xpath_search(self, root, xpath):
        """Búsqueda manual paso a paso del XPath"""
        values = []
        
        try:
            # Limpiar y dividir XPath
            clean_xpath = xpath.strip().lstrip('./')
            if clean_xpath.startswith('//'):
                clean_xpath = clean_xpath[2:]
            
            parts = [p for p in clean_xpath.split('/') if p]
            
            if not parts:
                return values
            
            current_elements = [root]
            
            for part in parts:
                if not current_elements:
                    break
                    
                new_elements = []
                
                if '[' in part and ']' in part:
                    # Manejar predicados como ROW[1]
                    tag_name = part.split('[')[0]
                    index_str = part.split('[')[1].split(']')[0]
                    
                    try:
                        index = int(index_str) - 1  # XPath es 1-indexado
                    except ValueError:
                        index = 0
                    
                    for elem in current_elements:
                        matching_children = []
                        for child in elem:
                            if child.tag.endswith(tag_name) or child.tag == tag_name:
                                matching_children.append(child)
                        
                        if matching_children and 0 <= index < len(matching_children):
                            new_elements.append(matching_children[index])
                else:
                    # Buscar elementos normales
                    for elem in current_elements:
                        for child in elem:
                            if child.tag.endswith(part) or child.tag == part:
                                new_elements.append(child)
                
                current_elements = new_elements
            
            # Extraer valores de los elementos finales
            for elem in current_elements:
                if elem.text and elem.text.strip():
                    values.append(elem.text.strip())
        
        except Exception as e:
            print(f"Error en manual_xpath_search: {e}")
        
        return values
    
    def show_step6_view(self):
        """Vista del Paso 6: Definir filtros para BD"""
        self.clear_main_content()
        self.create_content_header("🎯 Paso 6: Definir Filtros para BD", "Los registros de la BD a comparar se filtrarán por estos campos")
        
        content_frame = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        content_frame.grid_rowconfigure(2, weight=1)  # main_scroll_frame expandible
        content_frame.grid_columnconfigure(0, weight=1)
        
        # Información del paso anterior y XRef IDs disponibles
        xref_ids_count = len(self.extracted_xref_ids) if hasattr(self, 'extracted_xref_ids') else 0
        info_label = ctk.CTkLabel(
            content_frame,
            text=f"✅ XRef IDs encontrados: {xref_ids_count} valores únicos",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#10b981"
        )
        info_label.grid(row=0, column=0, pady=(15, 10), sticky="w", padx=20)
        
        # Mostrar XRef IDs disponibles si existen
        if hasattr(self, 'extracted_xref_ids') and self.extracted_xref_ids:
            xref_display_frame = ctk.CTkFrame(content_frame, fg_color="#f0f9ff", corner_radius=8)
            xref_display_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 10))
            xref_display_frame.grid_columnconfigure(0, weight=1)
            
            xref_title = ctk.CTkLabel(
                xref_display_frame,
                text="🔍 XRef IDs disponibles para filtros:",
                font=ctk.CTkFont(size=12, weight="bold"),
                text_color="#1e40af"
            )
            xref_title.pack(pady=(10, 5), padx=15, anchor="w")
            
            # Mostrar hasta 10 XRef IDs como ejemplo
            sample_xrefs = [item['value'] for item in self.extracted_xref_ids[:10]]
            xref_text = ", ".join(sample_xrefs)
            if len(self.extracted_xref_ids) > 10:
                xref_text += f"... (+{len(self.extracted_xref_ids) - 10} más)"
            
            xref_values_label = ctk.CTkLabel(
                xref_display_frame,
                text=xref_text,
                font=ctk.CTkFont(size=11),
                text_color="#374151",
                wraplength=600
            )
            xref_values_label.pack(pady=(0, 10), padx=15, anchor="w")
        
        # Frame principal scrollable
        main_scroll_frame = ctk.CTkScrollableFrame(content_frame, fg_color="transparent")
        main_scroll_frame.grid(row=2, column=0, sticky="nsew", padx=20, pady=10)
        main_scroll_frame.grid_columnconfigure(0, weight=1)
        
        # Formulario para agregar filtros
        filter_form_frame = ctk.CTkFrame(main_scroll_frame, fg_color="#f8fafc", corner_radius=8)
        filter_form_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        filter_form_frame.grid_columnconfigure(1, weight=1)
        filter_form_frame.grid_columnconfigure(2, weight=1)
        filter_form_frame.grid_columnconfigure(3, weight=1)
        
        form_title = ctk.CTkLabel(
            filter_form_frame,
            text="� Agregar Nuevo Filtro",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#1e293b"
        )
        form_title.grid(row=0, column=0, columnspan=4, pady=(15, 20), padx=20)
        
        # Tabla
        table_label = ctk.CTkLabel(filter_form_frame, text="Tabla:", font=ctk.CTkFont(size=12, weight="bold"))
        table_label.grid(row=1, column=0, sticky="w", padx=20, pady=(0, 5))
        
        self.filter_table_entry = ctk.CTkComboBox(
            filter_form_frame,
            values=["dispatch", "nfirs_notification", "nfirs_notification_apparatus", "nfirs_notification_personel", "otro"],
            width=150,
            height=35,
            state="readonly"
        )
        self.filter_table_entry.grid(row=2, column=0, padx=20, pady=(0, 15), sticky="ew")
        self.filter_table_entry.set("dispatch")  # Valor por defecto
        
        # Nombre del Campo
        field_label = ctk.CTkLabel(filter_form_frame, text="Nombre del Campo:", font=ctk.CTkFont(size=12, weight="bold"))
        field_label.grid(row=1, column=1, sticky="w", padx=20, pady=(0, 5))
        
        self.filter_field_entry = ctk.CTkEntry(
            filter_form_frame,
            width=200,
            height=35,
            placeholder_text="ej: xref_id, created_at"
        )
        self.filter_field_entry.grid(row=2, column=1, padx=20, pady=(0, 15), sticky="ew")
        
        # Valor
        value_label = ctk.CTkLabel(filter_form_frame, text="Valor:", font=ctk.CTkFont(size=12, weight="bold"))
        value_label.grid(row=1, column=2, sticky="w", padx=20, pady=(0, 5))
        
        self.filter_value_combo = ctk.CTkComboBox(
            filter_form_frame,
            width=250,
            height=35,
            values=["Escribir valor..."],
            state="normal"
        )
        self.filter_value_combo.grid(row=2, column=2, padx=20, pady=(0, 15), sticky="ew")
        
        # Cuando el campo es xref_id, cargar los valores encontrados
        def on_field_change(*args):
            field_value = self.filter_field_entry.get().lower().strip()
            if field_value in ['xref_id', 'dispatch_number'] and hasattr(self, 'extracted_xref_ids') and self.extracted_xref_ids:
                # Cargar xref_ids encontrados
                xref_values = [item['value'] for item in self.extracted_xref_ids[:20]]  # Limitar a 20 para el ComboBox
                self.filter_value_combo.configure(values=xref_values + ["Escribir valor..."])
                self.filter_value_combo.set("Seleccionar XRef ID...")
            else:
                self.filter_value_combo.configure(values=["Escribir valor..."])
                self.filter_value_combo.set("")
        
        self.filter_field_entry.bind("<KeyRelease>", on_field_change)
        
        # Si ya hay XRef IDs extraídos y el campo es xref_id, cargarlos automáticamente
        if hasattr(self, 'extracted_xref_ids') and self.extracted_xref_ids:
            # Pre-configurar para xref_id si hay valores disponibles
            self.filter_field_entry.insert(0, "xref_id")
            xref_values = [item['value'] for item in self.extracted_xref_ids[:20]]
            self.filter_value_combo.configure(values=xref_values + ["Escribir valor..."])
            self.filter_value_combo.set("Seleccionar XRef ID...")
        
        # Botón agregar filtro
        add_filter_btn = ctk.CTkButton(
            filter_form_frame,
            text="➕ Agregar Filtro",
            width=140,
            height=35,
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.add_filter
        )
        add_filter_btn.grid(row=3, column=0, columnspan=4, pady=(10, 20))
        
        # Lista de filtros configurados
        filters_list_frame = ctk.CTkFrame(main_scroll_frame, fg_color="#ffffff", corner_radius=8)
        filters_list_frame.grid(row=1, column=0, sticky="ew", pady=(20, 20))
        filters_list_frame.grid_columnconfigure(0, weight=1)
        
        self.filters_display_frame = ctk.CTkFrame(filters_list_frame, fg_color="transparent")
        self.filters_display_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 20))
        self.filters_display_frame.grid_columnconfigure(0, weight=1)
        
        # Actualizar display
        self.update_filters_display()
        
        # Frame para botones de acción
        action_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        action_frame.grid(row=3, column=0, pady=20, sticky="ew")
        action_frame.grid_columnconfigure(1, weight=1)
        
        # Botón volver
        back_btn = ctk.CTkButton(
            action_frame,
            text="⬅️ Volver al Paso 5",
            width=160,
            height=40,
            font=ctk.CTkFont(size=12),
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=lambda: self.handle_step_click("step5")
        )
        back_btn.grid(row=0, column=0, padx=(20, 10), sticky="w")
        
        # Botón continuar (solo si hay filtros manuales)
        if hasattr(self, 'filter_criteria') and len(self.filter_criteria) > 0:
            continue_btn = ctk.CTkButton(
                action_frame,
                text="➡️ Paso 7: Comparar",
                width=180,
                height=40,
                font=ctk.CTkFont(size=12, weight="bold"),
                command=lambda: self.handle_step_click("step7")
            )
            continue_btn.grid(row=0, column=2, padx=(10, 20), sticky="e")
    
    def add_filter(self):
        """Agregar un nuevo filtro"""
        table = self.filter_table_entry.get().strip()
        field = self.filter_field_entry.get().strip()
        value = self.filter_value_combo.get().strip()
        
        if not table or not field or not value or value in ["Escribir valor...", "Seleccionar XRef ID..."]:
            messagebox.showwarning("Advertencia", "Todos los campos son obligatorios")
            return
        
        # Verificar que no exista ya el mismo filtro
        for filter_item in self.filter_criteria:
            if (filter_item['table'] == table and 
                filter_item['field'] == field and 
                filter_item['value'] == value):
                messagebox.showwarning("Advertencia", f"El filtro '{table}.{field} = {value}' ya existe")
                return
        
        # Agregar el filtro
        self.filter_criteria.append({
            'table': table,
            'field': field,
            'value': value
        })
        
        # Si es el primer filtro agregado, actualizar progreso
        if len(self.filter_criteria) == 1:
            self.update_progress(6)  # Marcar Paso 6 como completado
        
        # Limpiar inputs
        self.filter_table_entry.set("dispatch")  # Resetear a valor por defecto
        self.filter_field_entry.delete(0, 'end')
        self.filter_value_combo.set("")
        
        # Actualizar displays
        self.update_filters_display()
        self.update_tracking_display()
        self.show_step6_view()  # Refrescar para mostrar botón continuar
        
        messagebox.showinfo("Éxito", f"Filtro '{table}.{field} = {value}' agregado correctamente")
    
    def remove_filter(self, index):
        """Eliminar un filtro"""
        if 0 <= index < len(self.filter_criteria):
            removed_filter = self.filter_criteria.pop(index)
            
            # Si no quedan filtros, actualizar progreso
            if len(self.filter_criteria) == 0:
                self.update_progress(5)  # Retroceder a paso 5
            
            self.update_filters_display()
            self.update_tracking_display()
            self.show_step6_view()  # Refrescar vista
            messagebox.showinfo("Filtro eliminado", f"Filtro '{removed_filter['table']}.{removed_filter['field']} = {removed_filter['value']}' eliminado")
    
    def update_filters_display(self):
        """Actualizar la visualización de filtros configurados"""
        # Limpiar contenido anterior
        for widget in self.filters_display_frame.winfo_children():
            widget.destroy()
        
        if not hasattr(self, 'filter_criteria') or not self.filter_criteria:
            # Mensaje cuando no hay filtros
            no_filters_label = ctk.CTkLabel(
                self.filters_display_frame,
                text="📝 No hay filtros configurados.\nAgrega al menos 1 filtro para continuar.",
                font=ctk.CTkFont(size=14),
                text_color="#6b7280"
            )
            no_filters_label.grid(row=0, column=0, pady=50)
            return
        
        # Título de la lista
        list_title = ctk.CTkLabel(
            self.filters_display_frame,
            text=f"🎯 Filtros Configurados ({len(self.filter_criteria)})",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#1f2937"
        )
        list_title.grid(row=0, column=0, pady=(10, 15), sticky="w")
        
        # Mostrar cada filtro
        for i, filter_item in enumerate(self.filter_criteria):
            filter_frame = ctk.CTkFrame(self.filters_display_frame, fg_color="#f1f5f9", corner_radius=6)
            filter_frame.grid(row=i+1, column=0, sticky="ew", pady=5)
            filter_frame.grid_columnconfigure(1, weight=1)
            
            # Contenido del filtro
            content_frame = ctk.CTkFrame(filter_frame, fg_color="transparent")
            content_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=15, pady=10)
            content_frame.grid_columnconfigure(1, weight=1)
            
            # Número del filtro
            number_label = ctk.CTkLabel(
                content_frame,
                text=f"{i+1}.",
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color="#2563eb",
                width=35
            )
            number_label.grid(row=0, column=0, sticky="w")
            
            # Información del filtro
            filter_text = f"📋 {filter_item['table']}.{filter_item['field']} = {filter_item['value']}"
            filter_info = ctk.CTkLabel(
                content_frame,
                text=filter_text,
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color="#059669"
            )
            filter_info.grid(row=0, column=1, sticky="w", padx=(10, 0))
            
            # Botón eliminar
            delete_btn = ctk.CTkButton(
                content_frame,
                text="🗑️",
                width=30,
                height=30,
                font=ctk.CTkFont(size=12),
                fg_color="#dc2626",
                hover_color="#b91c1c",
                command=lambda idx=i: self.remove_filter(idx)
            )
            delete_btn.grid(row=0, column=2, padx=(10, 0))
    
    def show_step7_view(self):
        """Vista del Paso 7: Comparar campos XML vs BD"""
        self.clear_main_content()
        self.create_content_header("⚖️ Paso 7: Comparar Campos", "Comparar campos XML vs Base de Datos")
        
        content_frame = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        
        # Resumen de configuración
        summary_frame = ctk.CTkFrame(content_frame, fg_color="#f0f9ff", corner_radius=8)
        summary_frame.pack(fill="x", padx=20, pady=20)
        
        summary_title = ctk.CTkLabel(
            summary_frame,
            text="📋 Resumen de Configuración",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#1d4ed8"
        )
        summary_title.pack(pady=(15, 10))
        
        # Información de campos objetivo
        fields_info = ctk.CTkLabel(
            summary_frame,
            text=f"🎯 Campos a comparar: {len(self.target_fields)} configurados",
            font=ctk.CTkFont(size=12),
            text_color="#374151"
        )
        fields_info.pack(pady=2)
        
        # Información de filtros
        filters_info = ctk.CTkLabel(
            summary_frame,
            text=f"🔍 Filtros BD: {len(self.filter_criteria)} configurados",
            font=ctk.CTkFont(size=12),
            text_color="#374151"
        )
        filters_info.pack(pady=2)
        
        # Información de XMLs
        xml_info = ctk.CTkLabel(
            summary_frame,
            text=f"📄 XRef IDs: {len(self.extracted_xref_ids) if hasattr(self, 'extracted_xref_ids') else 0} encontrados",
            font=ctk.CTkFont(size=12),
            text_color="#374151"
        )
        xml_info.pack(pady=(2, 15))
        
        # Instrucciones
        instruction_label = ctk.CTkLabel(
            content_frame,
            text="🚀 Todo está configurado correctamente.\nHaz clic en 'Iniciar Comparación' para procesar los datos.",
            font=ctk.CTkFont(size=14),
            text_color="#374151"
        )
        instruction_label.pack(pady=20)
        
        # Botón de ejecución
        execute_btn = ctk.CTkButton(
            content_frame,
            text="🚀 Iniciar Comparación",
            width=250,
            height=50,
            font=ctk.CTkFont(size=16, weight="bold"),
            fg_color="#10b981",
            hover_color="#059669",
            command=self.execute_comparison
        )
        execute_btn.pack(pady=20)
        
        # Frame para botones de navegación
        nav_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        nav_frame.pack(pady=20)
        
        # Botón volver
        back_btn = ctk.CTkButton(
            nav_frame,
            text="⬅️ Volver al Paso 6",
            width=160,
            height=40,
            font=ctk.CTkFont(size=12),
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=lambda: self.handle_step_click("step6")
        )
        back_btn.pack(side="left", padx=5)
    
    def execute_comparison(self):
        """Ejecutar la comparación de campos XML vs BD"""
        try:
            # Validar que toda la configuración esté lista
            if not self.validate_comparison_config():
                return
            
            # Mostrar ventana de progreso
            progress_window = self.create_progress_window("Ejecutando Comparación...")
            
            # Ejecutar comparación en hilo separado para no bloquear UI
            import threading
            comparison_thread = threading.Thread(target=self.run_comparison_process, args=(progress_window,))
            comparison_thread.daemon = True
            comparison_thread.start()
            
        except Exception as e:
            self.logger.error(f"Error iniciando comparación: {str(e)}")
            messagebox.showerror("Error", f"Error iniciando la comparación: {str(e)}")
    
    def validate_comparison_config(self):
        """Validar que toda la configuración esté completa"""
        errors = []
        
        # Verificar campos objetivo
        if not hasattr(self, 'target_fields') or not self.target_fields:
            errors.append("• No hay campos objetivo configurados (Paso 4)")
        
        # Verificar carpeta XML y XRef IDs extraídos
        if not hasattr(self, 'xml_folder_path') or not self.xml_folder_path:
            errors.append("• No se ha seleccionado carpeta XML (Paso 5)")
        
        if not hasattr(self, 'extracted_xref_ids') or not self.extracted_xref_ids:
            errors.append("• No se han extraído XRef IDs de los XMLs (Paso 5)")
        
        # Verificar filtros BD
        if not hasattr(self, 'filter_criteria') or not self.filter_criteria:
            errors.append("• No hay filtros configurados para BD (Paso 6)")
        
        # Verificar XPath
        if not hasattr(self, 'xref_xpath') or not self.xref_xpath:
            errors.append("• No se ha extraído XPath del XRef ID (Paso 3)")
        
        if errors:
            error_msg = "Configuración incompleta:\n\n" + "\n".join(errors)
            messagebox.showerror("Error de Configuración", error_msg)
            return False
        
        return True
    
    def create_progress_window(self, title):
        """Crear ventana de progreso"""
        progress_window = ctk.CTkToplevel(self.root)
        progress_window.title(title)
        progress_window.geometry("500x300")
        progress_window.transient(self.root)
        progress_window.grab_set()
        
        # Centrar ventana
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - (500 // 2)
        y = (progress_window.winfo_screenheight() // 2) - (300 // 2)
        progress_window.geometry(f"+{x}+{y}")
        
        # Contenido de la ventana
        main_frame = ctk.CTkFrame(progress_window)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        title_label = ctk.CTkLabel(main_frame, text=title, font=ctk.CTkFont(size=16, weight="bold"))
        title_label.pack(pady=(20, 10))
        
        # Barra de progreso
        progress_bar = ctk.CTkProgressBar(main_frame, width=400, height=20)
        progress_bar.pack(pady=20)
        progress_bar.set(0)
        
        # Texto de estado
        status_label = ctk.CTkLabel(main_frame, text="Iniciando comparación...", wraplength=400)
        status_label.pack(pady=10)
        
        # Almacenar referencias en variables de instancia
        self._progress_widgets = {
            'window': progress_window,
            'progress_bar': progress_bar,
            'status_label': status_label
        }
        
        return progress_window
    
    def run_comparison_process(self, progress_window):
        """Ejecutar el proceso de comparación en hilo separado"""
        try:
            self.update_progress_window(progress_window, 0.1, "Conectando a base de datos...")
            
            # Importar field_comparator con ruta absoluta
            import sys
            import os
            sys.path.append(os.path.dirname(__file__))
            from field_comparator import FieldComparator
            
            # Inicializar comparador con configuración de BD
            config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'db_config.json')
            if not os.path.exists(config_path):
                config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'config.json')
            
            comparator = FieldComparator(config_path)
            
            self.update_progress_window(progress_window, 0.2, "Configuración cargada, conectando a BD...")
            
            # Conectar a BD
            if not comparator.connect_to_db():
                raise Exception("No se pudo conectar a la base de datos")
            
            self.update_progress_window(progress_window, 0.3, "Procesando archivos XML...")
            
            # Obtener archivos XML de la carpeta
            if not self.xml_folder_path:
                raise Exception("No se ha seleccionado carpeta XML")
                
            xml_files = [f for f in os.listdir(self.xml_folder_path) if f.lower().endswith('.xml')]
            
            if not xml_files:
                raise Exception(f"No se encontraron archivos XML en: {self.xml_folder_path}")
            
            comparison_results = []
            total_files = len(xml_files)
            
            for i, xml_file in enumerate(xml_files):
                xml_path = os.path.join(self.xml_folder_path, xml_file)
                progress = 0.3 + (0.5 * (i + 1) / total_files)
                self.update_progress_window(progress_window, progress, f"Procesando {xml_file}...")
                
                # Procesar cada archivo XML
                file_results = self.process_xml_file(xml_path, xml_file, comparator)
                if file_results:
                    comparison_results.extend(file_results)
            
            self.update_progress_window(progress_window, 0.9, "Generando reporte...")
            
            # Generar reporte Excel
            report_path = self.generate_comparison_report(comparison_results)
            
            self.update_progress_window(progress_window, 1.0, "¡Comparación completada!")
            
            # Cerrar conexión BD
            comparator.close_db_connection()
            
            # Mostrar resultado
            self.root.after(1000, lambda: self.show_comparison_results(progress_window, report_path, len(comparison_results)))
            
            # Actualizar progreso
            self.update_progress(7)
            self.update_tracking_display()
            
        except Exception as e:
            error_msg = str(e)
            self.logger.error(f"Error en comparación: {error_msg}")
            self.root.after(100, lambda: self.show_comparison_error(progress_window, error_msg))
    
    def update_progress_window(self, window, progress, message):
        """Actualizar ventana de progreso desde hilo"""
        def update_ui():
            if (hasattr(self, '_progress_widgets') and 
                self._progress_widgets['window'] and 
                self._progress_widgets['window'].winfo_exists()):
                self._progress_widgets['progress_bar'].set(progress)
                self._progress_widgets['status_label'].configure(text=message)
                self._progress_widgets['window'].update()
        
        self.root.after(0, update_ui)
    
    def process_xml_file(self, xml_path, xml_filename, comparator):
        """Procesar un archivo XML individual"""
        try:
            # Extraer XRef ID del XML usando el XPath configurado
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            # Usar XPath para encontrar el XRef ID
            if not self.xref_xpath:
                print(f"No hay XPath configurado para XRef ID en {xml_filename}")
                return []
                
            xref_values = self.extract_values_from_xml(root, self.xref_xpath)
            
            if not xref_values:
                print(f"No se encontró XRef ID en {xml_filename} usando XPath: {self.xref_xpath}")
                return []
            
            xref_id = xref_values[0]  # Tomar el primer valor encontrado
            
            if not xref_id:
                print(f"XRef ID vacío en {xml_filename}")
                return []
            
            # Verificar si este XRef ID cumple con los filtros configurados
            if not self.xref_matches_filters(xref_id):
                print(f"XRef ID {xref_id} no cumple con los filtros configurados")
                return []
            
            # Buscar registro en BD
            db_record = self.find_db_record(xref_id, comparator)
            
            if not db_record:
                return [{
                    'xml_file': xml_filename,
                    'xref_id': xref_id,
                    'campo': 'REGISTRO_BD',
                    'valor_xml': 'N/A',
                    'valor_bd': 'NO_ENCONTRADO',
                    'estado': 'ERROR',
                    'filtros_usados': self.get_filters_display()
                }]
            
            # Comparar campos objetivo
            file_results = []
            
            # Agregar información de filtros usados
            filter_info = {
                'xml_file': xml_filename,
                'xref_id': xref_id,
                'campo': 'FILTROS_APLICADOS',
                'valor_xml': 'N/A',
                'valor_bd': self.get_filters_display(),
                'estado': 'INFO',
                'filtros_usados': self.get_filters_display()
            }
            file_results.append(filter_info)
            
            # Comparar cada campo objetivo
            for field_config in self.target_fields:
                field_result = self.compare_field(xml_path, db_record, field_config, xml_filename, xref_id)
                file_results.append(field_result)
            
            return file_results
            
        except Exception as e:
            print(f"Error procesando {xml_filename}: {str(e)}")
            return [{
                'xml_file': xml_filename,
                'xref_id': 'ERROR',
                'campo': 'PROCESAMIENTO',
                'valor_xml': 'ERROR',
                'valor_bd': str(e),
                'estado': 'ERROR',
                'filtros_usados': self.get_filters_display()
            }]
    
    def xref_matches_filters(self, xref_id):
        """Verificar si el XRef ID cumple con los filtros configurados"""
        # Por ahora, permitir todos los XRef IDs que estén en extracted_xref_ids
        # y que coincidan con los filtros de xref_id configurados
        
        for filter_item in self.filter_criteria:
            if filter_item['field'].lower() in ['xref_id', 'dispatch_number']:
                if filter_item['value'] == xref_id:
                    return True
        
        # Si no hay filtros específicos de xref_id, permitir todos los extraídos
        xref_filter_exists = any(f['field'].lower() in ['xref_id', 'dispatch_number'] for f in self.filter_criteria)
        if not xref_filter_exists:
            # Verificar que esté en los XRef IDs extraídos
            return any(item['value'] == xref_id for item in self.extracted_xref_ids)
        
        return False
    
    def find_db_record(self, xref_id, comparator):
        """Buscar registro en BD usando los filtros configurados"""
        try:
            cursor = comparator.db_connection.cursor()
            
            # Construir consulta SQL con los filtros
            where_conditions = []
            params = []
            
            for filter_item in self.filter_criteria:
                table_field = f"{filter_item['table']}.{filter_item['field']}"
                if filter_item['field'].lower() in ['xref_id', 'dispatch_number']:
                    # Para campos XRef, usar el XRef ID del XML
                    where_conditions.append(f"{table_field} = %s")
                    params.append(xref_id)
                else:
                    # Para otros campos, usar el valor configurado
                    where_conditions.append(f"{table_field} = %s")
                    params.append(filter_item['value'])
            
            # Obtener las tablas únicas para el FROM
            tables = list(set(f['table'] for f in self.filter_criteria))
            main_table = tables[0] if tables else 'dispatch'
            
            # Construir consulta
            query = f"SELECT * FROM {main_table}"
            if len(tables) > 1:
                # Si hay múltiples tablas, agregar JOINs (simplificado)
                for table in tables[1:]:
                    query += f" LEFT JOIN {table} ON {main_table}.xref_id = {table}.xref_id"
            
            if where_conditions:
                query += " WHERE " + " AND ".join(where_conditions)
            
            query += " LIMIT 1"
            
            cursor.execute(query, params)
            result = cursor.fetchone()
            
            if result:
                # Convertir a diccionario usando nombres de columnas
                columns = [desc[0] for desc in cursor.description]
                return dict(zip(columns, result))
            
            return None
            
        except Exception as e:
            print(f"Error buscando en BD para XRef ID {xref_id}: {str(e)}")
            return None
    
    def compare_field(self, xml_path, db_record, field_config, xml_filename, xref_id):
        """Comparar un campo específico entre XML y BD"""
        try:
            # Extraer valor del XML usando XPath
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            xml_elements = root.findall(field_config['xpath'].replace('//', './/'))
            xml_value = xml_elements[0].text if xml_elements and xml_elements[0].text else 'NULL'
            
            # Obtener valor de BD
            db_field = field_config['db_field']
            db_value = str(db_record.get(db_field, 'NULL')) if db_record else 'NULL'
            
            # Comparar valores
            if xml_value == db_value:
                estado = 'COINCIDE'
            elif xml_value == 'NULL' or db_value == 'NULL':
                estado = 'ERROR'
            else:
                estado = 'NO_COINCIDE'
            
            return {
                'xml_file': xml_filename,
                'xref_id': xref_id,
                'campo': db_field,
                'valor_xml': xml_value,
                'valor_bd': db_value,
                'estado': estado,
                'xpath_usado': field_config['xpath'],
                'filtros_usados': self.get_filters_display()
            }
            
        except Exception as e:
            return {
                'xml_file': xml_filename,
                'xref_id': xref_id,
                'campo': field_config['db_field'],
                'valor_xml': 'ERROR',
                'valor_bd': 'ERROR',
                'estado': 'ERROR',
                'xpath_usado': field_config.get('xpath', 'N/A'),
                'error': str(e),
                'filtros_usados': self.get_filters_display()
            }
    
    def get_filters_display(self):
        """Obtener texto de los filtros para mostrar en reporte"""
        if not hasattr(self, 'filter_criteria') or not self.filter_criteria:
            return "Sin filtros"
        
        filters_text = []
        for f in self.filter_criteria:
            filters_text.append(f"{f['table']}.{f['field']} = {f['value']}")
        
        return " | ".join(filters_text)
    
    def generate_comparison_report(self, comparison_results):
        """Generar reporte Excel con los resultados"""
        try:
            # Crear DataFrame
            df = pd.DataFrame(comparison_results)
            
            # Crear directorio de reportes si no existe
            reports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'reportes')
            os.makedirs(reports_dir, exist_ok=True)
            
            # Nombre del archivo
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"reporte_comparacion_campos_{timestamp}.xlsx"
            report_path = os.path.join(reports_dir, filename)
            
            # Crear archivo Excel con formato
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Comparacion_Campos', index=False)
                
                # Obtener workbook y worksheet para formatear
                workbook = writer.book
                worksheet = writer.sheets['Comparacion_Campos']
                
                # Definir estilos
                from openpyxl.styles import PatternFill, Font
                
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                blue_fill = PatternFill(start_color="C6E2FF", end_color="C6E2FF", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                
                # Aplicar colores según estado
                for row in range(2, len(df) + 2):  # Empezar desde fila 2 (después del header)
                    estado_cell = worksheet[f'E{row}']  # Columna 'estado'
                    estado_value = estado_cell.value
                    
                    if estado_value == 'COINCIDE':
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = green_fill
                    elif estado_value == 'NO_COINCIDE':
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = red_fill
                    elif estado_value == 'ERROR':
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = blue_fill
                    elif estado_value == 'INFO':
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = yellow_fill
                
                # Ajustar ancho de columnas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            return report_path
            
        except Exception as e:
            self.logger.error(f"Error generando reporte: {str(e)}")
            raise
    
    def show_comparison_results(self, progress_window, report_path, total_comparisons):
        """Mostrar resultados de la comparación"""
        if progress_window and progress_window.winfo_exists():
            progress_window.destroy()
        
        result_msg = f"""✅ ¡Comparación completada exitosamente!

📊 Resultados:
• Total de comparaciones: {total_comparisons}
• Reporte generado: {os.path.basename(report_path)}

📁 Ubicación del reporte:
{report_path}

🎨 Código de colores en Excel:
• 🟢 Verde: Valores coinciden
• 🔴 Rojo: Valores no coinciden  
• 🔵 Azul: Error en comparación
• 🟡 Amarillo: Información de filtros

¿Deseas abrir el reporte ahora?"""
        
        result = messagebox.askyesno("Comparación Completada", result_msg)
        if result:
            try:
                os.startfile(report_path)  # Windows
            except:
                import subprocess
                subprocess.run(['open', report_path])  # macOS/Linux
    
    def show_comparison_error(self, progress_window, error_message):
        """Mostrar error de comparación"""
        if progress_window and progress_window.winfo_exists():
            progress_window.destroy()
        
        messagebox.showerror("Error en Comparación", f"Error durante la comparación:\n\n{error_message}")
    
    def run(self):
        """Ejecutar la aplicación"""
        self.root.mainloop()

def main():
    """Función principal"""
    try:
        app = ModernWebGUI()
        app.run()
    except Exception as e:
        print(f"Error iniciando la aplicación: {e}")
        messagebox.showerror("Error", f"Error iniciando la aplicación: {e}")

if __name__ == "__main__":
    main()
