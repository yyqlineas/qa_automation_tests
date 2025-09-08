                        # Formato: "Nombre de fichero: xref_id"
                        combined_label = ctk.CTkLabel(
                            item_frame,
                            text=f"� {xref_file}: {xref_value}",
                            font=ctk.CTkFont(size=12, weight="bold"),
                            text_color="#1565c0",
                            anchor="w"
                        )
                        combined_label.pack(fill="x", padx=12, pady=8)
                    
            except Exception:
                pass
        
        # Frame para botones de acción (siempre visible al final con fondo destacado)
        action_frame = ctk.CTkFrame(content_frame, fg_color="#f8f9fa", corner_radius=8)
        action_frame.pack(side="bottom", fill="x", padx=10, pady=(5, 10))  # Reducido pady de (20, 15) a (5, 10)
        
        # Contenedor interno para centrar botones
        buttons_container = ctk.CTkFrame(action_frame, fg_color="transparent")
        buttons_container.pack(expand=True, pady=10)  # Reducido pady de 15 a 10
        
        # Botón volver (siempre visible)
        back_btn = ctk.CTkButton(
            buttons_container,
            text="⬅️ Volver al Paso 4",
            width=160,
            height=35,  # Reducido de 45 a 35
            font=ctk.CTkFont(size=12, weight="bold"),  # Reducido font de 13 a 12
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=lambda: self.handle_step_click("step4")
        )
        back_btn.pack(side="left", padx=(0, 15))
        
        # Botón continuar (solo si hay carpeta seleccionada)
        if self.xml_folder_path:
            continue_btn = ctk.CTkButton(
                buttons_container,
                text="➡️ Paso 6: Verificar BD",
                width=200,
                height=35,  # Reducido de 45 a 35
                font=ctk.CTkFont(size=12, weight="bold"),  # Reducido font de 13 a 12
                fg_color="#059669",
                hover_color="#047857",
                command=lambda: self.handle_step_click("step6")
            )
            continue_btn.pack(side="left", padx=(15, 0))
    
    def select_xml_folder(self):
        """Seleccionar carpeta de archivos XML"""
        folder_path = filedialog.askdirectory(
            title="Seleccionar carpeta con archivos XML"
        )
        
        if folder_path:
            # Verificar que la carpeta contenga archivos XML
            try:
                xml_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.xml')]
                if len(xml_files) == 0:
                    messagebox.showwarning("Advertencia", "La carpeta seleccionada no contiene archivos XML")
                    return
                
                self.xml_folder_path = folder_path
                self.update_progress(5)  # Actualizar barra de progreso
                
                # Extraer xref_ids de los XMLs
                self.extract_xref_ids_from_xmls()
                
                self.show_step5_view()  # Refrescar vista
                self.update_tracking_display()  # Actualizar panel de seguimiento
                messagebox.showinfo("Éxito", f"Carpeta seleccionada: {len(xml_files)} archivos XML encontrados")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al acceder a la carpeta: {str(e)}")
    
    def extract_xref_ids_from_xmls(self):
        """Extraer xref_ids de los archivos XML usando XPath específico del JSON"""
        if not self.xml_folder_path:
            return
        
        self.extracted_xref_ids = []
        
        # Solo procesar si tenemos el XPath específico del JSON
        if not hasattr(self, 'xref_xpath') or not self.xref_xpath:
            print("No se encontró XPath para xref_id en el JSON")
            return
        
        try:
            # Obtener todos los archivos XML de la carpeta
            xml_files = [f for f in os.listdir(self.xml_folder_path) if f.lower().endswith('.xml')]
            print(f"Procesando {len(xml_files)} archivos XML con XPath: {self.xref_xpath}")
            
            for xml_file in xml_files:
                xml_path = os.path.join(self.xml_folder_path, xml_file)
                
                try:
                    # Parsear el archivo XML
                    tree = ET.parse(xml_path)
                    root = tree.getroot()
                    
                    # Usar el XPath específico del JSON
                    try:
                        # Extraer valores usando diferentes métodos de búsqueda
                        xref_values = self.extract_values_from_xml(root, self.xref_xpath)
                        
                        for value in xref_values:
                            if value and value.strip():
                                xref_data = {
                                    'value': value.strip(),
                                    'file': xml_file,
                                    'type': 'xpath_match',
                                    'xpath': self.xref_xpath
                                }
                                # Verificar que no esté duplicado
                                if not any(item['value'] == value.strip() for item in self.extracted_xref_ids):
                                    self.extracted_xref_ids.append(xref_data)
                                    print(f"✅ Encontrado XRef ID: {value.strip()} en {xml_file}")
                    
                    except Exception as e:
                        print(f"Error procesando XPath {self.xref_xpath} en {xml_file}: {e}")
                
                except ET.ParseError as e:
                    print(f"Error parsing XML file {xml_file}: {e}")
                except Exception as e:
                    print(f"Error processing XML file {xml_file}: {e}")
        
        except Exception as e:
            print(f"Error accessing XML folder: {e}")
            messagebox.showerror("Error", f"Error al acceder a los archivos XML: {str(e)}")
    
        print(f"Total XRef IDs extraídos usando XPath específico: {len(self.extracted_xref_ids)}")
    
    def extract_values_from_xml(self, root, xpath):
        """Extraer valores del XML usando diferentes métodos de búsqueda"""
        values = []
        
        try:
            # Método 1: Buscar usando XPath directo con findall
            if xpath.startswith('//'):
                # XPath absoluto: //FMPDSORESULT/ROW[1]/CAD
                try:
                    elements = root.findall(xpath[2:])  # Quitar // del inicio
                    for elem in elements:
                        if elem.text:
                            values.append(elem.text)
                except Exception as e:
                    print(f"Error con findall: {e}")
            
            # Método 2: Buscar paso a paso manualmente
            if not values:
                values.extend(self.manual_xpath_search(root, xpath))
            
            # Método 3: Buscar por tag final en todo el árbol
            if not values:
                final_tag = xpath.split('/')[-1]
                if '[' in final_tag:
                    final_tag = final_tag.split('[')[0]
                
                for elem in root.iter():
                    if elem.tag.endswith(final_tag) or elem.tag == final_tag:
                        if elem.text and elem.text.strip():
                            values.append(elem.text)
                            break  # Solo tomar el primero encontrado
            
        except Exception as e:
            print(f"Error en extract_values_from_xml: {e}")
        
        return values
    
    def manual_xpath_search(self, root, xpath):
        """Búsqueda manual paso a paso del XPath"""
        values = []
        
        try:
            # Limpiar y dividir XPath
            clean_xpath = xpath.strip().lstrip('./')
            if clean_xpath.startswith('//'):
                clean_xpath = clean_xpath[2:]
            
            parts = [p for p in clean_xpath.split('/') if p]
            
            if not parts:
                return values
            
            current_elements = [root]
            
            for part in parts:
                if not current_elements:
                    break
                    
                new_elements = []
                
                if '[' in part and ']' in part:
                    # Manejar predicados como ROW[1]
                    tag_name = part.split('[')[0]
                    index_str = part.split('[')[1].split(']')[0]
                    
                    try:
                        index = int(index_str) - 1  # XPath es 1-indexado
                    except ValueError:
                        index = 0
                    
                    for elem in current_elements:
                        matching_children = []
                        for child in elem:
                            if child.tag.endswith(tag_name) or child.tag == tag_name:
                                matching_children.append(child)
                        
                        if matching_children and 0 <= index < len(matching_children):
                            new_elements.append(matching_children[index])
                else:
                    # Buscar elementos normales
                    for elem in current_elements:
                        for child in elem:
                            if child.tag.endswith(part) or child.tag == part:
                                new_elements.append(child)
                
                current_elements = new_elements
            
            # Extraer valores de los elementos finales
            for elem in current_elements:
                if elem.text and elem.text.strip():
                    values.append(elem.text.strip())
        
        except Exception as e:
            print(f"Error en manual_xpath_search: {e}")
        
        return values
    
    def show_step6_view(self):
        """Vista del Paso 6: Definir filtros para BD"""
        self.clear_main_content()
        self.create_content_header("🎯 Paso 6: Definir Filtros para BD", "Los registros de la BD a comparar se filtrarán por estos campos")
        
        content_frame = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        content_frame.grid_rowconfigure(1, weight=1)  # main_scroll_frame expandible
        content_frame.grid_columnconfigure(0, weight=1)
        
        # Información del paso anterior y XRef IDs disponibles
        xref_ids_count = len(self.extracted_xref_ids) if hasattr(self, 'extracted_xref_ids') else 0
        info_label = ctk.CTkLabel(
            content_frame,
            text=f"✅ XRef IDs encontrados: {xref_ids_count} valores únicos\n🔍 Los XRef IDs están disponibles en el Paso 5 para su revisión",
            font=ctk.CTkFont(size=12),
            text_color="#10b981",
            justify="left"
        )
        info_label.grid(row=0, column=0, pady=(15, 20), sticky="w", padx=20)
        
        # Frame principal scrollable
        main_scroll_frame = ctk.CTkScrollableFrame(content_frame, fg_color="transparent")
        main_scroll_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        main_scroll_frame.grid_columnconfigure(0, weight=1)
        
        # Formulario para agregar filtros
        filter_form_frame = ctk.CTkFrame(main_scroll_frame, fg_color="#f8fafc", corner_radius=8)
        filter_form_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        filter_form_frame.grid_columnconfigure(1, weight=1)
        filter_form_frame.grid_columnconfigure(2, weight=1)
        filter_form_frame.grid_columnconfigure(3, weight=1)
        
        form_title = ctk.CTkLabel(
            filter_form_frame,
            text="� Agregar Nuevo Filtro",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#1e293b"
        )
        form_title.grid(row=0, column=0, columnspan=4, pady=(15, 20), padx=20)
        
        # Tabla
        table_label = ctk.CTkLabel(filter_form_frame, text="Tabla:", font=ctk.CTkFont(size=12, weight="bold"))
        table_label.grid(row=1, column=0, sticky="w", padx=20, pady=(0, 5))
        
        self.filter_table_entry = ctk.CTkComboBox(
            filter_form_frame,
            values=["dispatch", "nfirs_notification", "nfirs_notification_apparatus", "nfirs_notification_personel", "otro"],
            width=150,
            height=35,
            state="readonly"
        )
        self.filter_table_entry.grid(row=2, column=0, padx=20, pady=(0, 15), sticky="ew")
        self.filter_table_entry.set("dispatch")  # Valor por defecto
        
        # Nombre del Campo
        field_label = ctk.CTkLabel(filter_form_frame, text="Nombre del Campo:", font=ctk.CTkFont(size=12, weight="bold"))
        field_label.grid(row=1, column=1, sticky="w", padx=20, pady=(0, 5))
        
        self.filter_field_entry = ctk.CTkEntry(
            filter_form_frame,
            width=200,
            height=35,
            placeholder_text="ej: xref_id, created_at"
        )
        self.filter_field_entry.grid(row=2, column=1, padx=20, pady=(0, 15), sticky="ew")
        
        # Valor
        value_label = ctk.CTkLabel(filter_form_frame, text="Valor:", font=ctk.CTkFont(size=12, weight="bold"))
        value_label.grid(row=1, column=2, sticky="w", padx=20, pady=(0, 5))
        
        self.filter_value_combo = ctk.CTkComboBox(
            filter_form_frame,
            width=250,
            height=35,
            values=["Escribir valor..."],
            state="normal"
        )
        self.filter_value_combo.grid(row=2, column=2, padx=20, pady=(0, 15), sticky="ew")
        
        # Cuando el campo es xref_id, cargar los valores encontrados
        def on_field_change(*args):
            field_value = self.filter_field_entry.get().lower().strip()
            if field_value in ['xref_id', 'dispatch_number'] and hasattr(self, 'extracted_xref_ids') and self.extracted_xref_ids:
                # Cargar xref_ids encontrados
                xref_values = [item['value'] for item in self.extracted_xref_ids[:20]]  # Limitar a 20 para el ComboBox
                self.filter_value_combo.configure(values=xref_values + ["Escribir valor..."])
                self.filter_value_combo.set("Seleccionar XRef ID...")
            else:
                self.filter_value_combo.configure(values=["Escribir valor..."])
                self.filter_value_combo.set("")
        
        self.filter_field_entry.bind("<KeyRelease>", on_field_change)
        
        # Si ya hay XRef IDs extraídos y el campo es xref_id, cargarlos automáticamente
        if hasattr(self, 'extracted_xref_ids') and self.extracted_xref_ids:
            # Pre-configurar para xref_id si hay valores disponibles
            self.filter_field_entry.insert(0, "xref_id")
            xref_values = [item['value'] for item in self.extracted_xref_ids[:20]]
            self.filter_value_combo.configure(values=xref_values + ["Escribir valor..."])
            self.filter_value_combo.set("Seleccionar XRef ID...")
        
        # Botón agregar filtro
        add_filter_btn = ctk.CTkButton(
            filter_form_frame,
            text="➕ Agregar Filtro",
            width=140,
            height=35,
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.add_filter
        )
        add_filter_btn.grid(row=3, column=0, columnspan=4, pady=(10, 20))
        
        # Lista de filtros configurados
        filters_list_frame = ctk.CTkFrame(main_scroll_frame, fg_color="#ffffff", corner_radius=8)
        filters_list_frame.grid(row=1, column=0, sticky="ew", pady=(20, 20))
        filters_list_frame.grid_columnconfigure(0, weight=1)
        
        self.filters_display_frame = ctk.CTkFrame(filters_list_frame, fg_color="transparent")
        self.filters_display_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 20))
        self.filters_display_frame.grid_columnconfigure(0, weight=1)
        
        # Actualizar display
        self.update_filters_display()
        
        # Frame para botones de acción
        action_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        action_frame.grid(row=3, column=0, pady=20, sticky="ew")
        action_frame.grid_columnconfigure(1, weight=1)
        
        # Botón volver
        back_btn = ctk.CTkButton(
            action_frame,
            text="⬅️ Volver al Paso 5",
            width=160,
            height=40,
            font=ctk.CTkFont(size=12),
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=lambda: self.handle_step_click("step5")
        )
        back_btn.grid(row=0, column=0, padx=(20, 10), sticky="w")
        
        # Botón continuar (solo si hay filtros manuales)
        if hasattr(self, 'filter_criteria') and len(self.filter_criteria) > 0:
            continue_btn = ctk.CTkButton(
                action_frame,
                text="➡️ Paso 7: Comparar",
                width=180,
                height=40,
                font=ctk.CTkFont(size=12, weight="bold"),
                command=lambda: self.handle_step_click("step7")
            )
            continue_btn.grid(row=0, column=2, padx=(10, 20), sticky="e")
    
    def add_filter(self):
        """Agregar un nuevo filtro"""
        table = self.filter_table_entry.get().strip()
        field = self.filter_field_entry.get().strip()
        value = self.filter_value_combo.get().strip()
        
        if not table or not field or not value or value in ["Escribir valor...", "Seleccionar XRef ID..."]:
            messagebox.showwarning("Advertencia", "Todos los campos son obligatorios")
            return
        
        # Verificar que no exista ya el mismo filtro
        for filter_item in self.filter_criteria:
            if (filter_item['table'] == table and 
                filter_item['field'] == field and 
                filter_item['value'] == value):
                messagebox.showwarning("Advertencia", f"El filtro '{table}.{field} = {value}' ya existe")
                return
        
        # Agregar el filtro
        self.filter_criteria.append({
            'table': table,
            'field': field,
            'value': value
        })
        
        # Si es el primer filtro agregado, actualizar progreso
        if len(self.filter_criteria) == 1:
            self.update_progress(6)  # Marcar Paso 6 como completado
        
        # Limpiar inputs
        self.filter_table_entry.set("dispatch")  # Resetear a valor por defecto
        self.filter_field_entry.delete(0, 'end')
        self.filter_value_combo.set("")
        
        # Actualizar displays
        self.update_filters_display()
        self.update_tracking_display()
        self.show_step6_view()  # Refrescar para mostrar botón continuar
        
        messagebox.showinfo("Éxito", f"Filtro '{table}.{field} = {value}' agregado correctamente")
    
    def remove_filter(self, index):
        """Eliminar un filtro"""
        if 0 <= index < len(self.filter_criteria):
            removed_filter = self.filter_criteria.pop(index)
            
            # Si no quedan filtros, actualizar progreso
            if len(self.filter_criteria) == 0:
                self.update_progress(5)  # Retroceder a paso 5
            
            self.update_filters_display()
            self.update_tracking_display()
            self.show_step6_view()  # Refrescar vista
            messagebox.showinfo("Filtro eliminado", f"Filtro '{removed_filter['table']}.{removed_filter['field']} = {removed_filter['value']}' eliminado")
    
    def update_filters_display(self):
        """Actualizar la visualización de filtros configurados"""
        # Limpiar contenido anterior
        for widget in self.filters_display_frame.winfo_children():
            widget.destroy()
        
        if not hasattr(self, 'filter_criteria') or not self.filter_criteria:
            # Mensaje cuando no hay filtros
            no_filters_label = ctk.CTkLabel(
                self.filters_display_frame,
                text="📝 No hay filtros configurados.\nAgrega al menos 1 filtro para continuar.",
                font=ctk.CTkFont(size=14),
                text_color="#6b7280"
            )
            no_filters_label.grid(row=0, column=0, pady=50)
            return
        
        # Título de la lista
        list_title = ctk.CTkLabel(
            self.filters_display_frame,
            text=f"🎯 Filtros Configurados ({len(self.filter_criteria)})",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#1f2937"
        )
        list_title.grid(row=0, column=0, pady=(10, 15), sticky="w")
        
        # Mostrar cada filtro
        for i, filter_item in enumerate(self.filter_criteria):
            filter_frame = ctk.CTkFrame(self.filters_display_frame, fg_color="#f1f5f9", corner_radius=6)
            filter_frame.grid(row=i+1, column=0, sticky="ew", pady=5)
            filter_frame.grid_columnconfigure(1, weight=1)
            
            # Contenido del filtro
            content_frame = ctk.CTkFrame(filter_frame, fg_color="transparent")
            content_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=15, pady=10)
            content_frame.grid_columnconfigure(1, weight=1)
            
            # Número del filtro
            number_label = ctk.CTkLabel(
                content_frame,
                text=f"{i+1}.",
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color="#2563eb",
                width=35
            )
            number_label.grid(row=0, column=0, sticky="w")
            
            # Información del filtro
            filter_text = f"📋 {filter_item['table']}.{filter_item['field']} = {filter_item['value']}"
            filter_info = ctk.CTkLabel(
                content_frame,
                text=filter_text,
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color="#059669"
            )
            filter_info.grid(row=0, column=1, sticky="w", padx=(10, 0))
            
            # Botón eliminar
            delete_btn = ctk.CTkButton(
                content_frame,
                text="🗑️",
                width=30,
                height=30,
                font=ctk.CTkFont(size=12),
                fg_color="#dc2626",
                hover_color="#b91c1c",
                command=lambda idx=i: self.remove_filter(idx)
            )
            delete_btn.grid(row=0, column=2, padx=(10, 0))
    
    def show_step7_view(self):
        """Vista del Paso 7: Comparar campos XML vs BD"""
        self.clear_main_content()
        self.create_content_header("⚖️ Paso 7: Comparar Campos", "Comparar campos XML vs Base de Datos")
        
        content_frame = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        
        # Resumen de configuración
        summary_frame = ctk.CTkFrame(content_frame, fg_color="#f0f9ff", corner_radius=8)
        summary_frame.pack(fill="x", padx=20, pady=20)
        
        summary_title = ctk.CTkLabel(
            summary_frame,
            text="📋 Resumen de Configuración",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#1d4ed8"
        )
        summary_title.pack(pady=(15, 10))
        
        # Información de campos objetivo
        fields_info = ctk.CTkLabel(
            summary_frame,
            text=f"🎯 Campos a comparar: {len(self.target_fields)} configurados",
            font=ctk.CTkFont(size=12),
            text_color="#374151"
        )
        fields_info.pack(pady=2)
        
        # Información de filtros
        filters_info = ctk.CTkLabel(
            summary_frame,
            text=f"🔍 Filtros BD: {len(self.filter_criteria)} configurados",
            font=ctk.CTkFont(size=12),
            text_color="#374151"
        )
        filters_info.pack(pady=2)
        
        # Información de XMLs
        xml_info = ctk.CTkLabel(
            summary_frame,
            text=f"📄 XRef IDs: {len(self.extracted_xref_ids) if hasattr(self, 'extracted_xref_ids') else 0} encontrados",
            font=ctk.CTkFont(size=12),
            text_color="#374151"
        )
        xml_info.pack(pady=(2, 15))
        
        # Instrucciones
        instruction_label = ctk.CTkLabel(
            content_frame,
            text="🚀 Todo está configurado correctamente.\nHaz clic en 'Iniciar Comparación' para procesar los datos.",
            font=ctk.CTkFont(size=14),
            text_color="#374151"
        )
        instruction_label.pack(pady=20)
        
        # Botón de ejecución
        execute_btn = ctk.CTkButton(
            content_frame,
            text="🚀 Iniciar Comparación",
            width=250,
            height=50,
            font=ctk.CTkFont(size=16, weight="bold"),
            fg_color="#10b981",
            hover_color="#059669",
            command=self.execute_comparison
        )
        execute_btn.pack(pady=20)
        
        # Frame para botones de navegación
        nav_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        nav_frame.pack(pady=20)
        
        # Botón volver
        back_btn = ctk.CTkButton(
            nav_frame,
            text="⬅️ Volver al Paso 6",
            width=160,
            height=40,
            font=ctk.CTkFont(size=12),
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=lambda: self.handle_step_click("step6")
        )
        back_btn.pack(side="left", padx=5)
    
    def execute_comparison(self):
        """Ejecutar la comparación de campos XML vs BD"""
        try:
            # Validar que toda la configuración esté lista
            if not self.validate_comparison_config():
                return
            
            # Mostrar ventana de progreso
            progress_window = self.create_progress_window("Ejecutando Comparación...")
            
            # Ejecutar comparación en hilo separado para no bloquear UI
            import threading
            comparison_thread = threading.Thread(target=self.run_comparison_process, args=(progress_window,))
            comparison_thread.daemon = True
            comparison_thread.start()
            
        except Exception as e:
            self.logger.error(f"Error iniciando comparación: {str(e)}")
            messagebox.showerror("Error", f"Error iniciando la comparación: {str(e)}")
    
    def validate_comparison_config(self):
        """Validar que toda la configuración esté completa"""
        errors = []
        
        # Verificar campos objetivo
        if not hasattr(self, 'target_fields') or not self.target_fields:
            errors.append("• No hay campos objetivo configurados (Paso 4)")
        
        # Verificar carpeta XML y XRef IDs extraídos
        if not hasattr(self, 'xml_folder_path') or not self.xml_folder_path:
            errors.append("• No se ha seleccionado carpeta XML (Paso 5)")
        
        if not hasattr(self, 'extracted_xref_ids') or not self.extracted_xref_ids:
            errors.append("• No se han extraído XRef IDs de los XMLs (Paso 5)")
        
        # Verificar filtros BD
        if not hasattr(self, 'filter_criteria') or not self.filter_criteria:
            errors.append("• No hay filtros configurados para BD (Paso 6)")
        
        # Verificar XPath
        if not hasattr(self, 'xref_xpath') or not self.xref_xpath:
            errors.append("• No se ha extraído XPath del XRef ID (Paso 3)")
        
        if errors:
            error_msg = "Configuración incompleta:\n\n" + "\n".join(errors)
            messagebox.showerror("Error de Configuración", error_msg)
            return False
        
        return True
    
    def create_progress_window(self, title):
        """Crear ventana de progreso"""
        progress_window = ctk.CTkToplevel(self.root)
        progress_window.title(title)
        progress_window.geometry("500x300")
        progress_window.transient(self.root)
        progress_window.grab_set()
        
        # Centrar ventana
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - (500 // 2)
        y = (progress_window.winfo_screenheight() // 2) - (300 // 2)
        progress_window.geometry(f"+{x}+{y}")
        
        # Contenido de la ventana
        main_frame = ctk.CTkFrame(progress_window)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        title_label = ctk.CTkLabel(main_frame, text=title, font=ctk.CTkFont(size=16, weight="bold"))
        title_label.pack(pady=(20, 10))
        
        # Barra de progreso
        progress_bar = ctk.CTkProgressBar(main_frame, width=400, height=20)
        progress_bar.pack(pady=20)
        progress_bar.set(0)
        
        # Texto de estado
        status_label = ctk.CTkLabel(main_frame, text="Iniciando comparación...", wraplength=400)
        status_label.pack(pady=10)
        
        # Almacenar referencias en variables de instancia
        self._progress_widgets = {
            'window': progress_window,
            'progress_bar': progress_bar,
            'status_label': status_label
        }
        
        return progress_window
    
    def run_comparison_process(self, progress_window):
        """Ejecutar el proceso de comparación en hilo separado"""
        try:
            self.update_progress_window(progress_window, 0.1, "Conectando a base de datos...")
            
            # Verificar que tenemos configuración de BD
            if not self.db_config or not all(key in self.db_config for key in ['host', 'port', 'database', 'user', 'password']):
                raise Exception("Configuración de BD incompleta. Por favor configure la conexión a BD.")
            
            # Importar field_comparator con ruta absoluta
            import sys
            import os
            sys.path.append(os.path.dirname(__file__))
            from field_comparator import FieldComparator
            
            # Crear archivo temporal de configuración para FieldComparator
            temp_config = {
                "database": self.db_config
            }
            
            # Guardar configuración temporal
            temp_config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'temp_db_config.json')
            with open(temp_config_path, 'w', encoding='utf-8') as f:
                json.dump(temp_config, f, indent=4, ensure_ascii=False)
            
            # FieldComparator espera el archivo de configuración
            comparator = FieldComparator(temp_config_path)
            
            self.update_progress_window(progress_window, 0.2, "Configuración cargada, conectando a BD...")
            
            # Conectar a BD con mejor manejo de errores
            try:
                if not comparator.connect_to_db():
                    raise Exception(f"Falló la conexión a BD: {self.db_config['host']}:{self.db_config['port']}/{self.db_config['database']}")
            except Exception as db_error:
                raise Exception(f"Error de conexión a BD: {str(db_error)}")
            finally:
                # Limpiar archivo temporal
                if os.path.exists(temp_config_path):
                    os.remove(temp_config_path)
            
            self.update_progress_window(progress_window, 0.3, "Procesando archivos XML...")
            
            # Obtener archivos XML de la carpeta
            if not self.xml_folder_path:
                raise Exception("No se ha seleccionado carpeta XML")
                
            xml_files = [f for f in os.listdir(self.xml_folder_path) if f.lower().endswith('.xml')]
            
            if not xml_files:
                raise Exception(f"No se encontraron archivos XML en: {self.xml_folder_path}")
            
            comparison_results = []
            total_files = len(xml_files)
            
            for i, xml_file in enumerate(xml_files):
                xml_path = os.path.join(self.xml_folder_path, xml_file)
                progress = 0.3 + (0.5 * (i + 1) / total_files)
                self.update_progress_window(progress_window, progress, f"Procesando {xml_file}...")
                
                # Procesar cada archivo XML
                file_results = self.process_xml_file(xml_path, xml_file, comparator)
                if file_results:
                    comparison_results.extend(file_results)
            
            self.update_progress_window(progress_window, 0.9, "Generando reporte...")
            
            # Generar reporte Excel
            report_path = self.generate_comparison_report(comparison_results)
            
            self.update_progress_window(progress_window, 1.0, "¡Comparación completada!")
            
            # Cerrar conexión BD
            comparator.close_db_connection()
            
            # Mostrar resultado
            self.root.after(1000, lambda: self.show_comparison_results(progress_window, report_path, len(comparison_results)))
            
            # Actualizar progreso
            self.update_progress(7)
            self.update_tracking_display()
            
        except Exception as e:
            error_msg = str(e)
            self.logger.error(f"Error en comparación: {error_msg}")
            self.root.after(100, lambda: self.show_comparison_error(progress_window, error_msg))
    
    def update_progress_window(self, window, progress, message):
        """Actualizar ventana de progreso desde hilo"""
        def update_ui():
            if (hasattr(self, '_progress_widgets') and 
                self._progress_widgets['window'] and 
                self._progress_widgets['window'].winfo_exists()):
                self._progress_widgets['progress_bar'].set(progress)
                self._progress_widgets['status_label'].configure(text=message)
                self._progress_widgets['window'].update()
        
        self.root.after(0, update_ui)
    
    def process_xml_file(self, xml_path, xml_filename, comparator):
        """Procesar un archivo XML individual"""
        try:
            # Extraer XRef ID del XML usando el XPath configurado
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            # Usar XPath para encontrar el XRef ID
            if not self.xref_xpath:
                print(f"No hay XPath configurado para XRef ID en {xml_filename}")
                return []
                
            xref_values = self.extract_values_from_xml(root, self.xref_xpath)
            
            if not xref_values:
                print(f"No se encontró XRef ID en {xml_filename} usando XPath: {self.xref_xpath}")
                return []
            
            xref_id = xref_values[0]  # Tomar el primer valor encontrado
            
            if not xref_id:
                print(f"XRef ID vacío en {xml_filename}")
                return []
            
            # Verificar si este XRef ID cumple con los filtros configurados
            if not self.xref_matches_filters(xref_id):
                print(f"XRef ID {xref_id} no cumple con los filtros configurados")
                return []
            
            # Buscar registro en BD
            db_record = self.find_db_record(xref_id, comparator)
            
            if not db_record:
                return [{
                    'xml_file': xml_filename,
                    'batt_dept_id': xref_id,
                    'target_field': 'REGISTRO_BD',
                    'xpath': 'N/A',
                    'valor_xml': 'N/A',
                    'valor_bd': 'NO_ENCONTRADO',
                    'coincide': False,
                    'observaciones': f'No se encontró registro en BD con filtros: {self.get_filters_display()}'
                }]
            
            # Comparar solo los campos objetivo definidos en el Paso 4
            file_results = []
            
            # Comparar cada campo objetivo
            for field_config in self.target_fields:
                field_result = self.compare_field(xml_path, db_record, field_config, xml_filename, xref_id)
                file_results.append(field_result)
            
            return file_results
            
        except Exception as e:
            print(f"Error procesando {xml_filename}: {str(e)}")
            return [{
                'xml_file': xml_filename,
                'batt_dept_id': 'ERROR',
                'target_field': 'PROCESAMIENTO',
                'xpath': 'N/A',
                'valor_xml': 'ERROR',
                'valor_bd': str(e),
                'coincide': False,
                'observaciones': f'Error de procesamiento: {str(e)}'
            }]
    
    def xref_matches_filters(self, xref_id):
        """Verificar si el XRef ID cumple con los filtros configurados"""
        # Por ahora, permitir todos los XRef IDs que estén en extracted_xref_ids
        # y que coincidan con los filtros de xref_id configurados
        
        for filter_item in self.filter_criteria:
            if filter_item['field'].lower() in ['xref_id', 'dispatch_number']:
                if filter_item['value'] == xref_id:
                    return True
        
        # Si no hay filtros específicos de xref_id, permitir todos los extraídos
        xref_filter_exists = any(f['field'].lower() in ['xref_id', 'dispatch_number'] for f in self.filter_criteria)
        if not xref_filter_exists:
            # Verificar que esté en los XRef IDs extraídos
            return any(item['value'] == xref_id for item in self.extracted_xref_ids)
        
        return False
    
    def find_db_record(self, xref_id, comparator):
        """Buscar registro en BD usando el XRef ID extraído del XML"""
        try:
            cursor = comparator.db_connection.cursor()
            
            # CORRECCIÓN: Buscar usando dispatch.dispatch_number que es STRING
            # No intentar convertir xref_id a entero, mantenerlo como string
            print(f"🔍 Buscando registro para XRef ID: {xref_id}")
            
            # Construir consulta base usando dispatch.dispatch_number
            base_query = """
                SELECT d.*, n.*, f.*, h.*, v.*, t.*, a.*
                FROM dispatch d
                LEFT JOIN nfirs_notification n ON d.dispatch_number = n.dispatch_number
                LEFT JOIN nfirs_fire_incident f ON d.dispatch_number = f.dispatch_number  
                LEFT JOIN nfirs_hazmat h ON d.dispatch_number = h.dispatch_number
                LEFT JOIN nfirs_vehicle v ON d.dispatch_number = v.dispatch_number
                LEFT JOIN nfirs_training t ON d.dispatch_number = t.dispatch_number
                LEFT JOIN nfirs_apparatus a ON d.dispatch_number = a.dispatch_number
                WHERE d.dispatch_number = %s
            """
            
            # Agregar filtros adicionales si están configurados
            additional_conditions = []
            params = [xref_id]  # El primer parámetro siempre es el xref_id
            
            if hasattr(self, 'filter_criteria') and self.filter_criteria:
                for filter_item in self.filter_criteria:
                    # Solo agregar filtros que NO sean para dispatch_number 
                    # (ya que ese es nuestro filtro principal)
                    if filter_item['field'].lower() not in ['dispatch_number', 'xref_id']:
                        table = filter_item['table']
                        field = filter_item['field']
                        value = filter_item['value']
                        
                        # Mapear tabla a alias
                        table_alias_map = {
                            'dispatch': 'd',
                            'nfirs_notification': 'n', 
                            'nfirs_fire_incident': 'f',
                            'nfirs_hazmat': 'h',
                            'nfirs_vehicle': 'v',
                            'nfirs_training': 't',
                            'nfirs_apparatus': 'a'
                        }
                        
                        alias = table_alias_map.get(table, table)
                        additional_conditions.append(f"{alias}.{field} = %s")
                        params.append(value)
            
            # Construir consulta final
            final_query = base_query
            if additional_conditions:
                final_query += " AND " + " AND ".join(additional_conditions)
            
            final_query += " LIMIT 1"
            
            print(f"🔍 Consulta SQL: {final_query}")
            print(f"📋 Parámetros: {params}")
            
            cursor.execute(final_query, params)
            result = cursor.fetchone()
            
            if result:
                # Obtener nombres de columnas y crear diccionario
                columns = [desc[0] for desc in cursor.description]
                record_dict = dict(zip(columns, result))
                print(f"✅ Registro encontrado para XRef {xref_id}")
                return record_dict
            else:
                print(f"❌ No se encontró registro para XRef {xref_id}")
                return None
                
        except Exception as e:
            print(f"❌ Error en búsqueda BD para XRef {xref_id}: {str(e)}")
            self.logger.error(f"Error en find_db_record: {str(e)}")
            return None
            
        except Exception as e:
            print(f"❌ Error buscando en BD para XRef ID {xref_id}: {str(e)}")
            return None
    
    def compare_field(self, xml_path, db_record, field_config, xml_filename, xref_id):
        """Comparar un campo específico entre XML y BD"""
        try:
            # Extraer valor del XML usando XPath correcto
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            # Usar el método extract_values_from_xml que ya maneja namespaces
            xml_values = self.extract_values_from_xml(root, field_config['xpath'])
            xml_value = xml_values[0] if xml_values and xml_values[0] else None
            
            # Obtener valor de BD del registro encontrado
            db_field = field_config['db_field']
            db_value = db_record.get(db_field) if db_record else None
            
            # Normalizar valores para comparación
            xml_str = str(xml_value).strip() if xml_value is not None else ""
            db_str = str(db_value).strip() if db_value is not None else ""
            
            # Comparar valores (exacta, case-sensitive)
            coincide_bool = (xml_str == db_str) and xml_str != "" and db_str != ""
            
            # Verificar errores
            if "ERROR" in str(xml_value) or "ERROR" in str(db_value):
                coincide = False
            else:
                coincide = coincide_bool
            
            # Generar observaciones detalladas
            observaciones = self._get_comparison_notes(xml_value, db_value)
            
            # Estructura del reporte según requerimientos del usuario
            return {
                'xml_file': xml_filename,           # Archivo XML
                'batt_dept_id': xref_id,            # XRef ID del XML  
                'target_field': db_field,           # Campo objetivo en BD
                'xpath': field_config['xpath'],     # XPath usado
                'valor_xml': xml_value,             # Valor del XML
                'valor_bd': db_value,               # Valor de la BD
                'coincide': coincide,               # TRUE/FALSE
                'observaciones': observaciones      # Notas detalladas
            }
            
        except Exception as e:
            # Estructura de error según requerimientos del usuario
            return {
                'xml_file': xml_filename,
                'batt_dept_id': xref_id,
                'target_field': field_config['db_field'],
                'xpath': field_config.get('xpath', 'N/A'),
                'valor_xml': f"ERROR: {str(e)}",
                'valor_bd': 'ERROR',
                'coincide': False,
                'observaciones': f"Error de procesamiento: {str(e)}"
            }
    
    def _get_comparison_notes(self, xml_value, db_value):
        """Generar notas sobre la comparación (igual que XML BD Comparator)"""
        # Verificar si hay errores
        if "ERROR" in str(xml_value):
            return "Error al obtener valor XML"
        elif "ERROR" in str(db_value):
            return "Error al obtener valor BD"
        
        # Verificar valores nulos
        xml_is_null = xml_value is None or str(xml_value).strip() == ""
        db_is_null = db_value is None or str(db_value).strip() == ""
        
        if xml_is_null and db_is_null:
            return "Ambos valores son nulos"
        elif xml_is_null:
            return "Valor XML es nulo"
        elif db_is_null:
            return "Valor BD es nulo"
        
        # Comparar valores no nulos
        xml_str = str(xml_value).strip()
        db_str = str(db_value).strip()
        
        if xml_str == db_str:
            return "Valores coinciden"
        else:
            return f"Valores diferentes: XML='{xml_str}' vs BD='{db_str}'"
    
    def get_filters_display(self):
        """Obtener texto de los filtros para mostrar en reporte"""
        if not hasattr(self, 'filter_criteria') or not self.filter_criteria:
            return "Sin filtros"
        
        filters_text = []
        for f in self.filter_criteria:
            filters_text.append(f"{f['table']}.{f['field']} = {f['value']}")
        
        return " | ".join(filters_text)
    
    def generate_comparison_report(self, comparison_results):
        """Generar reporte Excel con los resultados (formato XML BD Comparator)"""
        try:
            # Crear DataFrame
            df = pd.DataFrame(comparison_results)
            
            # Crear directorio de reportes si no existe
            reports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'reportes')
            os.makedirs(reports_dir, exist_ok=True)
            
            # Nombre del archivo
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"reporte_comparacion_campos_{timestamp}.xlsx"
            report_path = os.path.join(reports_dir, filename)
            
            # Crear archivo Excel con formato
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Resultados', index=False)
                
                # Crear resumen
                total_comparaciones = len(comparison_results)
                coincidencias = len([r for r in comparison_results if r.get('coincide', False) == True])
                errores = len([r for r in comparison_results if 'ERROR' in str(r.get('valor_xml', '')) or 'ERROR' in str(r.get('valor_bd', ''))])
                diferencias = total_comparaciones - coincidencias - errores
                
                resumen = pd.DataFrame({
                    'Métrica': ['Total de comparaciones', 'Coincidencias', 'Diferencias', 'Errores', 'Porcentaje de coincidencias'],
                    'Valor': [total_comparaciones, coincidencias, diferencias, errores, f"{(coincidencias/total_comparaciones*100):.2f}%" if total_comparaciones > 0 else "0%"]
                })
                resumen.to_excel(writer, sheet_name='Resumen', index=False)
                
                # Obtener workbook y worksheet para formatear
                workbook = writer.book
                worksheet = writer.sheets['Resultados']
                
                # Ajustar ancho de columnas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar formato de colores igual que XML BD Comparator
            self._apply_color_formatting(report_path, df)
            
            return report_path
            
        except Exception as e:
            print(f"Error generando reporte: {str(e)}")
            return None
    
    def _apply_color_formatting(self, report_path, df):
        """Aplicar formato de colores al reporte (igual que XML BD Comparator)"""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill
            
            workbook = openpyxl.load_workbook(report_path)
            worksheet = workbook['Resultados']
            
            # Definir colores iguales a XML BD Comparator
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Rojo claro
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Amarillo claro
            blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")   # Azul claro
            
            # Buscar la columna de "observaciones"
            observaciones_col = None
            for col_idx, cell in enumerate(worksheet[1], 1):  # Primera fila (encabezados)
                if cell.value and 'observacion' in str(cell.value).lower():
                    observaciones_col = col_idx
                    break
            
            if not observaciones_col:
                print("⚠️ No se encontró la columna 'observaciones' para aplicar colores")
                return
            
            # Aplicar colores según el contenido de observaciones
            for row_idx in range(2, len(df) + 2):  # Empezar desde la fila 2 (datos, no encabezados)
                cell = worksheet.cell(row=row_idx, column=observaciones_col)
                if cell.value:
                    observacion = str(cell.value).lower()
                    
                    if "coinciden" in observacion:
                        # Verde para coincidencias
                        cell.fill = green_fill
                    elif "error" in observacion:
                        # Amarillo para errores
                        cell.fill = yellow_fill
                    elif "diferentes" in observacion:
                        # Rojo para diferencias
                        cell.fill = red_fill
                    elif "nulo" in observacion or "nulos" in observacion:
                        # Azul para valores nulos
                        cell.fill = blue_fill
            
            # Guardar cambios
            workbook.save(report_path)
            print("✅ Formato de colores aplicado al reporte")
            
        except Exception as e:
            print(f"⚠️ Error aplicando colores al reporte: {str(e)}")
    
    def show_comparison_results(self, progress_window, report_path, total_comparisons):
        """Mostrar resultados de la comparación"""
        if progress_window and progress_window.winfo_exists():
            progress_window.destroy()
        
        result_msg = f"""✅ ¡Comparación completada exitosamente!

📊 Resultados:
• Total de comparaciones: {total_comparisons}
• Reporte generado: {os.path.basename(report_path)}

📁 Ubicación del reporte:
{report_path}

🎨 Código de colores en Excel:
• 🟢 Verde: Valores coinciden
• 🔴 Rojo: Valores no coinciden  
• 🔵 Azul: Error en comparación
• 🟡 Amarillo: Información de filtros

¿Deseas abrir el reporte ahora?"""
        
        result = messagebox.askyesno("Comparación Completada", result_msg)
        if result:
            try:
                os.startfile(report_path)  # Windows
            except:
                import subprocess
                subprocess.run(['open', report_path])  # macOS/Linux
    
    def show_comparison_error(self, progress_window, error_message):
        """Mostrar error de comparación"""
        if progress_window and progress_window.winfo_exists():
            progress_window.destroy()
        
        messagebox.showerror("Error en Comparación", f"Error durante la comparación:\n\n{error_message}")
    
    def run(self):
        """Ejecutar la aplicación"""
        self.root.mainloop()

def main():
    """Función principal"""
    try:
        app = ModernWebGUI()
        app.run()
    except Exception as e:
        print(f"Error iniciando la aplicación: {e}")
        messagebox.showerror("Error", f"Error iniciando la aplicación: {e}")

if __name__ == "__main__":
    main()
            except Exception:
                pass
        
        # Frame para botones de acción (siempre visible al final con fondo destacado)
        action_frame = ctk.CTkFrame(content_frame, fg_color="#f8f9fa", corner_radius=8)
        action_frame.pack(side="bottom", fill="x", padx=10, pady=(5, 10))  # Reducido pady de (20, 15) a (5, 10)
        
        # Contenedor interno para centrar botones
        buttons_container = ctk.CTkFrame(action_frame, fg_color="transparent")
        buttons_container.pack(expand=True, pady=10)  # Reducido pady de 15 a 10
        
        # Botón volver (siempre visible)
        back_btn = ctk.CTkButton(
            buttons_container,
            text="⬅️ Volver al Paso 4",
            width=160,
            height=35,  # Reducido de 45 a 35
            font=ctk.CTkFont(size=12, weight="bold"),  # Reducido font de 13 a 12
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=lambda: self.handle_step_click("step4")
        )
        back_btn.pack(side="left", padx=(0, 15))
        
        # Botón continuar (solo si hay carpeta seleccionada)
        if self.xml_folder_path:
            continue_btn = ctk.CTkButton(
                buttons_container,
                text="➡️ Paso 6: Verificar BD",
                width=200,
                height=35,  # Reducido de 45 a 35
                font=ctk.CTkFont(size=12, weight="bold"),  # Reducido font de 13 a 12
                fg_color="#059669",
                hover_color="#047857",
                command=lambda: self.handle_step_click("step6")
            )
            continue_btn.pack(side="left", padx=(15, 0))
    
    def select_xml_folder(self):
        """Seleccionar carpeta de archivos XML"""
        folder_path = filedialog.askdirectory(
            title="Seleccionar carpeta con archivos XML"
        )
        
        if folder_path:
            # Verificar que la carpeta contenga archivos XML
            try:
                xml_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.xml')]
                if len(xml_files) == 0:
                    messagebox.showwarning("Advertencia", "La carpeta seleccionada no contiene archivos XML")
                    return
                
                self.xml_folder_path = folder_path
                self.update_progress(5)  # Actualizar barra de progreso
                
                # Extraer xref_ids de los XMLs
                self.extract_xref_ids_from_xmls()
                
                self.show_step5_view()  # Refrescar vista
                self.update_tracking_display()  # Actualizar panel de seguimiento
                messagebox.showinfo("Éxito", f"Carpeta seleccionada: {len(xml_files)} archivos XML encontrados")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al acceder a la carpeta: {str(e)}")
    
    def extract_xref_ids_from_xmls(self):
        """Extraer xref_ids de los archivos XML usando XPath específico del JSON"""
        if not self.xml_folder_path:
            return
        
        self.extracted_xref_ids = []
        
        # Solo procesar si tenemos el XPath específico del JSON
        if not hasattr(self, 'xref_xpath') or not self.xref_xpath:
            print("No se encontró XPath para xref_id en el JSON")
            return
        
        try:
            # Obtener todos los archivos XML de la carpeta
            xml_files = [f for f in os.listdir(self.xml_folder_path) if f.lower().endswith('.xml')]
            print(f"Procesando {len(xml_files)} archivos XML con XPath: {self.xref_xpath}")
            
            for xml_file in xml_files:
                xml_path = os.path.join(self.xml_folder_path, xml_file)
                
                try:
                    # Parsear el archivo XML
                    tree = ET.parse(xml_path)
                    root = tree.getroot()
                    
                    # Usar el XPath específico del JSON
                    try:
                        # Extraer valores usando diferentes métodos de búsqueda
                        xref_values = self.extract_values_from_xml(root, self.xref_xpath)
                        
                        for value in xref_values:
                            if value and value.strip():
                                xref_data = {
                                    'value': value.strip(),
                                    'file': xml_file,
                                    'type': 'xpath_match',
                                    'xpath': self.xref_xpath
                                }
                                # Verificar que no esté duplicado
                                if not any(item['value'] == value.strip() for item in self.extracted_xref_ids):
                                    self.extracted_xref_ids.append(xref_data)
                                    print(f"✅ Encontrado XRef ID: {value.strip()} en {xml_file}")
                    
                    except Exception as e:
                        print(f"Error procesando XPath {self.xref_xpath} en {xml_file}: {e}")
                
                except ET.ParseError as e:
                    print(f"Error parsing XML file {xml_file}: {e}")
                except Exception as e:
                    print(f"Error processing XML file {xml_file}: {e}")
        
        except Exception as e:
            print(f"Error accessing XML folder: {e}")
            messagebox.showerror("Error", f"Error al acceder a los archivos XML: {str(e)}")
    
        print(f"Total XRef IDs extraídos usando XPath específico: {len(self.extracted_xref_ids)}")
    
    def extract_values_from_xml(self, root, xpath):
        """Extraer valores del XML usando diferentes métodos de búsqueda"""
        values = []
        
        try:
            # Método 1: Buscar usando XPath directo con findall
            if xpath.startswith('//'):
                # XPath absoluto: //FMPDSORESULT/ROW[1]/CAD
                try:
                    elements = root.findall(xpath[2:])  # Quitar // del inicio
                    for elem in elements:
                        if elem.text:
                            values.append(elem.text)
                except Exception as e:
                    print(f"Error con findall: {e}")
            
            # Método 2: Buscar paso a paso manualmente
            if not values:
                values.extend(self.manual_xpath_search(root, xpath))
            
            # Método 3: Buscar por tag final en todo el árbol
            if not values:
                final_tag = xpath.split('/')[-1]
                if '[' in final_tag:
                    final_tag = final_tag.split('[')[0]
                
                for elem in root.iter():
                    if elem.tag.endswith(final_tag) or elem.tag == final_tag:
                        if elem.text and elem.text.strip():
                            values.append(elem.text)
                            break  # Solo tomar el primero encontrado
            
        except Exception as e:
            print(f"Error en extract_values_from_xml: {e}")
        
        return values
    
    def manual_xpath_search(self, root, xpath):
        """Búsqueda manual paso a paso del XPath"""
        values = []
        
        try:
            # Limpiar y dividir XPath
            clean_xpath = xpath.strip().lstrip('./')
            if clean_xpath.startswith('//'):
                clean_xpath = clean_xpath[2:]
            
            parts = [p for p in clean_xpath.split('/') if p]
            
            if not parts:
                return values
            
            current_elements = [root]
            
            for part in parts:
                if not current_elements:
                    break
                    
                new_elements = []
                
                if '[' in part and ']' in part:
                    # Manejar predicados como ROW[1]
                    tag_name = part.split('[')[0]
                    index_str = part.split('[')[1].split(']')[0]
                    
                    try:
                        index = int(index_str) - 1  # XPath es 1-indexado
                    except ValueError:
                        index = 0
                    
                    for elem in current_elements:
                        matching_children = []
                        for child in elem:
                            if child.tag.endswith(tag_name) or child.tag == tag_name:
                                matching_children.append(child)
                        
                        if matching_children and 0 <= index < len(matching_children):
                            new_elements.append(matching_children[index])
                else:
                    # Buscar elementos normales
                    for elem in current_elements:
                        for child in elem:
                            if child.tag.endswith(part) or child.tag == part:
                                new_elements.append(child)
                
                current_elements = new_elements
            
            # Extraer valores de los elementos finales
            for elem in current_elements:
                if elem.text and elem.text.strip():
                    values.append(elem.text.strip())
        
        except Exception as e:
            print(f"Error en manual_xpath_search: {e}")
        
        return values
    
    def show_step6_view(self):
        """Vista del Paso 6: Definir filtros para BD"""
        self.clear_main_content()
        self.create_content_header("🎯 Paso 6: Definir Filtros para BD", "Los registros de la BD a comparar se filtrarán por estos campos")
        
        content_frame = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        content_frame.grid_rowconfigure(1, weight=1)  # main_scroll_frame expandible
        content_frame.grid_columnconfigure(0, weight=1)
        
        # Información del paso anterior y XRef IDs disponibles
        xref_ids_count = len(self.extracted_xref_ids) if hasattr(self, 'extracted_xref_ids') else 0
        info_label = ctk.CTkLabel(
            content_frame,
            text=f"✅ XRef IDs encontrados: {xref_ids_count} valores únicos\n🔍 Los XRef IDs están disponibles en el Paso 5 para su revisión",
            font=ctk.CTkFont(size=12),
            text_color="#10b981",
            justify="left"
        )
        info_label.grid(row=0, column=0, pady=(15, 20), sticky="w", padx=20)
        
        # Frame principal scrollable
        main_scroll_frame = ctk.CTkScrollableFrame(content_frame, fg_color="transparent")
        main_scroll_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        main_scroll_frame.grid_columnconfigure(0, weight=1)
        
        # Formulario para agregar filtros
        filter_form_frame = ctk.CTkFrame(main_scroll_frame, fg_color="#f8fafc", corner_radius=8)
        filter_form_frame.grid(row=0, column=0, sticky="ew", pady=(0, 20))
        filter_form_frame.grid_columnconfigure(1, weight=1)
        filter_form_frame.grid_columnconfigure(2, weight=1)
        filter_form_frame.grid_columnconfigure(3, weight=1)
        
        form_title = ctk.CTkLabel(
            filter_form_frame,
            text="� Agregar Nuevo Filtro",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#1e293b"
        )
        form_title.grid(row=0, column=0, columnspan=4, pady=(15, 20), padx=20)
        
        # Tabla
        table_label = ctk.CTkLabel(filter_form_frame, text="Tabla:", font=ctk.CTkFont(size=12, weight="bold"))
        table_label.grid(row=1, column=0, sticky="w", padx=20, pady=(0, 5))
        
        self.filter_table_entry = ctk.CTkComboBox(
            filter_form_frame,
            values=["dispatch", "nfirs_notification", "nfirs_notification_apparatus", "nfirs_notification_personel", "otro"],
            width=150,
            height=35,
            state="readonly"
        )
        self.filter_table_entry.grid(row=2, column=0, padx=20, pady=(0, 15), sticky="ew")
        self.filter_table_entry.set("dispatch")  # Valor por defecto
        
        # Nombre del Campo
        field_label = ctk.CTkLabel(filter_form_frame, text="Nombre del Campo:", font=ctk.CTkFont(size=12, weight="bold"))
        field_label.grid(row=1, column=1, sticky="w", padx=20, pady=(0, 5))
        
        self.filter_field_entry = ctk.CTkEntry(
            filter_form_frame,
            width=200,
            height=35,
            placeholder_text="ej: xref_id, created_at"
        )
        self.filter_field_entry.grid(row=2, column=1, padx=20, pady=(0, 15), sticky="ew")
        
        # Valor
        value_label = ctk.CTkLabel(filter_form_frame, text="Valor:", font=ctk.CTkFont(size=12, weight="bold"))
        value_label.grid(row=1, column=2, sticky="w", padx=20, pady=(0, 5))
        
        self.filter_value_combo = ctk.CTkComboBox(
            filter_form_frame,
            width=250,
            height=35,
            values=["Escribir valor..."],
            state="normal"
        )
        self.filter_value_combo.grid(row=2, column=2, padx=20, pady=(0, 15), sticky="ew")
        
        # Cuando el campo es xref_id, cargar los valores encontrados
        def on_field_change(*args):
            field_value = self.filter_field_entry.get().lower().strip()
            if field_value in ['xref_id', 'dispatch_number'] and hasattr(self, 'extracted_xref_ids') and self.extracted_xref_ids:
                # Cargar xref_ids encontrados
                xref_values = [item['value'] for item in self.extracted_xref_ids[:20]]  # Limitar a 20 para el ComboBox
                self.filter_value_combo.configure(values=xref_values + ["Escribir valor..."])
                self.filter_value_combo.set("Seleccionar XRef ID...")
            else:
                self.filter_value_combo.configure(values=["Escribir valor..."])
                self.filter_value_combo.set("")
        
        self.filter_field_entry.bind("<KeyRelease>", on_field_change)
        
        # Si ya hay XRef IDs extraídos y el campo es xref_id, cargarlos automáticamente
        if hasattr(self, 'extracted_xref_ids') and self.extracted_xref_ids:
            # Pre-configurar para xref_id si hay valores disponibles
            self.filter_field_entry.insert(0, "xref_id")
            xref_values = [item['value'] for item in self.extracted_xref_ids[:20]]
            self.filter_value_combo.configure(values=xref_values + ["Escribir valor..."])
            self.filter_value_combo.set("Seleccionar XRef ID...")
        
        # Botón agregar filtro
        add_filter_btn = ctk.CTkButton(
            filter_form_frame,
            text="➕ Agregar Filtro",
            width=140,
            height=35,
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self.add_filter
        )
        add_filter_btn.grid(row=3, column=0, columnspan=4, pady=(10, 20))
        
        # Lista de filtros configurados
        filters_list_frame = ctk.CTkFrame(main_scroll_frame, fg_color="#ffffff", corner_radius=8)
        filters_list_frame.grid(row=1, column=0, sticky="ew", pady=(20, 20))
        filters_list_frame.grid_columnconfigure(0, weight=1)
        
        self.filters_display_frame = ctk.CTkFrame(filters_list_frame, fg_color="transparent")
        self.filters_display_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 20))
        self.filters_display_frame.grid_columnconfigure(0, weight=1)
        
        # Actualizar display
        self.update_filters_display()
        
        # Frame para botones de acción
        action_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        action_frame.grid(row=3, column=0, pady=20, sticky="ew")
        action_frame.grid_columnconfigure(1, weight=1)
        
        # Botón volver
        back_btn = ctk.CTkButton(
            action_frame,
            text="⬅️ Volver al Paso 5",
            width=160,
            height=40,
            font=ctk.CTkFont(size=12),
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=lambda: self.handle_step_click("step5")
        )
        back_btn.grid(row=0, column=0, padx=(20, 10), sticky="w")
        
        # Botón continuar (solo si hay filtros manuales)
        if hasattr(self, 'filter_criteria') and len(self.filter_criteria) > 0:
            continue_btn = ctk.CTkButton(
                action_frame,
                text="➡️ Paso 7: Comparar",
                width=180,
                height=40,
                font=ctk.CTkFont(size=12, weight="bold"),
                command=lambda: self.handle_step_click("step7")
            )
            continue_btn.grid(row=0, column=2, padx=(10, 20), sticky="e")
    
    def add_filter(self):
        """Agregar un nuevo filtro"""
        table = self.filter_table_entry.get().strip()
        field = self.filter_field_entry.get().strip()
        value = self.filter_value_combo.get().strip()
        
        if not table or not field or not value or value in ["Escribir valor...", "Seleccionar XRef ID..."]:
            messagebox.showwarning("Advertencia", "Todos los campos son obligatorios")
            return
        
        # Verificar que no exista ya el mismo filtro
        for filter_item in self.filter_criteria:
            if (filter_item['table'] == table and 
                filter_item['field'] == field and 
                filter_item['value'] == value):
                messagebox.showwarning("Advertencia", f"El filtro '{table}.{field} = {value}' ya existe")
                return
        
        # Agregar el filtro
        self.filter_criteria.append({
            'table': table,
            'field': field,
            'value': value
        })
        
        # Si es el primer filtro agregado, actualizar progreso
        if len(self.filter_criteria) == 1:
            self.update_progress(6)  # Marcar Paso 6 como completado
        
        # Limpiar inputs
        self.filter_table_entry.set("dispatch")  # Resetear a valor por defecto
        self.filter_field_entry.delete(0, 'end')
        self.filter_value_combo.set("")
        
        # Actualizar displays
        self.update_filters_display()
        self.update_tracking_display()
        self.show_step6_view()  # Refrescar para mostrar botón continuar
        
        messagebox.showinfo("Éxito", f"Filtro '{table}.{field} = {value}' agregado correctamente")
    
    def remove_filter(self, index):
        """Eliminar un filtro"""
        if 0 <= index < len(self.filter_criteria):
            removed_filter = self.filter_criteria.pop(index)
            
            # Si no quedan filtros, actualizar progreso
            if len(self.filter_criteria) == 0:
                self.update_progress(5)  # Retroceder a paso 5
            
            self.update_filters_display()
            self.update_tracking_display()
            self.show_step6_view()  # Refrescar vista
            messagebox.showinfo("Filtro eliminado", f"Filtro '{removed_filter['table']}.{removed_filter['field']} = {removed_filter['value']}' eliminado")
    
    def update_filters_display(self):
        """Actualizar la visualización de filtros configurados"""
        # Limpiar contenido anterior
        for widget in self.filters_display_frame.winfo_children():
            widget.destroy()
        
        if not hasattr(self, 'filter_criteria') or not self.filter_criteria:
            # Mensaje cuando no hay filtros
            no_filters_label = ctk.CTkLabel(
                self.filters_display_frame,
                text="📝 No hay filtros configurados.\nAgrega al menos 1 filtro para continuar.",
                font=ctk.CTkFont(size=14),
                text_color="#6b7280"
            )
            no_filters_label.grid(row=0, column=0, pady=50)
            return
        
        # Título de la lista
        list_title = ctk.CTkLabel(
            self.filters_display_frame,
            text=f"🎯 Filtros Configurados ({len(self.filter_criteria)})",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#1f2937"
        )
        list_title.grid(row=0, column=0, pady=(10, 15), sticky="w")
        
        # Mostrar cada filtro
        for i, filter_item in enumerate(self.filter_criteria):
            filter_frame = ctk.CTkFrame(self.filters_display_frame, fg_color="#f1f5f9", corner_radius=6)
            filter_frame.grid(row=i+1, column=0, sticky="ew", pady=5)
            filter_frame.grid_columnconfigure(1, weight=1)
            
            # Contenido del filtro
            content_frame = ctk.CTkFrame(filter_frame, fg_color="transparent")
            content_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=15, pady=10)
            content_frame.grid_columnconfigure(1, weight=1)
            
            # Número del filtro
            number_label = ctk.CTkLabel(
                content_frame,
                text=f"{i+1}.",
                font=ctk.CTkFont(size=14, weight="bold"),
                text_color="#2563eb",
                width=35
            )
            number_label.grid(row=0, column=0, sticky="w")
            
            # Información del filtro
            filter_text = f"📋 {filter_item['table']}.{filter_item['field']} = {filter_item['value']}"
            filter_info = ctk.CTkLabel(
                content_frame,
                text=filter_text,
                font=ctk.CTkFont(size=13, weight="bold"),
                text_color="#059669"
            )
            filter_info.grid(row=0, column=1, sticky="w", padx=(10, 0))
            
            # Botón eliminar
            delete_btn = ctk.CTkButton(
                content_frame,
                text="🗑️",
                width=30,
                height=30,
                font=ctk.CTkFont(size=12),
                fg_color="#dc2626",
                hover_color="#b91c1c",
                command=lambda idx=i: self.remove_filter(idx)
            )
            delete_btn.grid(row=0, column=2, padx=(10, 0))
    
    def show_step7_view(self):
        """Vista del Paso 7: Comparar campos XML vs BD"""
        self.clear_main_content()
        self.create_content_header("⚖️ Paso 7: Comparar Campos", "Comparar campos XML vs Base de Datos")
        
        content_frame = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        
        # Resumen de configuración
        summary_frame = ctk.CTkFrame(content_frame, fg_color="#f0f9ff", corner_radius=8)
        summary_frame.pack(fill="x", padx=20, pady=20)
        
        summary_title = ctk.CTkLabel(
            summary_frame,
            text="📋 Resumen de Configuración",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#1d4ed8"
        )
        summary_title.pack(pady=(15, 10))
        
        # Información de campos objetivo
        fields_info = ctk.CTkLabel(
            summary_frame,
            text=f"🎯 Campos a comparar: {len(self.target_fields)} configurados",
            font=ctk.CTkFont(size=12),
            text_color="#374151"
        )
        fields_info.pack(pady=2)
        
        # Información de filtros
        filters_info = ctk.CTkLabel(
            summary_frame,
            text=f"🔍 Filtros BD: {len(self.filter_criteria)} configurados",
            font=ctk.CTkFont(size=12),
            text_color="#374151"
        )
        filters_info.pack(pady=2)
        
        # Información de XMLs
        xml_info = ctk.CTkLabel(
            summary_frame,
            text=f"📄 XRef IDs: {len(self.extracted_xref_ids) if hasattr(self, 'extracted_xref_ids') else 0} encontrados",
            font=ctk.CTkFont(size=12),
            text_color="#374151"
        )
        xml_info.pack(pady=(2, 15))
        
        # Instrucciones
        instruction_label = ctk.CTkLabel(
            content_frame,
            text="🚀 Todo está configurado correctamente.\nHaz clic en 'Iniciar Comparación' para procesar los datos.",
            font=ctk.CTkFont(size=14),
            text_color="#374151"
        )
        instruction_label.pack(pady=20)
        
        # Botón de ejecución
        execute_btn = ctk.CTkButton(
            content_frame,
            text="🚀 Iniciar Comparación",
            width=250,
            height=50,
            font=ctk.CTkFont(size=16, weight="bold"),
            fg_color="#10b981",
            hover_color="#059669",
            command=self.execute_comparison
        )
        execute_btn.pack(pady=20)
        
        # Frame para botones de navegación
        nav_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        nav_frame.pack(pady=20)
        
        # Botón volver
        back_btn = ctk.CTkButton(
            nav_frame,
            text="⬅️ Volver al Paso 6",
            width=160,
            height=40,
            font=ctk.CTkFont(size=12),
            fg_color="#6b7280",
            hover_color="#4b5563",
            command=lambda: self.handle_step_click("step6")
        )
        back_btn.pack(side="left", padx=5)
    
    def execute_comparison(self):
        """Ejecutar la comparación de campos XML vs BD"""
        try:
            # Validar que toda la configuración esté lista
            if not self.validate_comparison_config():
                return
            
            # Mostrar ventana de progreso
            progress_window = self.create_progress_window("Ejecutando Comparación...")
            
            # Ejecutar comparación en hilo separado para no bloquear UI
            import threading
            comparison_thread = threading.Thread(target=self.run_comparison_process, args=(progress_window,))
            comparison_thread.daemon = True
            comparison_thread.start()
            
        except Exception as e:
            self.logger.error(f"Error iniciando comparación: {str(e)}")
            messagebox.showerror("Error", f"Error iniciando la comparación: {str(e)}")
    
    def validate_comparison_config(self):
        """Validar que toda la configuración esté completa"""
        errors = []
        
        # Verificar campos objetivo
        if not hasattr(self, 'target_fields') or not self.target_fields:
            errors.append("• No hay campos objetivo configurados (Paso 4)")
        
        # Verificar carpeta XML y XRef IDs extraídos
        if not hasattr(self, 'xml_folder_path') or not self.xml_folder_path:
            errors.append("• No se ha seleccionado carpeta XML (Paso 5)")
        
        if not hasattr(self, 'extracted_xref_ids') or not self.extracted_xref_ids:
            errors.append("• No se han extraído XRef IDs de los XMLs (Paso 5)")
        
        # Verificar filtros BD
        if not hasattr(self, 'filter_criteria') or not self.filter_criteria:
            errors.append("• No hay filtros configurados para BD (Paso 6)")
        
        # Verificar XPath
        if not hasattr(self, 'xref_xpath') or not self.xref_xpath:
            errors.append("• No se ha extraído XPath del XRef ID (Paso 3)")
        
        if errors:
            error_msg = "Configuración incompleta:\n\n" + "\n".join(errors)
            messagebox.showerror("Error de Configuración", error_msg)
            return False
        
        return True
    
    def create_progress_window(self, title):
        """Crear ventana de progreso"""
        progress_window = ctk.CTkToplevel(self.root)
        progress_window.title(title)
        progress_window.geometry("500x300")
        progress_window.transient(self.root)
        progress_window.grab_set()
        
        # Centrar ventana
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - (500 // 2)
        y = (progress_window.winfo_screenheight() // 2) - (300 // 2)
        progress_window.geometry(f"+{x}+{y}")
        
        # Contenido de la ventana
        main_frame = ctk.CTkFrame(progress_window)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        title_label = ctk.CTkLabel(main_frame, text=title, font=ctk.CTkFont(size=16, weight="bold"))
        title_label.pack(pady=(20, 10))
        
        # Barra de progreso
        progress_bar = ctk.CTkProgressBar(main_frame, width=400, height=20)
        progress_bar.pack(pady=20)
        progress_bar.set(0)
        
        # Texto de estado
        status_label = ctk.CTkLabel(main_frame, text="Iniciando comparación...", wraplength=400)
        status_label.pack(pady=10)
        
        # Almacenar referencias en variables de instancia
        self._progress_widgets = {
            'window': progress_window,
            'progress_bar': progress_bar,
            'status_label': status_label
        }
        
        return progress_window
    
    def run_comparison_process(self, progress_window):
        """Ejecutar el proceso de comparación en hilo separado"""
        try:
            self.update_progress_window(progress_window, 0.1, "Conectando a base de datos...")
            
            # Verificar que tenemos configuración de BD
            if not self.db_config or not all(key in self.db_config for key in ['host', 'port', 'database', 'user', 'password']):
                raise Exception("Configuración de BD incompleta. Por favor configure la conexión a BD.")
            
            # Importar field_comparator con ruta absoluta
            import sys
            import os
            sys.path.append(os.path.dirname(__file__))
            from field_comparator import FieldComparator
            
            # Crear archivo temporal de configuración para FieldComparator
            temp_config = {
                "database": self.db_config
            }
            
            # Guardar configuración temporal
            temp_config_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'temp_db_config.json')
            with open(temp_config_path, 'w', encoding='utf-8') as f:
                json.dump(temp_config, f, indent=4, ensure_ascii=False)
            
            # FieldComparator espera el archivo de configuración
            comparator = FieldComparator(temp_config_path)
            
            self.update_progress_window(progress_window, 0.2, "Configuración cargada, conectando a BD...")
            
            # Conectar a BD con mejor manejo de errores
            try:
                if not comparator.connect_to_db():
                    raise Exception(f"Falló la conexión a BD: {self.db_config['host']}:{self.db_config['port']}/{self.db_config['database']}")
            except Exception as db_error:
                raise Exception(f"Error de conexión a BD: {str(db_error)}")
            finally:
                # Limpiar archivo temporal
                if os.path.exists(temp_config_path):
                    os.remove(temp_config_path)
            
            self.update_progress_window(progress_window, 0.3, "Procesando archivos XML...")
            
            # Obtener archivos XML de la carpeta
            if not self.xml_folder_path:
                raise Exception("No se ha seleccionado carpeta XML")
                
            xml_files = [f for f in os.listdir(self.xml_folder_path) if f.lower().endswith('.xml')]
            
            if not xml_files:
                raise Exception(f"No se encontraron archivos XML en: {self.xml_folder_path}")
            
            comparison_results = []
            total_files = len(xml_files)
            
            for i, xml_file in enumerate(xml_files):
                xml_path = os.path.join(self.xml_folder_path, xml_file)
                progress = 0.3 + (0.5 * (i + 1) / total_files)
                self.update_progress_window(progress_window, progress, f"Procesando {xml_file}...")
                
                # Procesar cada archivo XML
                file_results = self.process_xml_file(xml_path, xml_file, comparator)
                if file_results:
                    comparison_results.extend(file_results)
            
            self.update_progress_window(progress_window, 0.9, "Generando reporte...")
            
            # Generar reporte Excel
            report_path = self.generate_comparison_report(comparison_results)
            
            self.update_progress_window(progress_window, 1.0, "¡Comparación completada!")
            
            # Cerrar conexión BD
            comparator.close_db_connection()
            
            # Mostrar resultado
            self.root.after(1000, lambda: self.show_comparison_results(progress_window, report_path, len(comparison_results)))
            
            # Actualizar progreso
            self.update_progress(7)
            self.update_tracking_display()
            
        except Exception as e:
            error_msg = str(e)
            self.logger.error(f"Error en comparación: {error_msg}")
            self.root.after(100, lambda: self.show_comparison_error(progress_window, error_msg))
    
    def update_progress_window(self, window, progress, message):
        """Actualizar ventana de progreso desde hilo"""
        def update_ui():
            if (hasattr(self, '_progress_widgets') and 
                self._progress_widgets['window'] and 
                self._progress_widgets['window'].winfo_exists()):
                self._progress_widgets['progress_bar'].set(progress)
                self._progress_widgets['status_label'].configure(text=message)
                self._progress_widgets['window'].update()
        
        self.root.after(0, update_ui)
    
    def process_xml_file(self, xml_path, xml_filename, comparator):
        """Procesar un archivo XML individual"""
        try:
            # Extraer XRef ID del XML usando el XPath configurado
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            # Usar XPath para encontrar el XRef ID
            if not self.xref_xpath:
                print(f"No hay XPath configurado para XRef ID en {xml_filename}")
                return []
                
            xref_values = self.extract_values_from_xml(root, self.xref_xpath)
            
            if not xref_values:
                print(f"No se encontró XRef ID en {xml_filename} usando XPath: {self.xref_xpath}")
                return []
            
            xref_id = xref_values[0]  # Tomar el primer valor encontrado
            
            if not xref_id:
                print(f"XRef ID vacío en {xml_filename}")
                return []
            
            # Verificar si este XRef ID cumple con los filtros configurados
            if not self.xref_matches_filters(xref_id):
                print(f"XRef ID {xref_id} no cumple con los filtros configurados")
                return []
            
            # Buscar registro en BD
            db_record = self.find_db_record(xref_id, comparator)
            
            if not db_record:
                return [{
                    'xml_file': xml_filename,
                    'batt_dept_id': xref_id,
                    'target_field': 'REGISTRO_BD',
                    'xpath': 'N/A',
                    'valor_xml': 'N/A',
                    'valor_bd': 'NO_ENCONTRADO',
                    'coincide': False,
                    'observaciones': f'No se encontró registro en BD con filtros: {self.get_filters_display()}'
                }]
            
            # Comparar solo los campos objetivo definidos en el Paso 4
            file_results = []
            
            # Comparar cada campo objetivo
            for field_config in self.target_fields:
                field_result = self.compare_field(xml_path, db_record, field_config, xml_filename, xref_id)
                file_results.append(field_result)
            
            return file_results
            
        except Exception as e:
            print(f"Error procesando {xml_filename}: {str(e)}")
            return [{
                'xml_file': xml_filename,
                'batt_dept_id': 'ERROR',
                'target_field': 'PROCESAMIENTO',
                'xpath': 'N/A',
                'valor_xml': 'ERROR',
                'valor_bd': str(e),
                'coincide': False,
                'observaciones': f'Error de procesamiento: {str(e)}'
            }]
    
    def xref_matches_filters(self, xref_id):
        """Verificar si el XRef ID cumple con los filtros configurados"""
        # Por ahora, permitir todos los XRef IDs que estén en extracted_xref_ids
        # y que coincidan con los filtros de xref_id configurados
        
        for filter_item in self.filter_criteria:
            if filter_item['field'].lower() in ['xref_id', 'dispatch_number']:
                if filter_item['value'] == xref_id:
                    return True
        
        # Si no hay filtros específicos de xref_id, permitir todos los extraídos
        xref_filter_exists = any(f['field'].lower() in ['xref_id', 'dispatch_number'] for f in self.filter_criteria)
        if not xref_filter_exists:
            # Verificar que esté en los XRef IDs extraídos
            return any(item['value'] == xref_id for item in self.extracted_xref_ids)
        
        return False
    
    def find_db_record(self, xref_id, comparator):
        """Buscar registro en BD usando el XRef ID extraído del XML"""
        try:
            cursor = comparator.db_connection.cursor()
            
            # CORRECCIÓN: Buscar usando dispatch.dispatch_number que es STRING
            # No intentar convertir xref_id a entero, mantenerlo como string
            print(f"🔍 Buscando registro para XRef ID: {xref_id}")
            
            # Construir consulta base usando dispatch.dispatch_number
            base_query = """
                SELECT d.*, n.*, f.*, h.*, v.*, t.*, a.*
                FROM dispatch d
                LEFT JOIN nfirs_notification n ON d.dispatch_number = n.dispatch_number
                LEFT JOIN nfirs_fire_incident f ON d.dispatch_number = f.dispatch_number  
                LEFT JOIN nfirs_hazmat h ON d.dispatch_number = h.dispatch_number
                LEFT JOIN nfirs_vehicle v ON d.dispatch_number = v.dispatch_number
                LEFT JOIN nfirs_training t ON d.dispatch_number = t.dispatch_number
                LEFT JOIN nfirs_apparatus a ON d.dispatch_number = a.dispatch_number
                WHERE d.dispatch_number = %s
            """
            
            # Agregar filtros adicionales si están configurados
            additional_conditions = []
            params = [xref_id]  # El primer parámetro siempre es el xref_id
            
            if hasattr(self, 'filter_criteria') and self.filter_criteria:
                for filter_item in self.filter_criteria:
                    # Solo agregar filtros que NO sean para dispatch_number 
                    # (ya que ese es nuestro filtro principal)
                    if filter_item['field'].lower() not in ['dispatch_number', 'xref_id']:
                        table = filter_item['table']
                        field = filter_item['field']
                        value = filter_item['value']
                        
                        # Mapear tabla a alias
                        table_alias_map = {
                            'dispatch': 'd',
                            'nfirs_notification': 'n', 
                            'nfirs_fire_incident': 'f',
                            'nfirs_hazmat': 'h',
                            'nfirs_vehicle': 'v',
                            'nfirs_training': 't',
                            'nfirs_apparatus': 'a'
                        }
                        
                        alias = table_alias_map.get(table, table)
                        additional_conditions.append(f"{alias}.{field} = %s")
                        params.append(value)
            
            # Construir consulta final
            final_query = base_query
            if additional_conditions:
                final_query += " AND " + " AND ".join(additional_conditions)
            
            final_query += " LIMIT 1"
            
            print(f"🔍 Consulta SQL: {final_query}")
            print(f"📋 Parámetros: {params}")
            
            cursor.execute(final_query, params)
            result = cursor.fetchone()
            
            if result:
                # Obtener nombres de columnas y crear diccionario
                columns = [desc[0] for desc in cursor.description]
                record_dict = dict(zip(columns, result))
                print(f"✅ Registro encontrado para XRef {xref_id}")
                return record_dict
            else:
                print(f"❌ No se encontró registro para XRef {xref_id}")
                return None
                
        except Exception as e:
            print(f"❌ Error en búsqueda BD para XRef {xref_id}: {str(e)}")
            self.logger.error(f"Error en find_db_record: {str(e)}")
            return None
            
        except Exception as e:
            print(f"❌ Error buscando en BD para XRef ID {xref_id}: {str(e)}")
            return None
    
    def compare_field(self, xml_path, db_record, field_config, xml_filename, xref_id):
        """Comparar un campo específico entre XML y BD"""
        try:
            # Extraer valor del XML usando XPath correcto
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            # Usar el método extract_values_from_xml que ya maneja namespaces
            xml_values = self.extract_values_from_xml(root, field_config['xpath'])
            xml_value = xml_values[0] if xml_values and xml_values[0] else None
            
            # Obtener valor de BD del registro encontrado
            db_field = field_config['db_field']
            db_value = db_record.get(db_field) if db_record else None
            
            # Normalizar valores para comparación
            xml_str = str(xml_value).strip() if xml_value is not None else ""
            db_str = str(db_value).strip() if db_value is not None else ""
            
            # Comparar valores (exacta, case-sensitive)
            coincide_bool = (xml_str == db_str) and xml_str != "" and db_str != ""
            
            # Verificar errores
            if "ERROR" in str(xml_value) or "ERROR" in str(db_value):
                coincide = False
            else:
                coincide = coincide_bool
            
            # Generar observaciones detalladas
            observaciones = self._get_comparison_notes(xml_value, db_value)
            
            # Estructura del reporte según requerimientos del usuario
            return {
                'xml_file': xml_filename,           # Archivo XML
                'batt_dept_id': xref_id,            # XRef ID del XML  
                'target_field': db_field,           # Campo objetivo en BD
                'xpath': field_config['xpath'],     # XPath usado
                'valor_xml': xml_value,             # Valor del XML
                'valor_bd': db_value,               # Valor de la BD
                'coincide': coincide,               # TRUE/FALSE
                'observaciones': observaciones      # Notas detalladas
            }
            
        except Exception as e:
            # Estructura de error según requerimientos del usuario
            return {
                'xml_file': xml_filename,
                'batt_dept_id': xref_id,
                'target_field': field_config['db_field'],
                'xpath': field_config.get('xpath', 'N/A'),
                'valor_xml': f"ERROR: {str(e)}",
                'valor_bd': 'ERROR',
                'coincide': False,
                'observaciones': f"Error de procesamiento: {str(e)}"
            }
    
    def _get_comparison_notes(self, xml_value, db_value):
        """Generar notas sobre la comparación (igual que XML BD Comparator)"""
        # Verificar si hay errores
        if "ERROR" in str(xml_value):
            return "Error al obtener valor XML"
        elif "ERROR" in str(db_value):
            return "Error al obtener valor BD"
        
        # Verificar valores nulos
        xml_is_null = xml_value is None or str(xml_value).strip() == ""
        db_is_null = db_value is None or str(db_value).strip() == ""
        
        if xml_is_null and db_is_null:
            return "Ambos valores son nulos"
        elif xml_is_null:
            return "Valor XML es nulo"
        elif db_is_null:
            return "Valor BD es nulo"
        
        # Comparar valores no nulos
        xml_str = str(xml_value).strip()
        db_str = str(db_value).strip()
        
        if xml_str == db_str:
            return "Valores coinciden"
        else:
            return f"Valores diferentes: XML='{xml_str}' vs BD='{db_str}'"
    
    def get_filters_display(self):
        """Obtener texto de los filtros para mostrar en reporte"""
        if not hasattr(self, 'filter_criteria') or not self.filter_criteria:
            return "Sin filtros"
        
        filters_text = []
        for f in self.filter_criteria:
            filters_text.append(f"{f['table']}.{f['field']} = {f['value']}")
        
        return " | ".join(filters_text)
    
    def generate_comparison_report(self, comparison_results):
        """Generar reporte Excel con los resultados (formato XML BD Comparator)"""
        try:
            # Crear DataFrame
            df = pd.DataFrame(comparison_results)
            
            # Crear directorio de reportes si no existe
            reports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'reportes')
            os.makedirs(reports_dir, exist_ok=True)
            
            # Nombre del archivo
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"reporte_comparacion_campos_{timestamp}.xlsx"
            report_path = os.path.join(reports_dir, filename)
            
            # Crear archivo Excel con formato
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Resultados', index=False)
                
                # Crear resumen
                total_comparaciones = len(comparison_results)
                coincidencias = len([r for r in comparison_results if r.get('coincide', False) == True])
                errores = len([r for r in comparison_results if 'ERROR' in str(r.get('valor_xml', '')) or 'ERROR' in str(r.get('valor_bd', ''))])
                diferencias = total_comparaciones - coincidencias - errores
                
                resumen = pd.DataFrame({
                    'Métrica': ['Total de comparaciones', 'Coincidencias', 'Diferencias', 'Errores', 'Porcentaje de coincidencias'],
                    'Valor': [total_comparaciones, coincidencias, diferencias, errores, f"{(coincidencias/total_comparaciones*100):.2f}%" if total_comparaciones > 0 else "0%"]
                })
                resumen.to_excel(writer, sheet_name='Resumen', index=False)
                
                # Obtener workbook y worksheet para formatear
                workbook = writer.book
                worksheet = writer.sheets['Resultados']
                
                # Ajustar ancho de columnas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar formato de colores igual que XML BD Comparator
            self._apply_color_formatting(report_path, df)
            
            return report_path
            
        except Exception as e:
            print(f"Error generando reporte: {str(e)}")
            return None
    
    def _apply_color_formatting(self, report_path, df):
        """Aplicar formato de colores al reporte (igual que XML BD Comparator)"""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill
            
            workbook = openpyxl.load_workbook(report_path)
            worksheet = workbook['Resultados']
            
            # Definir colores iguales a XML BD Comparator
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Rojo claro
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Amarillo claro
            blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")   # Azul claro
            
            # Buscar la columna de "observaciones"
            observaciones_col = None
            for col_idx, cell in enumerate(worksheet[1], 1):  # Primera fila (encabezados)
                if cell.value and 'observacion' in str(cell.value).lower():
                    observaciones_col = col_idx
                    break
            
            if not observaciones_col:
                print("⚠️ No se encontró la columna 'observaciones' para aplicar colores")
                return
            
            # Aplicar colores según el contenido de observaciones
            for row_idx in range(2, len(df) + 2):  # Empezar desde la fila 2 (datos, no encabezados)
                cell = worksheet.cell(row=row_idx, column=observaciones_col)
                if cell.value:
                    observacion = str(cell.value).lower()
                    
                    if "coinciden" in observacion:
                        # Verde para coincidencias
                        cell.fill = green_fill
                    elif "error" in observacion:
                        # Amarillo para errores
                        cell.fill = yellow_fill
                    elif "diferentes" in observacion:
                        # Rojo para diferencias
                        cell.fill = red_fill
                    elif "nulo" in observacion or "nulos" in observacion:
                        # Azul para valores nulos
                        cell.fill = blue_fill
            
            # Guardar cambios
            workbook.save(report_path)
            print("✅ Formato de colores aplicado al reporte")
            
        except Exception as e:
            print(f"⚠️ Error aplicando colores al reporte: {str(e)}")
    
    def show_comparison_results(self, progress_window, report_path, total_comparisons):
        """Mostrar resultados de la comparación"""
        if progress_window and progress_window.winfo_exists():
            progress_window.destroy()
        
        result_msg = f"""✅ ¡Comparación completada exitosamente!

📊 Resultados:
• Total de comparaciones: {total_comparisons}
• Reporte generado: {os.path.basename(report_path)}

📁 Ubicación del reporte:
{report_path}

🎨 Código de colores en Excel:
• 🟢 Verde: Valores coinciden
• 🔴 Rojo: Valores no coinciden  
• 🔵 Azul: Error en comparación
• 🟡 Amarillo: Información de filtros

¿Deseas abrir el reporte ahora?"""
        
        result = messagebox.askyesno("Comparación Completada", result_msg)
        if result:
            try:
                os.startfile(report_path)  # Windows
            except:
                import subprocess
                subprocess.run(['open', report_path])  # macOS/Linux
    
    def show_comparison_error(self, progress_window, error_message):
        """Mostrar error de comparación"""
        if progress_window and progress_window.winfo_exists():
            progress_window.destroy()
        
        messagebox.showerror("Error en Comparación", f"Error durante la comparación:\n\n{error_message}")
    
    def run(self):
        """Ejecutar la aplicación"""
        self.root.mainloop()

def main():
    """Función principal"""
    try:
        app = ModernWebGUI()
        app.run()
    except Exception as e:
        print(f"Error iniciando la aplicación: {e}")
        messagebox.showerror("Error", f"Error iniciando la aplicación: {e}")

if __name__ == "__main__":
    main()
