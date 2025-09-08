import psycopg2
import json
import os
import logging
import xml.etree.ElementTree as ET
from datetime import datetime
import pandas as pd
from typing import Dict, List, Optional, Tuple, Any

class FieldComparator:
    def __init__(self, config_path: str):
        self.config_path = config_path
        self.config: Optional[Dict[str, Any]] = None
        self.db_connection = None
        self.logger = logging.getLogger(__name__)  # Inicializar logger aquí
        self.setup_logging()
        self.load_config()
        
    def setup_logging(self):
        """Configura el logging para el comparador"""
        # Crear directorio de logs si no existe
        log_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'logs')
        os.makedirs(log_dir, exist_ok=True)
        
        # Configurar logging
        log_filename = f"field_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        log_path = os.path.join(log_dir, log_filename)
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_path, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        # self.logger ya está inicializado en __init__
        
    def load_config(self):
        """Carga la configuración desde el archivo JSON"""
        try:
            with open(self.config_path, 'r', encoding='utf-8') as f:
                self.config = json.load(f)
            self.logger.info("Configuración cargada exitosamente")
        except Exception as e:
            self.logger.error(f"Error al cargar configuración: {str(e)}")
            raise
    
    def connect_to_db(self) -> bool:
        """Establece conexión con la base de datos PostgreSQL"""
        try:
            # Verificar si tenemos configuración de database directa o dentro de 'database'
            if 'database' in self.config:
                db_config = self.config['database']
            else:
                # Configuración directa (db_config.json)
                db_config = self.config
            
            # Usar connection_string si está disponible, sino construir la conexión
            if 'connection_string' in db_config and db_config['connection_string']:
                conn_str = db_config['connection_string']
            else:
                # Manejar diferentes nombres de campos para compatibilidad
                username = db_config.get('username', db_config.get('user', ''))
                password = db_config.get('password', '')
                host = db_config.get('server', db_config.get('host', ''))
                port = db_config.get('port', 5432)
                database = db_config.get('database', '')
                
                conn_str = (
                    f"postgresql://{username}:{password}"
                    f"@{host}:{port}/{database}"
                )
            
            self.db_connection = psycopg2.connect(conn_str)
            self.logger.info("Conexión a base de datos PostgreSQL establecida")
            return True
            
        except Exception as e:
            self.logger.error(f"Error al conectar con PostgreSQL: {str(e)}")
            return False
    
    def close_db_connection(self):
        """Cierra la conexión con la base de datos"""
        if self.db_connection:
            self.db_connection.close()
            self.logger.info("Conexión a base de datos cerrada")
    
    def load_sql_query(self) -> str:
        """Carga la consulta SQL desde el archivo"""
        try:
            sql_file_path = os.path.join(
                os.path.dirname(os.path.dirname(__file__)), 
                'sql', 
                self.config['comparison']['sql_query_file']
            )
            
            with open(sql_file_path, 'r', encoding='utf-8') as f:
                query = f.read()
            
            self.logger.info(f"Consulta SQL cargada desde: {sql_file_path}")
            return query
            
        except Exception as e:
            self.logger.error(f"Error al cargar consulta SQL: {str(e)}")
            raise
    
    def get_database_records(self) -> pd.DataFrame:
        """Obtiene los registros de la base de datos según la consulta configurada"""
        try:
            self.logger.info("=== OBTENIENDO REGISTROS DE BD ===")
            
            query = self.load_sql_query()
            self.logger.info(f"Consulta SQL:")
            self.logger.info(query)
            
            df = pd.read_sql(query, self.db_connection)
            self.logger.info(f"Registros obtenidos: {len(df)}")
            
            if not df.empty:
                self.logger.info("Columnas obtenidas:")
                for col in df.columns:
                    self.logger.info(f"  - {col}")
                
                self.logger.info("Primeros registros:")
                for i, (_, row) in enumerate(df.head(3).iterrows()):
                    self.logger.info(f"  Registro {i+1}: {dict(row)}")
            else:
                self.logger.warning("¡No se obtuvieron registros de la BD!")
                
            return df
            
        except Exception as e:
            self.logger.error(f"Error al obtener registros de PostgreSQL: {str(e)}")
            raise
    
    def extract_xml_value(self, xml_path: str, xpath: str) -> Optional[str]:
        """Extrae un valor específico de un archivo XML usando XPath con fallback automático"""
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            self.logger.debug(f"XML Root tag: {root.tag}")
            self.logger.debug(f"Buscando con XPath: {xpath}")
            
            # Verificar si hay concatenación con '+'
            if ' + ' in xpath:
                self.logger.info(f"Detectada concatenación en XPath: {xpath}")
                parts = xpath.split(' + ')
                values = []
                
                for part in parts:
                    part = part.strip()
                    self.logger.debug(f"Evaluando parte: {part}")
                    
                    # Extraer valor para cada parte del XPath
                    part_value = self._process_xpath_with_fallback(root, part)
                    if part_value:
                        values.append(part_value)
                        self.logger.debug(f"Valor encontrado en '{part}': '{part_value}'")
                
                # Concatenar todos los valores encontrados con " - "
                if values:
                    concatenated = " - ".join(values)
                    self.logger.info(f"Valor concatenado final: '{concatenated}'")
                    return concatenated
                else:
                    self.logger.warning(f"No se encontraron valores para ninguna parte de: {xpath}")
                    return ""
            
            # Verificar si hay operador OR (|) - procesar con fallback
            elif ' | ' in xpath:
                self.logger.info(f"Detectado operador OR en XPath: {xpath}")
                return self._process_xpath_with_fallback(root, xpath)
            
            else:
                # XPath simple
                return self._process_xpath_with_fallback(root, xpath)
                    
        except Exception as e:
            self.logger.error(f"Error al extraer valor del XML: {str(e)}")
            return ""
    
    def _process_xpath_with_fallback(self, root, xpath: str) -> str:
        """Procesa XPath con lógica de fallback automática"""
        try:
            # Si hay operador OR, dividir y procesar cada alternativa
            if ' | ' in xpath:
                alternatives = xpath.split(' | ')
                
                # Buscar el XPath principal (primero)
                primary_xpath = alternatives[0].strip()
                
                # Extraer XPath base removiendo condiciones
                clean_primary = self._clean_xpath_conditions(primary_xpath)
                self.logger.debug(f"XPath principal limpio: {clean_primary}")
                
                # Intentar encontrar valor en el XPath principal
                value = self._extract_from_xml(root, clean_primary)
                if value and value.strip():
                    self.logger.info(f"Valor encontrado en XPath principal '{clean_primary}': '{value}'")
                    return value
                
                # Si no encuentra en el principal, buscar en las alternativas
                for i, alternative in enumerate(alternatives[1:], 1):
                    alternative = alternative.strip()
                    clean_alternative = self._clean_xpath_conditions(alternative)
                    self.logger.debug(f"Probando alternativa {i}: {clean_alternative}")
                    
                    value = self._extract_from_xml(root, clean_alternative)
                    if value and value.strip():
                        self.logger.info(f"Valor encontrado en alternativa '{clean_alternative}': '{value}'")
                        return value
                
                self.logger.warning(f"No se encontraron valores en ninguna alternativa")
                return ""
            
            else:
                # XPath simple sin OR
                clean_xpath = self._clean_xpath_conditions(xpath)
                return self._extract_from_xml(root, clean_xpath)
                
        except Exception as e:
            self.logger.error(f"Error procesando XPath con fallback: {str(e)}")
            return ""
    
    def _clean_xpath_conditions(self, xpath: str) -> str:
        """Limpia condiciones complejas del XPath para hacer búsqueda simple"""
        xpath = xpath.strip()
        
        # Remover condiciones entre corchetes
        if '[' in xpath and ']' in xpath:
            xpath = xpath.split('[')[0]
        
        # Remover /text() si existe
        if xpath.endswith('/text()'):
            xpath = xpath[:-7]
            
        self.logger.debug(f"XPath limpio: {xpath}")
        return xpath
    
    def _extract_from_xml(self, root, xpath: str) -> str:
        """Extrae valor del XML usando múltiples estrategias de búsqueda"""
        try:
            xpath = xpath.strip()
            self.logger.debug(f"Extrayendo valor de XPath: {xpath}")
            
            # Estrategia 1: Búsqueda directa por tag name (más simple y efectiva)
            if xpath.startswith('//'):
                tag_name = xpath[2:].split('[')[0]  # Extraer solo el nombre del tag
                self.logger.debug(f"Buscando directamente el tag: {tag_name}")
                
                elements = root.findall(f".//{tag_name}")
                self.logger.debug(f"Encontrados {len(elements)} elementos para tag '{tag_name}'")
                
                for i, element in enumerate(elements):
                    self.logger.debug(f"Elemento {i}: tag='{element.tag}', text='{element.text}', attrib={element.attrib}")
                    if element.text and element.text.strip():
                        value = element.text.strip()
                        self.logger.info(f"Valor encontrado en '{tag_name}': '{value}'")
                        return value
            
            # Estrategia 2: Búsqueda con path completo 
            if xpath.startswith('//'):
                search_path = xpath[2:].split('[')[0]  # Remover // y condiciones
                elements = root.findall(f".//{search_path}")
                self.logger.debug(f"Búsqueda con path completo './/{search_path}': {len(elements)} elementos")
                
                for element in elements:
                    if element.text and element.text.strip():
                        value = element.text.strip()
                        self.logger.debug(f"Valor encontrado con path completo: '{value}'")
                        return value
            
            # Estrategia 3: Búsqueda recursiva manual
            def find_element_recursive(elem, target_tag):
                """Busca un elemento recursivamente en todo el árbol"""
                if elem.tag == target_tag or elem.tag.endswith(f"}}{target_tag}"):
                    return elem
                
                for child in elem:
                    result = find_element_recursive(child, target_tag)
                    if result is not None:
                        return result
                return None
            
            if xpath.startswith('//'):
                target_tag = xpath[2:].split('[')[0].split('/')[-1]
                self.logger.debug(f"Búsqueda recursiva manual para tag: {target_tag}")
                
                found_element = find_element_recursive(root, target_tag)
                if found_element is not None and found_element.text and found_element.text.strip():
                    value = found_element.text.strip()
                    self.logger.info(f"Valor encontrado con búsqueda recursiva: '{value}'")
                    return value
            
            # Estrategia 4: Imprimir estructura para debug
            self.logger.debug("No se encontró valor, imprimiendo estructura del XML:")
            self._debug_xml_structure(root, max_depth=3)
            
            self.logger.warning(f"No se encontró valor para {xpath} con ninguna estrategia")
            return ""
            
        except Exception as e:
            self.logger.error(f"Error extrayendo de XML: {str(e)}")
            return ""
    
    def _debug_xml_structure(self, element, depth=0, max_depth=3):
        """Imprime la estructura del XML para debugging"""
        if depth > max_depth:
            return
            
        indent = "  " * depth
        text_preview = ""
        if element.text and element.text.strip():
            text_preview = f" -> '{element.text.strip()[:50]}'"
        
        self.logger.debug(f"{indent}{element.tag}{text_preview}")
        
        for child in element:
            self._debug_xml_structure(child, depth + 1, max_depth)
    
    def analyze_xml_structure(self, xml_path: str, record_id: str) -> None:
        """Analiza la estructura del XML para ayudar a encontrar XPaths correctos"""
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            
            self.logger.info(f"=== ANÁLISIS DE ESTRUCTURA XML ===")
            self.logger.info(f"Archivo: {os.path.basename(xml_path)}")
            self.logger.info(f"Root tag: {root.tag}")
            self.logger.info(f"Root attributes: {root.attrib}")
            
            # Buscar elementos que contengan el record_id
            def search_for_id(element, path=""):
                current_path = f"{path}/{element.tag}" if path else element.tag
                
                # Verificar texto del elemento
                if element.text and str(record_id) in str(element.text):
                    self.logger.info(f"ID encontrado en texto: {current_path} = '{element.text}'")
                
                # Verificar atributos del elemento
                for attr, value in element.attrib.items():
                    if str(record_id) in str(value):
                        self.logger.info(f"ID encontrado en atributo: {current_path}/@{attr} = '{value}'")
                
                # Buscar recursivamente en hijos
                for child in element:
                    search_for_id(child, current_path)
            
            search_for_id(root)
            
            # Mostrar algunos elementos principales
            self.logger.info(f"Elementos principales:")
            for child in root:
                self.logger.info(f"  - {child.tag}: {child.text[:50] if child.text else 'No text'}...")
                for subchild in child:
                    self.logger.info(f"    - {subchild.tag}: {subchild.text[:50] if subchild.text else 'No text'}...")
            
        except Exception as e:
            self.logger.error(f"Error al analizar estructura XML: {str(e)}")
    
    def find_xml_for_record(self, xml_folder: str, record_id: str) -> Optional[str]:
        """Busca el archivo XML correspondiente a un registro específico"""
        try:
            # Listar todos los archivos XML en la carpeta
            xml_files = [f for f in os.listdir(xml_folder) if f.lower().endswith('.xml')]
            self.logger.info(f"Buscando XML para registro {record_id}")
            self.logger.info(f"Archivos XML encontrados en carpeta: {len(xml_files)}")
            
            for xml_file in xml_files:
                xml_path = os.path.join(xml_folder, xml_file)
                self.logger.info(f"Procesando archivo XML: {xml_file}")
                
                # Extraer el identificador del XML
                id_xpath = self.config['xml']['identificador_xpath']
                self.logger.info(f"Usando XPath identificador: {id_xpath}")
                
                xml_id = self.extract_xml_value(xml_path, id_xpath)
                self.logger.info(f"ID extraído del XML: '{xml_id}' vs Record ID: '{record_id}'")
                
                if xml_id and xml_id.strip() == str(record_id).strip():
                    self.logger.info(f"¡Coincidencia encontrada! XML: {xml_file}")
                    return xml_path
                else:
                    self.logger.info(f"No hay coincidencia. XML_ID='{xml_id}', Record_ID='{record_id}'")
            
            # Si no se encontró coincidencia, analizar el primer XML para ayudar con debugging
            if xml_files:
                self.logger.info("No se encontraron coincidencias. Analizando primer XML para debug:")
                self.analyze_xml_structure(os.path.join(xml_folder, xml_files[0]), record_id)
            
            self.logger.warning(f"No se encontró XML para el registro {record_id}")
            return None
            
        except Exception as e:
            self.logger.error(f"Error al buscar XML para registro {record_id}: {str(e)}")
            return None
    
    def compare_fields(self, xml_folder: str) -> List[Dict]:
        """Compara los campos de la BD con los valores del XML"""
        try:
            # Obtener registros de la base de datos
            db_records = self.get_database_records()
            
            # Configuración de comparación
            comparison_config = self.config['comparison']
            xml_config = self.config['xml']
            
            campos_comparar = comparison_config['campos_comparar']
            campo_id = comparison_config['campo_identificador']
            xpath_mappings = xml_config['xpath_mappings']
            
            results = []
            
            for _, record in db_records.iterrows():
                record_id = record[campo_id]
                
                # Buscar el XML correspondiente
                xml_path = self.find_xml_for_record(xml_folder, record_id)
                
                result = {
                    'registro_id': record_id,
                    'xml_encontrado': xml_path is not None,
                    'xml_archivo': os.path.basename(xml_path) if xml_path else 'No encontrado',
                    'coincidencias': {},
                    'diferencias': {},
                    'errores': []
                }
                
                if xml_path:
                    # Comparar cada campo configurado
                    for campo in campos_comparar:
                        if campo in xpath_mappings:
                            db_value = str(record[campo]).strip() if pd.notna(record[campo]) else ""
                            xml_value = self.extract_xml_value(xml_path, xpath_mappings[campo])
                            xml_value = xml_value.strip() if xml_value else ""
                            
                            if db_value == xml_value:
                                result['coincidencias'][campo] = {
                                    'bd_value': db_value,
                                    'xml_value': xml_value
                                }
                            else:
                                result['diferencias'][campo] = {
                                    'bd_value': db_value,
                                    'xml_value': xml_value
                                }
                        else:
                            result['errores'].append(f"No se encontró XPath para el campo {campo}")
                else:
                    result['errores'].append("Archivo XML no encontrado para este registro")
                
                results.append(result)
                
                # Log del progreso
                if len(results) % 10 == 0:
                    self.logger.info(f"Procesados {len(results)} registros...")
            
            self.logger.info(f"Comparación completada. Total de registros procesados: {len(results)}")
            return results
            
        except Exception as e:
            self.logger.error(f"Error durante la comparación: {str(e)}")
            raise
    
    def extract_xml_id(self, xml_file: str) -> Optional[str]:
        """Extrae el ID del archivo XML usando el XPath configurado"""
        try:
            self.logger.info(f"=== EXTRAYENDO ID DEL XML ===")
            self.logger.info(f"Archivo: {os.path.basename(xml_file)}")
            
            id_xpath = self.config['xml']['identificador_xpath']
            self.logger.info(f"XPath configurado: {id_xpath}")
            
            # Analizar la estructura del XML primero
            tree = ET.parse(xml_file)
            root = tree.getroot()
            self.logger.info(f"Root tag: {root.tag}")
            
            # Intentar extraer el ID
            xml_id = self.extract_xml_value(xml_file, id_xpath)
            self.logger.info(f"ID extraído: '{xml_id}'")
            
            # Si no encuentra el ID, intentar buscar elementos que contengan números
            if not xml_id:
                self.logger.warning("ID no encontrado con XPath, buscando alternativas...")
                
                # Buscar elementos que contengan números (posibles IDs)
                def search_numbers(element, path=""):
                    current_path = f"{path}/{element.tag}" if path else element.tag
                    
                    if element.text and element.text.strip().isdigit():
                        self.logger.info(f"Número encontrado en {current_path}: '{element.text}'")
                    
                    # Buscar en hijos
                    for child in element:
                        search_numbers(child, current_path)
                
                search_numbers(root)
                
                # Intentar con el nombre del archivo como último recurso
                filename = os.path.basename(xml_file)
                self.logger.info(f"Nombre del archivo: {filename}")
                
                # Buscar números en el nombre del archivo
                import re
                numbers_in_filename = re.findall(r'\d+', filename)
                if numbers_in_filename:
                    self.logger.info(f"Números en nombre del archivo: {numbers_in_filename}")
            
            return xml_id
            
        except Exception as e:
            self.logger.error(f"Error al extraer ID del XML: {str(e)}")
            return None
    
    def compare_with_single_file(self, xml_file: str) -> List[Dict]:
        """Compara los datos de la BD con un archivo XML específico"""
        try:
            self.logger.info(f"Comparando con archivo: {xml_file}")
            
            # Verificar que el archivo existe
            if not os.path.exists(xml_file):
                raise FileNotFoundError(f"El archivo XML no existe: {xml_file}")
            
            # Obtener datos de la base de datos
            db_data = self.get_database_records()
            
            if db_data.empty:
                self.logger.warning("No se encontraron registros en la base de datos")
                return []
            
            self.logger.info(f"Procesando {len(db_data)} registros de la BD")
            
            results = []
            
            # Como no usamos identificador, comparamos TODOS los registros de BD con el XML
            self.logger.info("=== COMPARANDO TODOS LOS REGISTROS DE BD CON XML ===")
            
            for index, (_, record) in enumerate(db_data.iterrows()):
                # Obtener campos de filtro para logging
                campos_filtro = self.config['comparison'].get('campos_filtro_bd', [])
                record_identifier_parts = []
                
                for campo_filtro in campos_filtro:
                    if campo_filtro in record:
                        valor = record[campo_filtro]
                        record_identifier_parts.append(f"{campo_filtro}={valor}")
                
                # Crear identificador para logging
                if record_identifier_parts:
                    record_identifier = " | ".join(record_identifier_parts)
                else:
                    record_id = record[self.config['comparison']['campo_identificador']]
                    record_identifier = f"ID={record_id}"
                
                self.logger.info(f"Procesando registro {index + 1}/{len(db_data)}: {record_identifier}")
                
                # Comparar cada campo configurado
                for campo in self.config['comparison']['campos_comparar']:
                    bd_value = record.get(campo, '')
                    
                    # Obtener XPath para este campo
                    xpath = self.config['xml']['xpath_mappings'].get(campo, '')
                    
                    if xpath:
                        self.logger.info(f"Extrayendo campo '{campo}' con XPath: {xpath}")
                        xml_value = self.extract_xml_value(xml_file, xpath)
                        self.logger.info(f"Valor extraído para '{campo}': '{xml_value}'")
                    else:
                        self.logger.warning(f"No se encontró XPath para el campo: {campo}")
                        xml_value = "XPATH_NO_CONFIGURADO"
                    
                    # Comparar valores
                    bd_str = str(bd_value).strip() if bd_value is not None else ""
                    xml_str = str(xml_value).strip() if xml_value is not None else ""
                    
                    # Determinar el tipo de coincidencia
                    if bd_str == xml_str:
                        if bd_str == "" and xml_str == "":
                            coincide = "EMPTY_MATCH"  # Ambos vacíos
                        else:
                            coincide = "TRUE"  # Coinciden con valores
                    else:
                        coincide = "FALSE"  # No coinciden
                    
                    # Crear resultado con columnas separadas para campos de filtro
                    result = {
                        'campo': campo,
                        'valor_bd': bd_value,
                        'valor_xml': xml_value,
                        'coincide': coincide,
                        'archivo_xml': os.path.basename(xml_file)
                    }
                    
                    # Agregar cada campo de filtro como columna separada
                    for campo_filtro in campos_filtro:
                        if campo_filtro in record:
                            result[campo_filtro] = record[campo_filtro]
                        else:
                            result[campo_filtro] = "N/A"
                    
                    # Si no hay campos de filtro, agregar el campo identificador principal
                    if not campos_filtro:
                        campo_id = self.config['comparison']['campo_identificador']
                        result[campo_id] = record[campo_id]
                    
                    results.append(result)
                    
                    self.logger.debug(f"Campo {campo}: BD='{bd_value}' | XML='{xml_value}' | Coincide: {coincide}")
            
            self.logger.info(f"Comparación completada. Resultados: {len(results)}")
            return results
            
        except Exception as e:
            self.logger.error(f"Error en comparación con archivo único: {str(e)}")
            raise
    
    def generate_simple_report(self, comparison_results: List[Dict]) -> str:
        """Genera un reporte en Excel con los resultados de la comparación simplificada"""
        try:
            self.logger.info("Generando reporte Excel...")
            
            # Crear DataFrame directamente con los resultados
            df = pd.DataFrame(comparison_results)
            
            if df.empty:
                raise Exception("No hay resultados para generar el reporte")
            
            # Reorganizar columnas: campos de filtro primero, luego campos de comparación
            campos_filtro = self.config['comparison'].get('campos_filtro_bd', [])
            
            # Definir el orden de las columnas
            columnas_ordenadas = []
            
            # 1. Campos de filtro de BD al inicio
            for campo in campos_filtro:
                if campo in df.columns:
                    columnas_ordenadas.append(campo)
            
            # 2. Si no hay campos de filtro, usar campo identificador
            if not campos_filtro:
                campo_id = self.config['comparison']['campo_identificador']
                if campo_id in df.columns:
                    columnas_ordenadas.append(campo_id)
            
            # 3. Campos de comparación
            columnas_comparacion = ['campo', 'valor_bd', 'valor_xml', 'coincide']
            for col in columnas_comparacion:
                if col in df.columns:
                    columnas_ordenadas.append(col)
            
            # 4. Archivo XML al final
            if 'archivo_xml' in df.columns:
                columnas_ordenadas.append('archivo_xml')
            
            # Reordenar DataFrame
            df = df[columnas_ordenadas]
            
            # Crear nombre del archivo con timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"reporte_comparacion_campos_{timestamp}.xlsx"
            
            # Crear la ruta completa
            reports_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'reportes')
            os.makedirs(reports_dir, exist_ok=True)
            
            report_path = os.path.join(reports_dir, filename)
            
            # Escribir a Excel con formato
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                # Hoja principal con resultados detallados
                df.to_excel(writer, sheet_name='Resultados Detallados', index=False)
                
                # Hoja de resumen
                # Obtener campos de filtro para el resumen
                campos_filtro = self.config['comparison'].get('campos_filtro_bd', [])
                
                # Crear información del primer registro como ejemplo
                registro_info = "N/A"
                if not df.empty:
                    info_parts = []
                    for campo in campos_filtro:
                        if campo in df.columns:
                            valor = df[campo].iloc[0]
                            info_parts.append(f"{campo}={valor}")
                    if info_parts:
                        registro_info = " | ".join(info_parts)
                    else:
                        # Fallback al campo identificador si no hay campos de filtro
                        campo_id = self.config['comparison']['campo_identificador']
                        if campo_id in df.columns:
                            registro_info = f"{campo_id}={df[campo_id].iloc[0]}"
                
                summary_data = {
                    'Métrica': [
                        'Total de campos comparados',
                        'Campos que coinciden con datos',
                        'Campos que difieren',
                        'Campos vacíos en ambos',
                        'Porcentaje de coincidencia total',
                        'Archivo XML analizado',
                        'Registros procesados',
                        'Ejemplo de registro'
                    ],
                    'Valor': [
                        len(df),
                        len(df[df['coincide'] == "TRUE"]),
                        len(df[df['coincide'] == "FALSE"]),
                        len(df[df['coincide'] == "EMPTY_MATCH"]),
                        f"{((len(df[df['coincide'] == 'TRUE']) + len(df[df['coincide'] == 'EMPTY_MATCH'])) / len(df) * 100):.1f}%" if len(df) > 0 else "0%",
                        df['archivo_xml'].iloc[0] if not df.empty else "N/A",
                        len(df) // len(self.config['comparison']['campos_comparar']) if len(self.config['comparison']['campos_comparar']) > 0 else 0,
                        registro_info
                    ]
                }
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Resumen', index=False)
                
                # Formatear las hojas
                workbook = writer.book
                
                # Formatear hoja de resultados
                worksheet = writer.sheets['Resultados Detallados']
                
                # Autoajustar columnas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Colorear filas según coincidencia
                from openpyxl.styles import PatternFill
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                blue_fill = PatternFill(start_color="CCDDFF", end_color="CCDDFF", fill_type="solid")
                
                # Encontrar la columna de 'coincide' dinámicamente
                coincide_col = None
                for col_idx, col_name in enumerate(df.columns, 1):
                    if col_name == 'coincide':
                        coincide_col = col_idx
                        break
                
                if coincide_col:
                    for row in range(2, worksheet.max_row + 1):  # Empezar desde fila 2 (después del header)
                        coincide_cell = worksheet.cell(row=row, column=coincide_col)
                        if coincide_cell.value == "TRUE":
                            for col in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row, column=col).fill = green_fill
                        elif coincide_cell.value == "FALSE":
                            for col in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row, column=col).fill = red_fill
                        elif coincide_cell.value == "EMPTY_MATCH":
                            for col in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row, column=col).fill = blue_fill
                
                # Crear hoja de leyenda
                leyenda_data = {
                    'Elemento': [
                        'Fila VERDE',
                        'Fila ROJA',
                        'Fila AZUL',
                        'Campos de filtro',
                        'Campo',
                        'Valor BD',
                        'Valor XML',
                        'Coincide'
                    ],
                    'Significado': [
                        'Los valores de BD y XML COINCIDEN con datos',
                        'Los valores de BD y XML NO COINCIDEN',
                        'Los valores de BD y XML están VACÍOS (ambos sin datos)',
                        'Campos usados para filtrar registros en BD (dispatch_number, created_at, batt_dept_id)',
                        'Campo que se está comparando entre BD y XML',
                        'Valor del campo en la base de datos',
                        'Valor extraído del archivo XML',
                        'TRUE = coinciden con datos, FALSE = no coinciden, EMPTY_MATCH = ambos vacíos'
                    ]
                }
                
                leyenda_df = pd.DataFrame(leyenda_data)
                leyenda_df.to_excel(writer, sheet_name='Leyenda', index=False)
                
                # Aplicar formato a la hoja de leyenda
                leyenda_worksheet = writer.sheets['Leyenda']
                
                # Aplicar colores de ejemplo
                from openpyxl.styles import PatternFill, Font
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                bold_font = Font(bold=True)
                
                # Colorear las filas de ejemplo
                for col in range(1, 3):  # Columnas A y B
                    leyenda_worksheet.cell(row=2, column=col).fill = green_fill  # Fila VERDE
                    leyenda_worksheet.cell(row=3, column=col).fill = red_fill   # Fila ROJA
                
                # Hacer los headers en negrita
                for col in range(1, 3):
                    leyenda_worksheet.cell(row=1, column=col).font = bold_font
                
                # Autoajustar columnas de leyenda
                for column in leyenda_worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 80)
                    leyenda_worksheet.column_dimensions[column_letter].width = adjusted_width
            
            self.logger.info(f"Reporte generado: {report_path}")
            return report_path
            
        except Exception as e:
            self.logger.error(f"Error al generar reporte: {str(e)}")
            raise
    
    def generate_report(self, comparison_results: List[Dict]) -> str:
        """Genera un reporte en Excel con los resultados de la comparación"""
        try:
            # Crear dataframes para diferentes hojas del reporte
            summary_data = []
            details_data = []
            errors_data = []
            
            total_records = len(comparison_results)
            xmls_found = sum(1 for r in comparison_results if r['xml_encontrado'])
            total_coincidencias = 0
            total_diferencias = 0
            
            for result in comparison_results:
                record_id = result['registro_id']
                xml_file = result['xml_archivo']
                xml_found = result['xml_encontrado']
                
                coincidencias = len(result['coincidencias'])
                diferencias = len(result['diferencias'])
                errores = len(result['errores'])
                
                total_coincidencias += coincidencias
                total_diferencias += diferencias
                
                # Datos del resumen
                summary_data.append({
                    'Registro_ID': record_id,
                    'XML_Archivo': xml_file,
                    'XML_Encontrado': 'Sí' if xml_found else 'No',
                    'Coincidencias': coincidencias,
                    'Diferencias': diferencias,
                    'Errores': errores
                })
                
                # Datos de detalles - coincidencias
                for campo, valores in result['coincidencias'].items():
                    details_data.append({
                        'Registro_ID': record_id,
                        'Campo': campo,
                        'Tipo': 'Coincidencia',
                        'Valor_BD': valores['bd_value'],
                        'Valor_XML': valores['xml_value'],
                        'XML_Archivo': xml_file
                    })
                
                # Datos de detalles - diferencias
                for campo, valores in result['diferencias'].items():
                    details_data.append({
                        'Registro_ID': record_id,
                        'Campo': campo,
                        'Tipo': 'Diferencia',
                        'Valor_BD': valores['bd_value'],
                        'Valor_XML': valores['xml_value'],
                        'XML_Archivo': xml_file
                    })
                
                # Datos de errores
                for error in result['errores']:
                    errors_data.append({
                        'Registro_ID': record_id,
                        'XML_Archivo': xml_file,
                        'Error': error
                    })
            
            # Crear DataFrames
            df_summary = pd.DataFrame(summary_data)
            df_details = pd.DataFrame(details_data)
            df_errors = pd.DataFrame(errors_data)
            
            # Crear estadísticas generales
            stats_data = [
                ['Total de registros procesados', total_records],
                ['XMLs encontrados', xmls_found],
                ['XMLs no encontrados', total_records - xmls_found],
                ['Total de coincidencias', total_coincidencias],
                ['Total de diferencias', total_diferencias],
                ['Porcentaje de XMLs encontrados', f"{(xmls_found/total_records)*100:.2f}%" if total_records > 0 else "0%"],
                ['Fecha de procesamiento', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            ]
            df_stats = pd.DataFrame(stats_data, columns=['Métrica', 'Valor'])
            
            # Generar archivo de reporte
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_filename = f"reporte_comparacion_campos_{timestamp}.xlsx"
            report_path = os.path.join(
                os.path.dirname(os.path.dirname(__file__)), 
                'reportes', 
                report_filename
            )
            
            # Escribir a Excel
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                df_stats.to_excel(writer, sheet_name='Estadísticas', index=False)
                df_summary.to_excel(writer, sheet_name='Resumen', index=False)
                if not df_details.empty:
                    df_details.to_excel(writer, sheet_name='Detalles', index=False)
                if not df_errors.empty:
                    df_errors.to_excel(writer, sheet_name='Errores', index=False)
            
            self.logger.info(f"Reporte generado: {report_path}")
            return report_path
            
        except Exception as e:
            self.logger.error(f"Error al generar reporte: {str(e)}")
            raise
    
    def run_comparison(self, xml_file: str) -> str:
        """Ejecuta la comparación completa y genera el reporte con un archivo XML específico"""
        try:
            self.logger.info(f"Iniciando comparación de campos con archivo: {xml_file}")
            
            # Conectar a la base de datos
            if not self.connect_to_db():
                raise Exception("No se pudo conectar a la base de datos")
            
            # Realizar la comparación con un archivo específico
            results = self.compare_with_single_file(xml_file)
            
            # Generar el reporte
            report_path = self.generate_simple_report(results)
            
            # Cerrar conexión
            self.close_db_connection()
            
            self.logger.info("Comparación completada exitosamente")
            return report_path
            
        except Exception as e:
            self.logger.error(f"Error durante la comparación: {str(e)}")
            if self.db_connection:
                self.close_db_connection()
            raise
