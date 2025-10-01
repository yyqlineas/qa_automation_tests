import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from lxml import etree
import psycopg2
import logging
from datetime import datetime
import os
import json

class XPathMapper:
    def __init__(self, config_file=None, mapping_file=None):
        self.config_file = config_file
        self.mapping_file = mapping_file
        self.mapping_data = None
        self.conn = None
        self.cursor = None
        self.config = None
        self.logger = self._setup_logger()
        
        # Cargar configuración primero
        if config_file:
            self._load_config()
        
        # Si se proporciona archivo de mapeo, cargarlo después de la configuración
        if mapping_file:
            self.mapping_file = mapping_file
            self.load_mapping_file()

    def _setup_logger(self):
        # Configurar el logger
        logger = logging.getLogger('XPathMapper')
        logger.setLevel(logging.INFO)
        
        # Crear el directorio de logs si no existe
        logs_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'logs')
        os.makedirs(logs_dir, exist_ok=True)
        
        # Crear el archivo de log con la fecha actual
        log_file = os.path.join(logs_dir, f'comparison_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
        # Configurar FileHandler con encoding UTF-8 para soportar emojis
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        
        # Configurar el formato del log
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        
        # Agregar el handler al logger
        logger.addHandler(file_handler)
        
        return logger

    def _process_xpath(self, xpath_raw):
        """Procesar rutas XPath desde diferentes formatos, incluyendo clausulas WHEN."""
        if pd.isna(xpath_raw) or not xpath_raw:
            return None
            
        xpath_str = str(xpath_raw).strip()
        
        # Si contiene "CAD Field" o similar, no es un XPath válido - skip
        if any(term in xpath_str for term in ['CAD Field', 'Field Notes', 'Unnamed:']):
            return None
        
        # Si ya es una ruta XPath válida (comienza con // o /)
        if xpath_str.startswith('//') or xpath_str.startswith('/'):
            return xpath_str
        
        # Detectar clausulas WHEN descriptivas
        if 'WHEN' in xpath_str.upper():
            return self._parse_when_clause(xpath_str)
        
        # Si contiene el operador de concatenación "+", procesarlo por partes
        if '+' in xpath_str:
            # Dividir por "+" y procesar cada parte individualmente
            parts = [part.strip() for part in xpath_str.split('+')]
            processed_parts = []
            
            for part in parts:
                # Procesar cada parte individualmente
                if '<' in part and '>' in part:
                    # Extraer los nombres de los elementos
                    import re
                    elements = re.findall(r'<([^>]+)>', part)
                    if elements:
                        # Construir la ruta XPath para esta parte
                        part_xpath = '//' + '/'.join(elements)
                        processed_parts.append(part_xpath)
                elif part.startswith('//') or part.startswith('/'):
                    processed_parts.append(part)
                elif '/' in part and not any(term in part for term in ['<', '>', 'Field']):
                    # Ruta de elementos separados por / - agregar //
                    processed_parts.append('//' + part)
                elif part and not any(char in part for char in ['/', '<', '>', 'Field']):
                    processed_parts.append(f'//{part}')
            
            # Unir las partes con " + " para mantener el formato de concatenación
            if processed_parts:
                return ' + '.join(processed_parts)
        
        # Si está en formato de elementos XML: <Incident> <IncidentNumber>
        if '<' in xpath_str and '>' in xpath_str:
            # Extraer los nombres de los elementos
            import re
            elements = re.findall(r'<([^>]+)>', xpath_str)
            if elements:
                # Construir la ruta XPath
                xpath = '//' + '/'.join(elements)
                return xpath
        
        # Si es una ruta con / pero no empieza con //, agregar //
        if '/' in xpath_str and not any(term in xpath_str for term in ['<', '>', 'Field']):
            return '//' + xpath_str
        
        # Si es un nombre simple, asumir que es un elemento raíz
        if xpath_str and not any(char in xpath_str for char in ['/', '<', '>', 'Field']):
            return f'//{xpath_str}'
        
        return None

    def _parse_when_clause(self, description):
        """
        Parsear clausulas WHEN descriptivas del Excel.
        Ejemplo: 'Populate the "<CombinedIncidentVehicleData><Vehicles><VehicleData><TimeCallCleared>" as the Unit Cancelled time, WHEN <CombinedIncidentVehicleData><Vehicles><VehicleData><UnitDispositions><UnitDisposition><Description> contains "Cancelled"'
        """
        try:
            import re
            
            self.logger.debug(f"Parseando clausula WHEN: {description}")
            
            # Inicializar variables para evitar errores de scope
            operator = None
            condition_value = None
            target_elements = []
            condition_elements = []
            
            # Extraer el elemento objetivo (antes de "WHEN")
            # Buscar patrones como "<Element1><Element2><Element3>"
            target_pattern = r'"?<([^>]+(?:><[^>]+)*)>"?[^"]*(?:as[^"]*)?WHEN'
            target_match = re.search(target_pattern, description, re.IGNORECASE)
            
            if not target_match:
                # Patrón alternativo más flexible
                target_pattern = r'<([^>]+(?:><[^>]+)*)>'
                target_matches = re.findall(target_pattern, description)
                if target_matches and len(target_matches) >= 2:
                    # Usar el primer match como target
                    target_elements = target_matches[0].split('><')
                    condition_elements = target_matches[1].split('><')
                    # Para este caso, asumir operador contains por defecto
                    operator = 'contains'
                    condition_value = ''  # Se definirá más adelante si es necesario
                else:
                    self.logger.warning(f"No se pudieron extraer elementos de: {description}")
                    return None
            else:
                target_raw = target_match.group(1)
                target_elements = target_raw.split('><')
                
                # Buscar los elementos de condición después de WHEN
                when_pattern = r'WHEN\s+<([^>]+(?:><[^>]+)*)>\s+(contains?|equals?|!=|=)\s*["\']([^"\']+)["\']'
                when_match = re.search(when_pattern, description, re.IGNORECASE)
                
                if not when_match:
                    self.logger.warning(f"No se pudo extraer condición WHEN de: {description}")
                    return '//' + '/'.join(target_elements)  # Devolver XPath simple sin condición
                
                condition_raw = when_match.group(1)
                operator = when_match.group(2).lower()
                condition_value = when_match.group(3)
                condition_elements = condition_raw.split('><')
            
            # Validar que operator esté definido
            if operator is None:
                self.logger.warning(f"No se pudo determinar el operador de la condición WHEN de: {description}")
                return '//' + '/'.join(target_elements)
            
            # Encontrar el elemento común más profundo entre target y condition
            common_base = []
            min_length = min(len(target_elements), len(condition_elements))
            
            for i in range(min_length):
                if target_elements[i] == condition_elements[i]:
                    common_base.append(target_elements[i])
                else:
                    break
            
            if not common_base:
                # Si no hay base común, usar el primer elemento del target
                common_base = [target_elements[0]]
            
            # Construir XPath relativo desde la base común
            remaining_target = target_elements[len(common_base):]
            remaining_condition = condition_elements[len(common_base):]
            
            # Construir el XPath condicional
            base_xpath = '//' + '/'.join(common_base)
            condition_path = '/'.join(remaining_condition) if remaining_condition else '/'.join(condition_elements)
            target_path = '/'.join(remaining_target) if remaining_target else '/'.join(target_elements)
            
            # Construir XPath según el operador
            if operator in ['contains', 'contain']:
                if remaining_condition:
                    xpath_with_condition = f"{base_xpath}[contains({condition_path}, '{condition_value}')]/{target_path}"
                else:
                    xpath_with_condition = f"//VehicleData[contains(.//Description, '{condition_value}')]/TimeCallCleared"
            elif operator in ['equals', 'equal', '=']:
                if remaining_condition:
                    xpath_with_condition = f"{base_xpath}[{condition_path} = '{condition_value}']/{target_path}"
                else:
                    xpath_with_condition = f"//VehicleData[.//Description = '{condition_value}']/{target_path}"
            elif operator == '!=':
                if remaining_condition:
                    xpath_with_condition = f"{base_xpath}[{condition_path} != '{condition_value}']/{target_path}"
                else:
                    xpath_with_condition = f"//VehicleData[.//Description != '{condition_value}']/{target_path}"
            else:
                self.logger.warning(f"Operador no soportado: {operator}")
                return '//' + '/'.join(target_elements)
            
            self.logger.info(f"XPath condicional generado: {xpath_with_condition}")
            return xpath_with_condition
                
        except Exception as e:
            self.logger.error(f"Error parseando clausula WHEN: {str(e)}")
            return None

    def _evaluate_xpath_with_conditions(self, root, xpath_str):
        """
        Evaluar XPath con soporte para condiciones condicionales.
        Analiza si hay condiciones en el XPath y extrae el valor correcto.
        """
        try:
            # Verificar si el XPath contiene condiciones múltiples separadas por "+"
            if '+' in xpath_str:
                xpath_parts = [part.strip() for part in xpath_str.split('+')]
                xml_values = []
                
                for xpath_part in xpath_parts:
                    part_value = self._evaluate_single_xpath_with_condition(root, xpath_part)
                    if part_value:  # Solo agregar si no está vacío
                        xml_values.append(part_value)
                
                # Concatenar los valores encontrados
                if xml_values:
                    return " + ".join(xml_values)
                else:
                    return None
            else:
                # XPath simple (sin concatenación)
                return self._evaluate_single_xpath_with_condition(root, xpath_str)
                
        except Exception as e:
            self.logger.error(f"Error evaluando XPath condicional '{xpath_str}': {str(e)}")
            return "ERROR_XPATH_CONDITION"

    def _evaluate_single_xpath_with_condition(self, root, xpath_part):
        """
        Evaluar un solo XPath que puede contener condiciones.
        Maneja tanto XPath tradicionales como clausulas descriptivas.
        """
        try:
            self.logger.debug(f"Evaluando XPath: {xpath_part}")
            
            # Método 1: Intentar XPath directo primero
            try:
                elements = root.xpath(xpath_part)
                if elements:
                    value = self._extract_element_value(elements[0])
                    if value:
                        self.logger.debug(f"XPath directo exitoso: '{value}'")
                        return value
            except Exception as e:
                self.logger.debug(f"XPath directo falló: {str(e)}")
            
            # Método 2: Si el XPath directo falla, intentar parsing manual
            # Buscar patrones como "//VehicleData[contains(.//Description, 'Cancelled')]/TimeCallCleared"
            import re
            
            # Detectar XPath con función contains()
            contains_pattern = r'(.+?)\[contains\(([^,]+),\s*[\'"]([^\'\"]+)[\'"]\)\]/(.+)'
            contains_match = re.match(contains_pattern, xpath_part)
            
            if contains_match:
                base_path = contains_match.group(1)
                condition_element = contains_match.group(2).strip()
                condition_value = contains_match.group(3)
                target_element = contains_match.group(4)
                
                self.logger.debug(f"Patron contains detectado - Base: {base_path}, Condicion: {condition_element} contains '{condition_value}', Target: {target_element}")
                
                # Buscar todos los elementos base
                base_elements = root.xpath(base_path)
                self.logger.debug(f"Encontrados {len(base_elements)} elementos base")
                
                for element in base_elements:
                    # Evaluar la condición contains
                    try:
                        if condition_element.startswith('.//'):
                            # Buscar en descendientes
                            condition_elements = element.xpath(condition_element)
                        else:
                            # Buscar relativo al elemento
                            condition_elements = element.xpath('.//' + condition_element)
                        
                        for cond_elem in condition_elements:
                            cond_value = self._extract_element_value(cond_elem)
                            self.logger.debug(f"Verificando condicion: '{cond_value}' contains '{condition_value}'")
                            
                            if condition_value.lower() in str(cond_value).lower():
                                # Condición se cumple, buscar el elemento target
                                target_elements = element.xpath('.//' + target_element)
                                if not target_elements:
                                    target_elements = element.xpath(target_element)
                                
                                if target_elements:
                                    target_value = self._extract_element_value(target_elements[0])
                                    self.logger.info(f"Condicion cumplida! Valor extraido: '{target_value}'")
                                    return target_value
                                else:
                                    self.logger.debug(f"Condicion cumplida pero elemento target '{target_element}' no encontrado")
                    except Exception as e:
                        self.logger.debug(f"Error evaluando condicion en elemento: {str(e)}")
                        continue
                
                self.logger.debug("Ninguna condicion se cumplio")
                return None
            
            # Método 3: XPath con otras condiciones [elemento = 'valor']
            condition_pattern = r'(.+?)\[([^=!<>]+)\s*(=|!=|>|<|>=|<=)\s*[\'"]([^\'\"]+)[\'"]\]/(.+)'
            condition_match = re.match(condition_pattern, xpath_part)
            
            if condition_match:
                base_path = condition_match.group(1)
                condition_element = condition_match.group(2).strip()
                operator = condition_match.group(3)
                condition_value = condition_match.group(4)
                target_element = condition_match.group(5)
                
                self.logger.debug(f"Patron condicional detectado - Base: {base_path}, Condicion: {condition_element} {operator} '{condition_value}', Target: {target_element}")
                
                base_elements = root.xpath(base_path)
                
                for element in base_elements:
                    try:
                        # Buscar elemento de condición
                        if condition_element.startswith('.//'):
                            condition_elements = element.xpath(condition_element)
                        else:
                            condition_elements = element.xpath('.//' + condition_element)
                        
                        for cond_elem in condition_elements:
                            cond_value = self._extract_element_value(cond_elem)
                            
                            # Evaluar la condición
                            if self._compare_values(cond_value, operator, condition_value):
                                # Buscar elemento target
                                target_elements = element.xpath('.//' + target_element)
                                if not target_elements:
                                    target_elements = element.xpath(target_element)
                                
                                if target_elements:
                                    target_value = self._extract_element_value(target_elements[0])
                                    self.logger.info(f"Condicion '{cond_value}' {operator} '{condition_value}' cumplida! Valor: '{target_value}'")
                                    return target_value
                    except Exception as e:
                        self.logger.debug(f"Error evaluando condicion: {str(e)}")
                        continue
                
                return None
            
            # Método 4: XPath simple sin condiciones
            try:
                elements = root.xpath(xpath_part)
                if elements:
                    return self._extract_element_value(elements[0])
            except:
                pass
            
            self.logger.debug(f"No se pudo evaluar el XPath: {xpath_part}")
            return None
                    
        except Exception as e:
            self.logger.error(f"Error evaluando XPath '{xpath_part}': {str(e)}")
            return None

    def _evaluate_condition(self, element, condition_str):
        """
        Evaluar una condición XPath contra un elemento.
        Soporta operadores: !=, =, <, >, <=, >=
        """
        try:
            import re
            
            # Extraer operador y valores de la condición
            # Ejemplo: "UnitCode != 'Batt_ID'"
            operators = ['!=', '>=', '<=', '=', '>', '<']
            
            for op in operators:
                if op in condition_str:
                    parts = condition_str.split(op, 1)
                    if len(parts) == 2:
                        left_expr = parts[0].strip()
                        right_value = parts[1].strip().strip("'\"")  # Remover comillas
                        
                        # Evaluar la expresión izquierda contra el elemento
                        left_elements = element.xpath(left_expr)
                        if left_elements:
                            left_actual_value = self._extract_element_value(left_elements[0])
                            
                            # Realizar la comparación
                            return self._compare_values(left_actual_value, op, right_value)
                        else:
                            # Si no se encuentra el elemento, considerar como valor vacío
                            return self._compare_values("", op, right_value)
            
            self.logger.warning(f"No se pudo parsear la condición: {condition_str}")
            return False
            
        except Exception as e:
            self.logger.error(f"Error evaluando condición '{condition_str}': {str(e)}")
            return False

    def _compare_values(self, actual_value, operator, expected_value):
        """Comparar dos valores usando el operador especificado."""
        try:
            actual_str = str(actual_value).strip() if actual_value is not None else ""
            expected_str = str(expected_value).strip()
            
            if operator == '!=':
                result = actual_str != expected_str
            elif operator == '=':
                result = actual_str == expected_str
            elif operator == '>':
                # Intentar comparación numérica, si falla usar string
                try:
                    result = float(actual_str) > float(expected_str)
                except:
                    result = actual_str > expected_str
            elif operator == '<':
                try:
                    result = float(actual_str) < float(expected_str)
                except:
                    result = actual_str < expected_str
            elif operator == '>=':
                try:
                    result = float(actual_str) >= float(expected_str)
                except:
                    result = actual_str >= expected_str
            elif operator == '<=':
                try:
                    result = float(actual_str) <= float(expected_str)
                except:
                    result = actual_str <= expected_str
            else:
                result = False
            
            self.logger.debug(f"Comparación: '{actual_str}' {operator} '{expected_str}' = {result}")
            return result
            
        except Exception as e:
            self.logger.error(f"Error comparando valores: {str(e)}")
            return False

    def _extract_element_value(self, element):
        """Extraer valor de un elemento XML."""
        try:
            if hasattr(element, 'text') and element.text:
                return element.text.strip()
            elif hasattr(element, 'tag'):
                # Si es un elemento, obtener todo el texto interno
                return ''.join(element.itertext()).strip()
            else:
                # Si es un atributo o valor directo
                return str(element).strip()
        except Exception as e:
            self.logger.error(f"Error extrayendo valor del elemento: {str(e)}")
            return ""

    def _load_config(self):
        """Cargar la configuración desde el archivo JSON."""
        try:
            with open(self.config_file, 'r') as f:
                self.config = json.load(f)
            self.logger.info(f"Configuración cargada exitosamente: {self.config_file}")
            return True
        except Exception as e:
            self.logger.error(f"Error al cargar la configuración: {str(e)}")
            return False

    def _normalize_field_name(self, field_name):
        """Normalizar nombres de campos reemplazando espacios con guiones bajos y corrigiendo nombres."""
        if not field_name:
            return field_name
        
        # Mapeo de corrección de nombres de campos
        field_corrections = {
            # Campos que no existen en dispatch pero sí en nfirs_notification
            'street_number': 'house_num',  # dispatch no lo tiene, nfirs sí
            'street_prefix': 'prefix_direction',  # dispatch no lo tiene, nfirs sí  
            'street_suffix': 'suffix_direction',  # dispatch no lo tiene, nfirs sí
        }
        
        normalized = str(field_name).strip().replace(' ', '_')
        
        # Aplicar correcciones específicas si existen
        if normalized in field_corrections:
            return field_corrections[normalized]
            
        return normalized
    
    def _extract_xml_value_with_special_logic(self, root, xpath_str, column_name, table_name):
        """
        Extraer valor XML con lógicas especiales según el tipo de campo.
        """
        # Detectar campos especiales
        field_key = f"{table_name}.{column_name}".lower()
        
        # 1. Lógica especial para status_code
        if column_name.lower() == 'status_code':
            return self._extract_status_code_value(root, xpath_str)
        
        # 2. Lógica especial para call_notes, narratives y comentarios  
        if column_name.lower() in ['call_notes', 'narratives', 'narrative', 'comment1', 'comment', 'comments']:
            return self._extract_narratives_value(root, xpath_str)
        
        # 3. Lógica especial para cross_street
        if 'cross_street' in column_name.lower():
            self.logger.info(f"Activando lógica especial de cross_street para campo: {column_name}")
            return self._extract_cross_street_value(root, xpath_str)
        
        # Para otros campos, usar la lógica estándar con fallback case-insensitive
        result = self._evaluate_xpath_with_conditions(root, xpath_str)
        
        # Si no se encontró nada, intentar con variaciones de mayúsculas/minúsculas
        if result is None and xpath_str:
            self.logger.debug(f"XPath '{xpath_str}' no encontró elementos, intentando con variaciones case-insensitive")
            result = self._try_case_insensitive_xpath(root, xpath_str)
        
        return result
    
    def _try_case_insensitive_xpath(self, root, xpath_str):
        """
        Intentar variaciones case-insensitive del XPath si el original falla.
        Genera todas las combinaciones posibles de case-insensitive basadas en elementos reales del XML.
        """
        try:
            import re
            from itertools import product
            
            self.logger.debug(f"Buscando variaciones case-insensitive para: {xpath_str}")
            
            # Extraer partes del XPath (elementos entre /)
            xpath_parts = [part.strip() for part in xpath_str.split('/') if part.strip()]
            
            # Obtener todos los elementos del XML para hacer matching case-insensitive
            all_elements = root.xpath("//*")
            element_names = set()
            
            for element in all_elements:
                element_names.add(element.tag)
                # También agregar elementos padre para navegación
                parent = element.getparent()
                if parent is not None:
                    element_names.add(parent.tag)
            
            self.logger.debug(f"Elementos encontrados en XML: {sorted(element_names)}")
            
            # Para cada parte del XPath, encontrar todas las variaciones case-insensitive posibles
            xpath_part_variations = []
            
            for xpath_part in xpath_parts:
                if xpath_part and not xpath_part.startswith('//'):
                    # Encontrar todas las variaciones case-insensitive para esta parte
                    part_variations = [xpath_part]  # Incluir original
                    
                    for xml_element in element_names:
                        if xpath_part.lower() == xml_element.lower() and xpath_part != xml_element:
                            part_variations.append(xml_element)
                    
                    xpath_part_variations.append(part_variations)
                    self.logger.debug(f"Parte '{xpath_part}' tiene variaciones: {part_variations}")
                else:
                    # Mantener partes especiales como '//' sin cambios
                    xpath_part_variations.append([xpath_part])
            
            # Generar todas las combinaciones posibles usando product de itertools
            all_combinations = list(product(*xpath_part_variations))
            
            # Crear XPaths completos para cada combinación
            variations_to_try = []
            
            for combination in all_combinations:
                # Reconstruir el XPath con esta combinación
                new_xpath = '/' + '/'.join(combination)
                if new_xpath != xpath_str:  # Solo agregar si es diferente al original
                    variations_to_try.append(new_xpath)
            
            # También agregar algunas variaciones específicas conocidas
            specific_variations = [
                xpath_str.replace('NfirsData', 'NFIRSData'),
                xpath_str.replace('nfirsdata', 'NFIRSData'),
                xpath_str.replace('Nfirs', 'NFIRS'),
                xpath_str.replace('nfirs', 'NFIRS'),
                xpath_str.replace('DATA', 'Data'),
                xpath_str.replace('data', 'Data'),
            ]
            
            # Combinar todas las variaciones y remover duplicados
            all_variations = variations_to_try + specific_variations
            unique_variations = []
            for variation in all_variations:
                if variation not in unique_variations and variation != xpath_str:
                    unique_variations.append(variation)
            
            self.logger.debug(f"Intentando {len(unique_variations)} variaciones case-insensitive")
            
            # Intentar cada variación
            for i, variation in enumerate(unique_variations):
                self.logger.debug(f"Probando variación {i+1}/{len(unique_variations)}: {variation}")
                try:
                    elements = root.xpath(variation)
                    if elements:
                        # Encontrado! usar la primera variación que funcione
                        self.logger.info(f"✅ XPath corregido funciona: '{xpath_str}' -> '{variation}'")
                        result = self._extract_element_value(elements[0])
                        if result:
                            return result
                except Exception as e:
                    self.logger.debug(f"Error probando variación '{variation}': {e}")
                    continue
            
            self.logger.warning(f"❌ Ninguna de las {len(unique_variations)} variaciones case-insensitive funcionó para: {xpath_str}")
            return None
            
        except Exception as e:
            self.logger.error(f"Error en fallback case-insensitive para {xpath_str}: {e}")
            return None
    
    def _try_case_insensitive_xpath_elements(self, root, xpath_str):
        """
        Intentar variaciones case-insensitive del XPath y devolver los elementos encontrados.
        Busca todos los elementos del XML y encuentra coincidencias case-insensitive.
        """
        try:
            import re
            
            # Extraer partes del XPath (elementos entre /)
            xpath_parts = [part.strip() for part in xpath_str.split('/') if part.strip()]
            
            # Obtener todos los elementos del XML para hacer matching case-insensitive
            all_elements = root.xpath("//*")  # Obtener todos los elementos
            element_names = set()
            
            for element in all_elements:
                element_names.add(element.tag)
                # También agregar elementos padre para navegación
                parent = element.getparent()
                if parent is not None:
                    element_names.add(parent.tag)
            
            # Crear variaciones del XPath original
            original_variations = [
                # Mantener XPath original
                xpath_str,
                # Casos específicos conocidos para compatibilidad
                xpath_str.replace('NfirsData', 'NFIRSData'),
                xpath_str.replace('nfirsdata', 'NFIRSData'),
                xpath_str.replace('Nfirs', 'NFIRS'),
                xpath_str.replace('nfirs', 'NFIRS'),
            ]
            
            # Generar variaciones case-insensitive basadas en elementos reales del XML
            for xpath_part in xpath_parts:
                if xpath_part and not xpath_part.startswith('//'):
                    # Buscar coincidencias case-insensitive en elementos reales
                    for xml_element in element_names:
                        if xpath_part.lower() == xml_element.lower() and xpath_part != xml_element:
                            # Crear nueva variación reemplazando la parte case-insensitive
                            new_variation = xpath_str.replace(xpath_part, xml_element)
                            if new_variation not in original_variations:
                                original_variations.append(new_variation)
            
            # Remover duplicados manteniendo orden
            unique_variations = []
            for variation in original_variations:
                if variation not in unique_variations:
                    unique_variations.append(variation)
            
            # Intentar cada variación
            for variation in unique_variations:
                if variation != xpath_str:  # Solo intentar variaciones diferentes al original
                    self.logger.debug(f"Intentando variación case-insensitive para elementos: {variation}")
                    elements = root.xpath(variation)
                    if elements:
                        # Encontrado! usar la primera variación que funcione
                        self.logger.info(f"✅ XPath corregido funciona para elementos: '{xpath_str}' -> '{variation}'")
                        return elements
            
            self.logger.debug(f"Ninguna variación case-insensitive funcionó para elementos: {xpath_str}")
            return None
            
        except Exception as e:
            self.logger.error(f"Error en fallback case-insensitive para elementos {xpath_str}: {e}")
            return None
    
    def _extract_status_code_value(self, root, xpath_str):
        """
        Lógica especial para status_code:
        - Si xpath tiene condicional WHEN, aplicar lógica
        - Si no, si xpath tiene valor = closed, sino = open por defecto
        """
        try:
            self.logger.debug(f"Procesando status_code con xpath: {xpath_str}")
            
            # Verificar si hay condiciones WHEN en el xpath
            if 'WHEN' in xpath_str.upper():
                return self._parse_when_condition_status(root, xpath_str)
            
            # Lógica por defecto: si el xpath existe y tiene valor = closed
            value = self._evaluate_xpath_with_conditions(root, xpath_str)
            if value and str(value).strip():
                # Si existe valor en el xpath, considerarlo como "closed"
                return "closed"
            else:
                # Si no se encontró valor, intentar con case-insensitive fallback
                if not value:
                    self.logger.debug(f"XPath '{xpath_str}' no encontró valor para status_code, intentando case-insensitive")
                    case_insensitive_value = self._try_case_insensitive_xpath(root, xpath_str)
                    if case_insensitive_value and str(case_insensitive_value).strip():
                        return "closed"
                
                # Si aún no hay valor, considerarlo como "open"
                return "open"
                
        except Exception as e:
            self.logger.error(f"Error procesando status_code: {e}")
            return "open"  # Valor por defecto
    
    def _extract_narratives_value(self, root, xpath_str):
        """
        Lógica especial para call_notes/narratives:
        - Concatenar múltiples valores: xpath+xpath+xpath (con espacios)
        - Manejar DateTime+Text en múltiples líneas
        """
        try:
            self.logger.debug(f"Procesando narratives con xpath: {xpath_str}")
            
            # Verificar si hay concatenación múltiple con +
            if '+' in xpath_str:
                return self._extract_concatenated_narratives(root, xpath_str)
            
            # Para xpath simple, obtener todos los valores que coincidan
            elements = root.xpath(xpath_str)
            if elements:
                values = []
                for element in elements:
                    value = self._extract_element_value(element)
                    if value and str(value).strip():
                        values.append(str(value).strip())
                
                if values:
                    # Concatenar con espacios entre valores
                    return " ".join(values)
            
            # Si no se encontró nada, intentar con variaciones case-insensitive
            if not elements:
                self.logger.debug(f"XPath '{xpath_str}' no encontró elementos para narratives, intentando con variaciones case-insensitive")
                case_insensitive_result = self._try_case_insensitive_xpath(root, xpath_str)
                if case_insensitive_result:
                    return case_insensitive_result
            
            return None
            
        except Exception as e:
            self.logger.error(f"Error procesando narratives: {e}")
            return None
    
    def _extract_concatenated_narratives(self, root, xpath_str):
        """
        Manejar concatenación compleja para narratives: DateTime+Text en múltiples líneas
        """
        try:
            xpath_parts = [part.strip() for part in xpath_str.split('+')]
            self.logger.debug(f"Partes de xpath para narratives: {xpath_parts}")
            
            # Verificar si es patrón DateTime+Text
            if len(xpath_parts) == 2:
                datetime_xpath = xpath_parts[0]
                text_xpath = xpath_parts[1]
                
                # Obtener todos los elementos DateTime
                datetime_elements = root.xpath(datetime_xpath)
                text_elements = root.xpath(text_xpath)
                
                # Si no se encontraron elementos, intentar con case-insensitive
                if not datetime_elements:
                    self.logger.debug(f"XPath DateTime '{datetime_xpath}' no encontró elementos, intentando case-insensitive")
                    datetime_fixed = self._try_case_insensitive_xpath_elements(root, datetime_xpath)
                    if datetime_fixed:
                        datetime_elements = datetime_fixed
                        
                if not text_elements:
                    self.logger.debug(f"XPath Text '{text_xpath}' no encontró elementos, intentando case-insensitive")
                    text_fixed = self._try_case_insensitive_xpath_elements(root, text_xpath)
                    if text_fixed:
                        text_elements = text_fixed
                
                # Crear líneas combinando DateTime + Text
                narrative_lines = []
                max_elements = max(len(datetime_elements), len(text_elements))
                
                for i in range(max_elements):
                    line_parts = []
                    
                    # Agregar DateTime si existe
                    if i < len(datetime_elements):
                        dt_value = self._extract_element_value(datetime_elements[i])
                        if dt_value:
                            line_parts.append(str(dt_value).strip())
                    
                    # Agregar Text si existe
                    if i < len(text_elements):
                        text_value = self._extract_element_value(text_elements[i])
                        if text_value:
                            line_parts.append(str(text_value).strip())
                    
                    if line_parts:
                        narrative_lines.append(" ".join(line_parts))

                if narrative_lines:
                    # Unir con saltos de línea
                    return "\n".join(narrative_lines)            # Concatenación simple con espacios
            else:
                values = []
                for xpath_part in xpath_parts:
                    elements = root.xpath(xpath_part)
                    
                    # Si no se encontraron elementos, intentar case-insensitive
                    if not elements:
                        self.logger.debug(f"XPath parte '{xpath_part}' no encontró elementos, intentando case-insensitive")
                        elements = self._try_case_insensitive_xpath_elements(root, xpath_part)
                        if not elements:
                            elements = []
                    
                    for element in elements:
                        value = self._extract_element_value(element)
                        if value and str(value).strip():
                            values.append(str(value).strip())
                
                if values:
                    return " ".join(values)
            
            return None
            
        except Exception as e:
            self.logger.error(f"Error en concatenación narratives: {e}")
            return None
    
    def _extract_cross_street_value(self, root, xpath_str):
        """
        Lógica especial para cross_street:
        - cross_street1 + cross_street2 = cross1/cross2
        - cross_street1 AND cross_street2 = cross1/cross2
        Soporta múltiples operadores de concatenación para flexibilidad
        """
        try:
            self.logger.debug(f"Procesando cross_street con xpath: {xpath_str}")
            
            # Detectar operadores de concatenación (+ o AND)
            xpath_parts = []
            
            if '+' in xpath_str:
                xpath_parts = [part.strip() for part in xpath_str.split('+')]
                self.logger.debug(f"Detectado operador '+', partes: {xpath_parts}")
            elif ' AND ' in xpath_str or ' and ' in xpath_str:
                # Detectar AND en mayúsculas o minúsculas
                if ' AND ' in xpath_str:
                    xpath_parts = [part.strip() for part in xpath_str.split(' AND ')]
                    self.logger.debug(f"Detectado operador 'AND' (mayúsculas), partes: {xpath_parts}")
                else:
                    xpath_parts = [part.strip() for part in xpath_str.split(' and ')]
                    self.logger.debug(f"Detectado operador 'and' (minúsculas), partes: {xpath_parts}")
            else:
                # DETECCIÓN AUTOMÁTICA: Buscar patrones de concatenación comunes
                # Ejemplo: //Path/LowCrossStreet/Path/HighCrossStreet -> //Path/LowCrossStreet + //Path/HighCrossStreet
                if 'LowCrossStreet' in xpath_str and 'HighCrossStreet' in xpath_str:
                    self.logger.info(f"Detectado patrón de cross_street sin operadores explícitos, intentando separación automática")
                    
                    # Buscar el patrón: ...LowCrossStreet/...HighCrossStreet
                    if '/LowCrossStreet/' in xpath_str and 'HighCrossStreet' in xpath_str:
                        # Ejemplo: //Incident/Location/LowCrossStreet/Incident/Location/HighCrossStreet
                        # Objetivo: //Incident/Location/LowCrossStreet + //Incident/Location/HighCrossStreet
                        
                        # Encontrar donde termina LowCrossStreet
                        low_end = xpath_str.find('/LowCrossStreet') + len('/LowCrossStreet')
                        low_part = xpath_str[:low_end]
                        
                        # Encontrar la base común para HighCrossStreet
                        # Buscar desde LowCrossStreet hasta HighCrossStreet
                        high_start_pos = xpath_str.find('HighCrossStreet')
                        if high_start_pos != -1:
                            # Tomar la parte base y añadir HighCrossStreet
                            base_path = xpath_str[:xpath_str.find('/LowCrossStreet')]
                            high_part = base_path + '/HighCrossStreet'
                            
                            xpath_parts = [low_part, high_part]
                            self.logger.info(f"Separación automática: {xpath_parts}")
            
            
            # Si encontramos operadores de concatenación, procesar múltiples XPaths
            if xpath_parts and len(xpath_parts) > 1:
                street_values = []
                
                for xpath_part in xpath_parts:
                    value = self._evaluate_xpath_with_conditions(root, xpath_part)
                    if value and str(value).strip():
                        street_values.append(str(value).strip())
                        self.logger.debug(f"Valor extraído de '{xpath_part}': '{value}'")
                
                if street_values:
                    # Unir con / entre valores
                    result = "/".join(street_values)
                    self.logger.info(f"Cross street combinado: {result}")
                    return result
                else:
                    self.logger.warning(f"No se encontraron valores para ninguna parte del cross_street")
            
            # Para xpath simple (sin operadores de concatenación)
            self.logger.debug(f"Procesando como XPath simple: {xpath_str}")
            return self._evaluate_xpath_with_conditions(root, xpath_str)
            
        except Exception as e:
            self.logger.error(f"Error procesando cross_street: {e}")
            return None
    
    def _parse_when_condition_status(self, root, xpath_str):
        """
        Parsear condiciones WHEN para status_code
        Ejemplo: WHEN //Path/Status = 'Complete' THEN 'closed' ELSE 'open'
        """
        try:
            import re
            
            # Patrón para WHEN condition THEN value ELSE value
            when_pattern = r'WHEN\s+(.+?)\s*=\s*[\'"]([^\'\"]+)[\'"]\s+THEN\s+[\'"]([^\'\"]+)[\'"]\s+ELSE\s+[\'"]([^\'\"]+)[\'"]'
            match = re.search(when_pattern, xpath_str, re.IGNORECASE)
            
            if match:
                condition_xpath = match.group(1).strip()
                condition_value = match.group(2)
                then_value = match.group(3)
                else_value = match.group(4)
                
                self.logger.debug(f"WHEN condition: {condition_xpath} = '{condition_value}' THEN '{then_value}' ELSE '{else_value}'")
                
                # Evaluar condición
                actual_value = self._evaluate_xpath_with_conditions(root, condition_xpath)
                
                # Si no se encontró valor, intentar con case-insensitive fallback
                if not actual_value:
                    self.logger.debug(f"XPath condición '{condition_xpath}' no encontró valor para status_code, intentando case-insensitive")
                    actual_value = self._try_case_insensitive_xpath(root, condition_xpath)
                
                if actual_value and str(actual_value).strip().lower() == condition_value.lower():
                    return then_value
                else:
                    return else_value
            
            # Si no se puede parsear, usar lógica por defecto
            default_value = self._evaluate_xpath_with_conditions(root, xpath_str)
            if default_value:
                return "closed"
            else:
                # Si no se encontró valor, intentar con case-insensitive fallback
                case_insensitive_value = self._try_case_insensitive_xpath(root, xpath_str)
                if case_insensitive_value:
                    return "closed"
                else:
                    return "open"
            
        except Exception as e:
            self.logger.error(f"Error parseando WHEN condition: {e}")
            return "open"
    
    def _extract_record_identifiers(self, root, xml_file):
        """
        Extraer identificadores únicos del XML (xref_id, dispatch_number, incident_number).
        Retorna un diccionario con los identificadores encontrados.
        """
        identifiers = {
            'xref_id': None,
            'dispatch_number': None, 
            'incident_number': None
        }
        
        try:
            self.logger.debug(f"Extrayendo identificadores del XML: {xml_file}")
            
            # Buscar xref_id y dispatch_number en el mapeo
            if hasattr(self, 'mapping_data') and self.mapping_data is not None:
                for _, row in self.mapping_data.iterrows():
                    try:
                        table_name = row.get('table_name', '')
                        column_name = row.get('column_name', '')
                        xpath = row.get('xpath', '')
                        
                        if pd.isna(table_name) or pd.isna(column_name) or pd.isna(xpath):
                            continue
                        
                        column_lower = str(column_name).lower()
                        
                        # Buscar xref_id
                        if 'xref' in column_lower and xpath:
                            value = self._evaluate_xpath_with_conditions(root, xpath)
                            if value:
                                identifiers['xref_id'] = str(value).strip()
                                self.logger.info(f"xref_id extraído: {identifiers['xref_id']}")
                        
                        # Buscar dispatch_number
                        elif 'dispatch_number' in column_lower and xpath:
                            value = self._evaluate_xpath_with_conditions(root, xpath)
                            if value:
                                identifiers['dispatch_number'] = str(value).strip()
                                self.logger.info(f"dispatch_number extraído: {identifiers['dispatch_number']}")
                        
                        # Buscar incident_number
                        elif 'incident_number' in column_lower and xpath:
                            value = self._evaluate_xpath_with_conditions(root, xpath)
                            if value:
                                identifiers['incident_number'] = str(value).strip()
                                self.logger.info(f"incident_number extraído: {identifiers['incident_number']}")
                                
                    except Exception as e:
                        self.logger.debug(f"Error extrayendo identificador de fila: {e}")
                        continue
            
            # Si no se encontraron en el mapeo, intentar rutas comunes
            if not any(identifiers.values()):
                self.logger.info("No se encontraron identificadores en mapeo, probando rutas comunes...")
                
                # Rutas XPath comunes para identificadores
                common_xpaths = {
                    'xref_id': [
                        '//XRefId',
                        '//xref_id', 
                        '//Xref_Id',
                        '//CADMasterCallTable/CADActiveCallTable/RelatedRecordNumber',
                        '//MasterIncidentNumber',
                        '//DispatchID'
                    ],
                    'dispatch_number': [
                        '//DispatchNumber',
                        '//FirstDueExport/NFIRSData/DispatchNumber',  # Específico para FirstDueExport
                        '//FirstDueExport/NfirsData/DispatchNumber',   # Fallback case-insensitive
                        '//dispatch_number',
                        '//CallNumber',
                        '//IncidentNumber',
                        '//CADNumber'
                    ],
                    'incident_number': [
                        '//IncidentNumber',
                        '//incident_number',
                        '//RunNumber',
                        '//CallNumber'
                    ]
                }
                
                for id_type, xpaths in common_xpaths.items():
                    if identifiers[id_type] is None:
                        for xpath in xpaths:
                            try:
                                value = self._evaluate_xpath_with_conditions(root, xpath)
                                if value:
                                    identifiers[id_type] = str(value).strip()
                                    self.logger.info(f"{id_type} extraído con xpath común '{xpath}': {identifiers[id_type]}")
                                    break
                            except Exception as e:
                                continue
            
            # Log de resultado
            found_ids = {k: v for k, v in identifiers.items() if v}
            if found_ids:
                self.logger.info(f"Identificadores encontrados para {xml_file}: {found_ids}")
            else:
                self.logger.warning(f"No se pudieron extraer identificadores de {xml_file}")
            
            return identifiers
            
        except Exception as e:
            self.logger.error(f"Error extrayendo identificadores de {xml_file}: {e}")
            return identifiers
    
    def _verify_record_exists_in_db(self, record_identifiers):
        """
        Verificar si existe al menos un registro en la BD que coincida con los identificadores del XML.
        Busca el valor del XML en AMBOS campos: dispatch.xref_id Y nfirs_notification.dispatch_number
        Retorna (found: bool, table_name: str or None) - la tabla donde se encontró el registro.
        """
        if not record_identifiers or not any(record_identifiers.values()):
            return False, None
        
        try:
            # Obtener filtros de configuración
            filters = self.config.get('filters', {}) if self.config else {}
            batt_dept_filter = filters.get('batt_dept_id', {})
            datetime_filter = filters.get('datetime', {})
            
            batt_dept_values = batt_dept_filter.get('values', [4611])
            batt_dept_ids = ','.join(str(val) for val in batt_dept_values)
            start_datetime = datetime_filter.get('start_datetime', '2025-09-22 00:00:00')
            
            # Extraer el valor principal del XML (puede ser xref_id o dispatch_number)
            xml_value = None
            if record_identifiers.get('xref_id'):
                xml_value = record_identifiers['xref_id']
            elif record_identifiers.get('dispatch_number'):
                xml_value = record_identifiers['dispatch_number']
            
            if not xml_value:
                self.logger.warning("No se encontró xref_id ni dispatch_number en los identificadores")
                return False, None
            
            self.logger.info(f"🔍 Buscando valor '{xml_value}' en dispatch.xref_id y nfirs_notification.dispatch_number")
            
            # Buscar PRIMERO en dispatch.xref_id
            dispatch_query = f"""
                SELECT 'dispatch' as tabla_encontrada, xref_id FROM dispatch d 
                WHERE d.batt_dept_id IN ({batt_dept_ids})
                AND d.created_at >= '{start_datetime}'
                AND d.xref_id = '{xml_value}'
                LIMIT 1
            """
            
            # Buscar SEGUNDO en nfirs_notification.dispatch_number
            nfirs_query = f"""
                SELECT 'nfirs_notification' as tabla_encontrada, dispatch_number FROM nfirs_notification n 
                WHERE n.batt_dept_id IN ({batt_dept_ids})
                AND n.created_at >= '{start_datetime}'
                AND n.dispatch_number = '{xml_value}'
                LIMIT 1
            """
            
            # Combinar ambas queries con UNION para buscar en ambas tablas
            combined_query = f"({dispatch_query}) UNION ({nfirs_query})"
            
            self.logger.debug(f"Query de búsqueda combinada: {combined_query}")
            
            if self.cursor is not None:
                self.cursor.execute(combined_query)
                results = self.cursor.fetchall()
                
                if results:
                    # Retornar el primer resultado encontrado
                    tabla_encontrada = results[0][0]
                    valor_encontrado = results[0][1]
                    self.logger.info(f"✅ Registro encontrado en {tabla_encontrada}: {valor_encontrado}")
                    return True, tabla_encontrada
                else:
                    self.logger.warning(f"❌ Valor '{xml_value}' NO encontrado en dispatch.xref_id ni en nfirs_notification.dispatch_number")
                    return False, None
            
            return False, None
            
        except Exception as e:
            self.logger.error(f"Error verificando existencia de registro: {e}")
            return False, None
    
    def load_mapping_file(self):
        """Cargar el archivo de mapeo Excel."""
        try:
            if not self.mapping_file:
                raise Exception("No se ha especificado el archivo de mapeo")
                
            # Leer el archivo Excel
            self.mapping_data = pd.read_excel(self.mapping_file)
            
            self.logger.info(f"Columnas encontradas en el Excel: {list(self.mapping_data.columns)}")
            
            # Definir columnas específicas basándose en la estructura real del Excel
            xpath_col = 'Data Source Mappings'
            
            # Columnas para CAD Dispatch (columnas 5 y 6)
            dispatch_table_col = 'CAD Dispatch '  # Columna 5: nombres de tabla
            dispatch_field_col = 'Unnamed: 6'    # Columna 6: nombres de campo
            
            # Columnas para CAD Incident (columnas 7 y 8)
            incident_table_col = 'CAD Incident'   # Columna 7: nombres de tabla
            incident_field_col = 'Unnamed: 8'    # Columna 8: nombres de campo
            
            self.logger.info(f"Usando columnas definidas - XPath: {xpath_col}")
            self.logger.info(f"Dispatch - Tabla: {dispatch_table_col} (col 5), Campo: {dispatch_field_col} (col 6)")
            self.logger.info(f"Incident - Tabla: {incident_table_col} (col 7), Campo: {incident_field_col} (col 8)")
            
            # Crear mapeos válidos leyendo directamente del Excel
            valid_mappings = []
            
            for idx, row in self.mapping_data.iterrows():
                # Saltar fila de headers (fila 0)
                if idx == 0:
                    continue
                    
                # Obtener ruta XPath
                xpath_raw = row.get(xpath_col, '') if xpath_col in self.mapping_data.columns else ''
                if pd.isna(xpath_raw) or not str(xpath_raw).strip():
                    continue
                
                # Procesar XPath
                xpath = self._process_xpath(xpath_raw)
                if not xpath:
                    continue
                
                # Procesar AMBAS columnas (CAD Dispatch y CAD Incident) por separado
                mappings_found = []
                
                # 1. Verificar CAD Dispatch (columnas 5 y 6)
                if dispatch_table_col in self.mapping_data.columns and dispatch_field_col in self.mapping_data.columns:
                    dispatch_table_val = row.get(dispatch_table_col, '')
                    dispatch_field_val = row.get(dispatch_field_col, '')
                    
                    if (not pd.isna(dispatch_table_val) and str(dispatch_table_val).strip() and 
                        str(dispatch_table_val).strip() not in ['CAD Table'] and
                        not pd.isna(dispatch_field_val) and str(dispatch_field_val).strip() and
                        str(dispatch_field_val).strip() not in ['CAD Column']):
                        
                        table_name = str(dispatch_table_val).strip()
                        field_original = str(dispatch_field_val).strip()
                        field_normalized = self._normalize_field_name(field_original)
                        
                        mappings_found.append({
                            'table_name': table_name,
                            'field_original': field_original,
                            'field_normalized': field_normalized,
                            'source': 'CAD_Dispatch'
                        })
                
                # 2. Verificar CAD Incident (columnas 7 y 8)
                if incident_table_col in self.mapping_data.columns and incident_field_col in self.mapping_data.columns:
                    incident_table_val = row.get(incident_table_col, '')
                    incident_field_val = row.get(incident_field_col, '')
                    
                    if (not pd.isna(incident_table_val) and str(incident_table_val).strip() and 
                        str(incident_table_val).strip() not in ['CAD Table'] and
                        not pd.isna(incident_field_val) and str(incident_field_val).strip() and
                        str(incident_field_val).strip() not in ['CAD Column']):
                        
                        table_name = str(incident_table_val).strip()
                        field_original = str(incident_field_val).strip()
                        field_normalized = self._normalize_field_name(field_original)
                        
                        mappings_found.append({
                            'table_name': table_name,
                            'field_original': field_original,
                            'field_normalized': field_normalized,
                            'source': 'CAD_Incident'
                        })
                
                # 3. Crear mapeos válidos para cada tabla/campo encontrado
                for mapping_info in mappings_found:
                    # Crear query SQL con filtros específicos
                    query = self._build_filtered_query(mapping_info['table_name'], mapping_info['field_normalized'])
                    
                    valid_mappings.append({
                        'xpath': xpath,
                        'query': query,
                        'xpath_raw': xpath_raw,
                        'table_name': mapping_info['table_name'],
                        'column_name': mapping_info['field_normalized'],
                        'column_original': mapping_info['field_original'],
                        'source': mapping_info['source'],
                        'row_index': idx
                    })
            
            if not valid_mappings:
                self.logger.warning("No se encontraron mapeos válidos. Usando mapeo de ejemplo.")
                # Crear un mapeo de ejemplo para probar
                valid_mappings = [{
                    'xpath': '//Incident/IncidentNumber',
                    'query': self._build_filtered_query('nfirs_notification', 'incident_number'),
                    'xpath_raw': '<Incident> <IncidentNumber>',
                    'table_name': 'nfirs_notification',
                    'column_name': 'incident_number',
                    'column_original': 'incident_number',
                    'source': 'Example'
                }]
            
            # Guardar los mapeos válidos tanto como lista como DataFrame
            self.valid_mappings = valid_mappings
            self.mapping_data = pd.DataFrame(valid_mappings)
            
            self.logger.info(f"Archivo de mapeo cargado exitosamente: {self.mapping_file}")
            self.logger.info(f"Número de mapeos válidos cargados: {len(valid_mappings)}")
            
            return True
        except Exception as e:
            self.logger.error(f"Error al cargar el archivo de mapeo: {str(e)}")
            return False

    def _build_filtered_query(self, table_name, column_name, record_identifiers=None):
        """Construir query SQL con filtros específicos para cada tabla usando configuración y identificadores del XML."""
        try:
            # Normalizar nombres de columnas reemplazando espacios con guiones bajos
            normalized_column = column_name.replace(' ', '_') if column_name else column_name
            
            # Obtener filtros de la configuración
            filters = self.config.get('filters', {}) if self.config else {}
            batt_dept_filter = filters.get('batt_dept_id', {})
            datetime_filter = filters.get('datetime', {})
            
            # Obtener valores de filtros
            batt_dept_values = batt_dept_filter.get('values', [4611])  # Valor por defecto para Obion TN
            batt_dept_ids = ','.join(str(val) for val in batt_dept_values)
            
            start_datetime = datetime_filter.get('start_datetime', '2025-09-22 00:00:00')  # Valor por defecto
            
            # Construir la query base con alias para evitar ambigüedades
            table_alias = table_name[0]  # Usar primera letra como alias (d para dispatch, n para nfirs_notification)
            query = f"SELECT {table_alias}.{normalized_column} FROM {table_name} {table_alias} WHERE 1=1"
            
            # PRIORIDAD: Si tenemos identificadores del XML, buscar el registro específico
            specific_record_found = False
            if record_identifiers:
                self.logger.debug(f"Intentando buscar registro específico con identificadores: {record_identifiers}")
                
                # Extraer el valor principal del XML
                xml_value = None
                if record_identifiers.get('xref_id'):
                    xml_value = record_identifiers['xref_id']
                elif record_identifiers.get('dispatch_number'):
                    xml_value = record_identifiers['dispatch_number']
                
                if xml_value:
                    # NUEVA LÓGICA: Buscar en los campos correctos según la tabla
                    if table_name.lower() == 'dispatch':
                        # En tabla dispatch buscar en campo xref_id
                        query += f" AND {table_alias}.xref_id = '{xml_value}'"
                        specific_record_found = True
                        self.logger.info(f"Buscando dispatch por xref_id: {xml_value}")
                    
                    elif table_name.lower() == 'nfirs_notification':
                        # En tabla nfirs_notification buscar en campo dispatch_number
                        query += f" AND {table_alias}.dispatch_number = '{xml_value}'"
                        specific_record_found = True
                        self.logger.info(f"Buscando nfirs_notification por dispatch_number: {xml_value}")
                
                # También buscar por incident_number si está disponible (solo para nfirs_notification)
                if not specific_record_found and table_name.lower() == 'nfirs_notification' and record_identifiers.get('incident_number'):
                    query += f" AND {table_alias}.incident_number = '{record_identifiers['incident_number']}'"
                    specific_record_found = True
                    self.logger.info(f"Buscando nfirs_notification por incident_number: {record_identifiers['incident_number']}")
            
            # Si NO encontramos un registro específico, usar filtros generales (registro más reciente)
            if not specific_record_found:
                self.logger.debug(f"No se encontró identificador específico, usando filtros generales para {table_name}")
                
                # Aplicar filtros específicos según la tabla
                if table_name.lower() == 'dispatch':
                    # Para tabla dispatch: filtrar por batt_dept_id y created_at
                    query += f" AND {table_alias}.batt_dept_id IN ({batt_dept_ids})"
                    query += f" AND {table_alias}.created_at >= '{start_datetime}'"
                    
                elif table_name.lower() == 'nfirs_notification':
                    # Para tabla nfirs_notification: filtrar por batt_dept_id y created_at
                    query += f" AND {table_alias}.batt_dept_id IN ({batt_dept_ids})"
                    query += f" AND {table_alias}.created_at >= '{start_datetime}'"
                    
                elif table_name.lower() == 'nfirs_notification_apparatus':
                    # Para nfirs_notification_apparatus: necesita JOIN con nfirs_notification
                    if record_identifiers and record_identifiers.get('incident_number'):
                        query = f"""SELECT nna.{normalized_column} 
                                   FROM nfirs_notification_apparatus nna
                                   INNER JOIN nfirs_notification nn ON nna.nfirs_notification_id = nn.id
                                   WHERE nn.incident_number = '{record_identifiers['incident_number']}'"""
                        specific_record_found = True
                    else:
                        query = f"""SELECT nna.{normalized_column} 
                                   FROM nfirs_notification_apparatus nna
                                   INNER JOIN nfirs_notification nn ON nna.nfirs_notification_id = nn.id
                                   WHERE nn.batt_dept_id IN ({batt_dept_ids})
                                   AND nn.created_at >= '{start_datetime}'
                                   ORDER BY nn.id DESC LIMIT 1"""
                               
                elif table_name.lower() == 'nfirs_notification_personnel':
                    # Para nfirs_notification_personnel: necesita JOIN con nfirs_notification
                    if record_identifiers and record_identifiers.get('incident_number'):
                        query = f"""SELECT nnp.{normalized_column} 
                                   FROM nfirs_notification_personnel nnp
                                   INNER JOIN nfirs_notification nn ON nnp.nfirs_notification_id = nn.id
                                   WHERE nn.incident_number = '{record_identifiers['incident_number']}'"""
                        specific_record_found = True
                    else:
                        query = f"""SELECT nnp.{normalized_column} 
                                   FROM nfirs_notification_personnel nnp
                                   INNER JOIN nfirs_notification nn ON nnp.nfirs_notification_id = nn.id
                                   WHERE nn.batt_dept_id IN ({batt_dept_ids})
                                   AND nn.created_at >= '{start_datetime}'
                                   ORDER BY nn.id DESC LIMIT 1"""
                               
                elif table_name.lower() == 'batt_dept':
                    # Para batt_dept, solo filtrar por ID (no tiene created_at)
                    query += f" AND {table_alias}.id IN ({batt_dept_ids})"
                    
                else:
                    # Para otras tablas, usar filtros generales
                    self.logger.warning(f"Tabla desconocida: {table_name}, usando filtros generales")
                    query += f" AND {table_alias}.batt_dept_id IN ({batt_dept_ids})"
                    query += f" AND {table_alias}.created_at >= '{start_datetime}'"
            
            # Limitar resultados (solo si no encontramos registro específico y no es JOIN con ORDER BY)
            if not specific_record_found and not ('JOIN' in query and 'ORDER BY' in query):
                query += f" ORDER BY {table_alias}.id DESC LIMIT 1"
            
            # Log del tipo de query generada
            if specific_record_found:
                self.logger.info(f"Query específica generada para {table_name}.{column_name}: {query}")
            else:
                self.logger.info(f"Query general (más reciente) generada para {table_name}.{column_name}: {query}")
            
            return query
            
        except Exception as e:
            self.logger.error(f"Error construyendo query filtrada: {str(e)}")
            # Query simple sin filtros como fallback
            escaped_column = f'"{column_name}"' if ' ' in column_name else column_name
            return f"SELECT {escaped_column} FROM {table_name} LIMIT 1"

    def connect_to_db(self):
        """Conectar a la base de datos PostgreSQL usando la configuración cargada."""
        try:
            self.logger.info("Iniciando conexión a BD...")
            
            if not self.config:
                self.logger.error("No se ha cargado la configuración")
                return False
                
            if 'database' not in self.config:
                self.logger.error("No se encuentra sección 'database' en configuración")
                return False
                
            db_config = self.config['database']
            self.logger.info(f"Intentando conectar a BD: {db_config.get('host', 'N/A')}:{db_config.get('port', 'N/A')} DB: {db_config.get('dbname', 'N/A')}")
            
            self.conn = psycopg2.connect(
                host=db_config['host'],
                port=db_config['port'], 
                database=db_config['dbname'],
                user=db_config['user'],
                password=db_config['password']
            )
            # Configurar autocommit para evitar errores de transacción
            self.conn.autocommit = True
            self.cursor = self.conn.cursor()
            self.logger.info("✅ Conexión a la base de datos establecida exitosamente")
            return True
        except Exception as e:
            self.logger.error(f"❌ Error al conectar a la base de datos: {str(e)}")
            return False

    def close_db_connection(self):
        """Cerrar la conexión a la base de datos."""
        if self.cursor:
            self.cursor.close()
        if self.conn:
            self.conn.close()
            self.logger.info("Conexión a la base de datos cerrada")

    def compare_xml_with_db(self, xml_folder_path):
        """Comparar archivos XML con la base de datos."""
        if self.mapping_data is None or self.conn is None:
            self.logger.error("Debe cargar el archivo de mapeo y conectarse a la BD primero")
            return None

        results = []
        
        # Crear directorio de reportes si no existe
        report_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'reportes')
        os.makedirs(report_dir, exist_ok=True)

        try:
            # Obtener lista de archivos XML (tanto .xml como .XML)
            xml_files = [f for f in os.listdir(xml_folder_path) if f.lower().endswith('.xml')]
            
            if not xml_files:
                self.logger.warning(f"No se encontraron archivos XML en: {xml_folder_path}")
                return None
            
            self.logger.info(f"Se encontraron {len(xml_files)} archivos XML para procesar")

            # Procesar cada archivo XML en la carpeta
            for xml_file in xml_files:
                xml_path = os.path.join(xml_folder_path, xml_file)
                self.logger.info(f"Procesando archivo: {xml_file}")

                try:
                    # Parsear el archivo XML con manejo robusto de errores
                    try:
                        tree = etree.parse(xml_path)
                        root = tree.getroot()
                    except etree.XMLSyntaxError as xml_error:
                        # Error de formato XML - intentar parsearlo de manera más tolerante
                        self.logger.warning(f"Error de formato XML en {xml_file}: {xml_error}")
                        self.logger.info(f"Intentando parseo más tolerante para {xml_file}")
                        
                        # Leer el archivo y intentar parsear con recuperación de errores
                        parser = etree.XMLParser(recover=True)
                        tree = etree.parse(xml_path, parser)
                        root = tree.getroot()
                        
                        if parser.error_log:
                            self.logger.warning(f"Errores recuperados en {xml_file}: {len(parser.error_log.filter_from_level(etree.ErrorLevels.WARNING))}")
                        
                        self.logger.info(f"XML parseado exitosamente con recuperación de errores: {xml_file}")
                    
                    if root is None:
                        raise Exception(f"No se pudo parsear el XML: {xml_file}")

                    # EXTRAER IDENTIFICADORES ÚNICOS DEL XML PARA BUSCAR REGISTRO ESPECÍFICO
                    record_identifiers = self._extract_record_identifiers(root, xml_file)
                    
                    self.logger.info(f"📄 PROCESANDO {xml_file}")
                    self.logger.info(f"🔍 Identificadores extraídos: {record_identifiers}")
                    
                    # Inicializar variables
                    matching_record_found = False
                    found_table = None
                    
                    if record_identifiers and any(record_identifiers.values()):
                        self.logger.info(f"Identificadores extraídos de {xml_file}: {record_identifiers}")
                        
                        # VERIFICAR SI EXISTE UN REGISTRO EN LA BD CON ESTOS IDENTIFICADORES
                        matching_record_found, found_table = self._verify_record_exists_in_db(record_identifiers)
                        
                        if not matching_record_found:
                            # No existe registro coincidente - crear entrada de "no encontrado"
                            # Extraer el valor principal del XML para el mensaje
                            xml_value = None
                            if record_identifiers.get('xref_id'):
                                xml_value = record_identifiers['xref_id']
                            elif record_identifiers.get('dispatch_number'):
                                xml_value = record_identifiers['dispatch_number']
                            
                            if xml_value:
                                error_message = f"No existe registro en la BD con xref_id/dispatch_number: '{xml_value}'"
                            else:
                                identifier_info = []
                                for key, value in record_identifiers.items():
                                    if value:
                                        identifier_info.append(f"{key}: {value}")
                                identifier_text = ", ".join(identifier_info) if identifier_info else "Sin identificadores válidos"
                                error_message = f"No existe registro en la BD con {identifier_text}"
                            
                            self.logger.warning(f"No se encontró registro en BD para {xml_file} - {error_message}")
                            
                            results.append({
                                'archivo': xml_file,
                                'tabla': 'N/A',
                                'campo': 'verificacion_registro',
                                'xpath': 'N/A',
                                'valor_xml': xml_value or "Sin identificadores válidos",
                                'valor_bd': 'N/A',
                                'coincide': False,
                                'observaciones': error_message
                            })
                            
                            # Saltar este XML y continuar con el siguiente
                            continue
                    else:
                        self.logger.warning(f"No se pudieron extraer identificadores de {xml_file}, usando registro más reciente")
                        found_table = None

                    # Si se encontró el registro, log de información sobre qué tabla se usará
                    if matching_record_found and found_table:
                        self.logger.info(f"Procesando comparación para {xml_file} usando registros de tabla: {found_table}")

                    # Procesar cada mapeo válido generado
                    for row in self.valid_mappings:
                        xpath = row['xpath']
                        
                        if not xpath:
                            self.logger.warning(f"XPath vacío en mapeo: {row}")
                            continue
                        
                        # Obtener valor del XML con soporte para XPath concatenados y condicionales
                        xml_value = None
                        try:
                            # Usar el método especializado que maneja lógicas especiales por campo
                            xml_value = self._extract_xml_value_with_special_logic(root, xpath, row['column_name'], row['table_name'])
                            
                            if xml_value is None:
                                self.logger.debug(f"No se encontró valor para XPath '{xpath}' en {xml_file}")
                                
                        except Exception as e:
                            self.logger.error(f"Error al procesar XPath '{xpath}' en {xml_file}: {str(e)}")
                            xml_value = "ERROR_XPATH"

                        # Obtener valor de la BD usando identificadores específicos del XML
                        db_value = None
                        try:
                            # VERIFICAR CURSOR ANTES DE USAR
                            if self.cursor is None:
                                self.logger.error("❌ CURSOR ES NONE - La conexión a BD no se estableció correctamente")
                                db_value = "ERROR_CONEXION"
                                continue
                                
                            # Construir query específica usando identificadores del XML
                            query = self._build_filtered_query(row['table_name'], row['column_name'], record_identifiers)
                            
                            if not query:
                                self.logger.warning(f"No se pudo construir query para {row['table_name']}.{row['column_name']}")
                                db_value = "ERROR_QUERY"
                            else:
                                # Ejecutar query
                                try:
                                    self.cursor.execute(query)
                                    result = self.cursor.fetchone()
                                    if result and result[0] is not None:
                                        db_value = str(result[0]).strip()
                                    else:
                                        db_value = None
                                except Exception as db_error:
                                    error_msg = str(db_error)
                                    self.logger.error(f"Error SQL: {error_msg}")
                                    if "does not exist" in error_msg or "column" in error_msg.lower():
                                        db_value = "CAMPO_NO_EXISTE"
                                    else:
                                        db_value = "ERROR_QUERY"
                        
                        except Exception as e:
                            self.logger.error(f"Error general: {str(e)}")
                            db_value = "ERROR_QUERY"

                        # Comparar valores con tratamiento especial para coordenadas
                        xml_str = str(xml_value) if xml_value is not None else ""
                        db_str = str(db_value) if db_value is not None else ""
                        
                        # Variables para almacenar valores para comparación (pueden ser diferentes a los originales)
                        xml_comparison_value = xml_str
                        db_comparison_value = db_str
                        
                        # Tratamiento especial para latitude y longitude
                        if row['column_name'].lower() in ['latitude', 'longitude']:
                            self.logger.info(f"🗺️ Procesando coordenada {row['column_name']}: XML='{xml_str}', DB='{db_str}'")
                            # Normalizar SOLO si ambos valores son válidos y numéricos
                            if xml_str and xml_str not in ["ERROR_XPATH", "ERROR_CONEXION", "ERROR_QUERY", "CAMPO_NO_EXISTE", ""]:
                                xml_comparison_value = self._normalize_coordinate(xml_str)
                                self.logger.info(f"🗺️ XML normalizado: '{xml_str}' -> '{xml_comparison_value}'")
                            
                            if db_str and db_str not in ["ERROR_XPATH", "ERROR_CONEXION", "ERROR_QUERY", "CAMPO_NO_EXISTE", "", "None"]:
                                db_comparison_value = self._normalize_coordinate(db_str)
                                self.logger.info(f"🗺️ BD normalizado: '{db_str}' -> '{db_comparison_value}'")
                            
                            match = (xml_comparison_value == db_comparison_value) and xml_comparison_value != "" and db_comparison_value != ""
                        else:
                            # Comparación normal para otros campos
                            match = (xml_comparison_value == db_comparison_value) and xml_comparison_value != "" and db_comparison_value != ""
                        
                        # Log detallado para debugging
                        self.logger.debug(f"Comparación - XPath: {xpath}, XML: '{xml_str}', DB: '{db_str}', Match: {match}")
                        
                        results.append({
                            'archivo': xml_file,
                            'tabla': row['table_name'],
                            'campo': row.get('column_original', row['column_name']), 
                            'xpath': xpath,
                            'valor_xml': xml_value,
                            'valor_bd': db_value,
                            'coincide': match,
                            'observaciones': self._get_comparison_notes(xml_value, db_value, row['column_name'])
                        })

                except Exception as e:
                    self.logger.error(f"Error al procesar el archivo {xml_file}: {str(e)}")
                    # Agregar entrada de error para este archivo
                    results.append({
                        'archivo': xml_file,
                        'tabla': 'ERROR',
                        'campo': 'ERROR',
                        'xpath': 'ERROR',
                        'valor_xml': f"Error al procesar archivo: {str(e)}",
                        'valor_bd': None,
                        'coincide': False,
                        'observaciones': 'Error de procesamiento'
                    })

            # Crear reporte Excel
            if results:
                report_path = os.path.join(report_dir, f'reporte_comparacion_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
                df = pd.DataFrame(results)
                
                # Crear resumen
                total_comparaciones = len(results)
                coincidencias = len([r for r in results if r['coincide']])
                errores = len([r for r in results if 'ERROR' in str(r['valor_xml']) or 'ERROR' in str(r['valor_bd'])])
                
                with pd.ExcelWriter(report_path) as writer:
                    # Hoja principal con los resultados
                    df.to_excel(writer, sheet_name='Resultados', index=False)
                    
                    # Hoja de resumen
                    resumen = pd.DataFrame({
                        'Métrica': ['Total de comparaciones', 'Coincidencias', 'Diferencias', 'Errores', 'Porcentaje de coincidencias'],
                        'Valor': [total_comparaciones, coincidencias, total_comparaciones - coincidencias - errores, errores, f"{(coincidencias/total_comparaciones*100):.2f}%" if total_comparaciones > 0 else "0%"]
                    })
                    resumen.to_excel(writer, sheet_name='Resumen', index=False)
                
                # Aplicar formato de colores
                self._apply_color_formatting(report_path, df)
                
                self.logger.info(f"Reporte generado: {report_path}")
                self.logger.info(f"Resumen: {coincidencias}/{total_comparaciones} coincidencias ({(coincidencias/total_comparaciones*100):.2f}%)")
                return report_path
            else:
                self.logger.warning("No se generó ningún resultado para el reporte")
                return None

        except Exception as e:
            self.logger.error(f"Error durante la comparación: {str(e)}")
            return None

    def _normalize_coordinate(self, coord_str):
        """
        Normalizar coordenadas (latitude/longitude) a 7 decimales.
        - Si tiene menos de 7 decimales, completar con ceros
        - Si tiene más de 7 decimales, redondear a 7 decimales
        """
        if not coord_str or coord_str.strip() == "" or coord_str in ["ERROR_XPATH", "ERROR_CONEXION", "ERROR_QUERY", "CAMPO_NO_EXISTE"]:
            return coord_str
        
        try:
            # Convertir a float para manejar la precisión
            coord_float = float(coord_str.strip())
            # Formatear a exactamente 7 decimales
            normalized = f"{coord_float:.7f}"
            return normalized
        except (ValueError, TypeError):
            # Si no se puede convertir a float, devolver el valor original
            return coord_str

    def _get_comparison_notes(self, xml_value, db_value, column_name=None):
        """Generar notas sobre la comparación."""
        # Verificar si hay errores
        if "ERROR" in str(xml_value):
            return "Error al obtener valor XML"
        elif "ERROR" in str(db_value) or "CAMPO_NO_EXISTE" in str(db_value):
            return "Error al obtener valor BD"
        
        # Verificar valores nulos
        xml_is_null = xml_value is None or str(xml_value).strip() == ""
        db_is_null = db_value is None or str(db_value).strip() == ""
        
        if xml_is_null and db_is_null:
            return "Ambos valores son nulos"
        elif xml_is_null:
            return "Valor XML es nulo"
        elif db_is_null:
            return "Valor BD es nulo"
        
        # Comparar valores no nulos
        xml_str = str(xml_value).strip()
        db_str = str(db_value).strip()
        
        # Tratamiento especial para coordenadas
        if column_name and column_name.lower() in ['latitude', 'longitude']:
            # Para coordenadas, comparar usando valores normalizados
            xml_normalized = self._normalize_coordinate(xml_str) if xml_str not in ["ERROR_XPATH", "ERROR_CONEXION", "ERROR_QUERY", "CAMPO_NO_EXISTE"] else xml_str
            db_normalized = self._normalize_coordinate(db_str) if db_str not in ["ERROR_XPATH", "ERROR_CONEXION", "ERROR_QUERY", "CAMPO_NO_EXISTE"] else db_str
            
            if xml_normalized == db_normalized:
                return "Valores coinciden (coordenadas normalizadas)"
            else:
                return f"Valores diferentes: XML='{xml_str}' vs BD='{db_str}'"
        else:
            # Comparación normal para otros campos
            if xml_str == db_str:
                return "Valores coinciden"
            else:
                return f"Valores diferentes: XML='{xml_str}' vs BD='{db_str}'"

    def _apply_color_formatting(self, report_path, df):
        """Aplicar formato de colores al reporte Excel."""
        try:
            # Cargar el archivo Excel generado
            workbook = openpyxl.load_workbook(report_path)
            worksheet = workbook['Resultados']
            
            # Definir colores
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Rojo claro
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Amarillo claro
            blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")   # Azul claro
            
            # Buscar la columna de "observaciones"
            observaciones_col = None
            for col_idx, cell in enumerate(worksheet[1], 1):  # Primera fila (encabezados)
                if cell.value and 'observacion' in str(cell.value).lower():
                    observaciones_col = col_idx
                    break
            
            if not observaciones_col:
                self.logger.warning("No se encontró la columna 'observaciones' para aplicar colores")
                return
            
            # Aplicar colores según el contenido de observaciones
            for row_idx in range(2, len(df) + 2):  # Empezar desde la fila 2 (datos, no encabezados)
                cell = worksheet.cell(row=row_idx, column=observaciones_col)
                if cell.value:
                    observacion = str(cell.value).lower()
                    
                    if "coinciden" in observacion:
                        # Verde para coincidencias - solo la celda de observaciones
                        cell.fill = green_fill
                    elif "error" in observacion:
                        # Amarillo para errores - solo la celda de observaciones
                        cell.fill = yellow_fill
                    elif "diferentes" in observacion:
                        # Rojo para diferencias - solo la celda de observaciones
                        cell.fill = red_fill
                    elif "nulo" in observacion or "nulos" in observacion:
                        # Azul para valores nulos - solo la celda de observaciones
                        cell.fill = blue_fill
            
            # Guardar cambios
            workbook.save(report_path)
            self.logger.info("Formato de colores aplicado al reporte Excel")
            
        except Exception as e:
            self.logger.error(f"Error aplicando formato de colores: {str(e)}")
            # El reporte se genera sin colores si hay error


if __name__ == "__main__":
    # Configuración para ejecución directa
    config_path = r"C:\FDSU\Automatizacion\Yatary_Pruebas\XML_BD_Comparator\config\config.json"
    mapping_path = r"C:\FDSU\Automatizacion\Yatary_Pruebas\XML_BD_Comparator\mappings\Humphreys_Co_TN_GeoConex_XML_Mappings_20250508.xlsx"
    xml_folder_path = r"C:\FDSU\Clients\Obion TN\example"
    
    # Crear instancia del comparador con archivos de configuración
    comparador = XPathMapper(config_file=config_path, mapping_file=mapping_path)
    
    # Verificar que la configuración se cargó
    if not comparador.config:
        print("❌ Error cargando configuración")
        exit(1)
    print("✅ Configuración cargada exitosamente")
    
    # Verificar que el mapeo se cargó
    if comparador.mapping_data is None:
        print("❌ Error cargando mapeo")
        exit(1)
    print("✅ Mapeo cargado exitosamente")
    
    # Conectar a base de datos
    if comparador.connect_to_db():
        print("✅ Conexión a base de datos establecida")
    else:
        print("❌ Error conectando a base de datos")
        exit(1)
    
    # Ejecutar comparación
    print(f"🔍 Iniciando comparación de XMLs en: {xml_folder_path}")
    results = comparador.compare_xml_with_db(xml_folder_path)
    
    if results:
        print(f"✅ Comparación completada. {len(results)} comparaciones realizadas")
    else:
        print("❌ No se obtuvieron resultados de la comparación")
    
    # Cerrar conexión
    comparador.close_db_connection()
    print("🔚 Proceso completado")
